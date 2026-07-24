# tests/test_generate_excel_report_history.py
"""
Test per generate_excel_report.py::save_report_with_history() — fix 2026-07-24
(richiesta esplicita dell'utente: "i file excel non si devono sovrascrivere",
un file/snapshot permanente per ogni esperimento lanciato, mantenendo comunque
un file/foglio "corrente" che si aggiorna a ogni run).

Nessuna dipendenza da torch (pure Python + openpyxl) — eseguiti realmente in
questo sandbox, non solo py_compile-verificati. Aggiunti dopo che una review
indipendente ha segnalato che la prima versione di questa funzione non aveva
copertura di test, nonostante fosse pienamente testabile senza torch (a
differenza della maggior parte del codice NVFLARE di questa sessione).
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

import generate_excel_report as ger  # noqa: E402


def _wb_with_marker(value: str) -> Workbook:
    wb = Workbook()
    wb.active["A1"] = value
    return wb


class TestSaveReportWithHistory:
    def test_writes_both_current_and_history_snapshot(self, tmp_path):
        output_path = tmp_path / "sweep1.xlsx"
        snapshot_path = ger.save_report_with_history(_wb_with_marker("v1"), output_path)

        assert output_path.exists(), "il file 'corrente' (output_path) deve sempre esistere"
        assert snapshot_path.exists(), "lo snapshot in history/ deve esistere"
        assert snapshot_path.parent == tmp_path / "history"
        assert snapshot_path.name.startswith("sweep1_")
        assert snapshot_path.suffix == ".xlsx"

    def test_current_file_is_overwritten_but_history_accumulates(self, tmp_path):
        """Il file 'corrente' riflette sempre l'ultima chiamata; history/ accumula
        uno snapshot per ogni chiamata, senza mai perdere quelli precedenti —
        il comportamento esplicitamente richiesto dall'utente."""
        output_path = tmp_path / "sweep1.xlsx"

        snap1 = ger.save_report_with_history(_wb_with_marker("run1"), output_path)
        snap2 = ger.save_report_with_history(_wb_with_marker("run2"), output_path)

        assert snap1 != snap2, "due chiamate devono produrre due snapshot distinti"
        assert snap1.exists() and snap2.exists(), (
            "il secondo salvataggio non deve cancellare/sovrascrivere il primo snapshot"
        )

        # Il file "corrente" riflette l'ultima chiamata (comportamento voluto:
        # è la vista aggregata sempre aggiornata, non un record permanente).
        current_wb = load_workbook(output_path)
        assert current_wb.active["A1"].value == "run2"

        # Ma ENTRAMBI gli snapshot in history/ restano quelli originali.
        assert load_workbook(snap1).active["A1"].value == "run1"
        assert load_workbook(snap2).active["A1"].value == "run2"

    def test_many_rapid_calls_never_collide(self, tmp_path):
        """Regressione per il bug di collisione trovato da review indipendente
        2026-07-24: la prima versione usava timestamp al secondo + contatore
        incrementale scelto con un controllo 'esiste già?' — un pattern
        check-then-write con una race condition reale fra processi concorrenti.
        Qui verifichiamo almeno che molte chiamate ravvicinate in un singolo
        processo (stesso secondo quasi certamente) producano comunque path
        tutti distinti, grazie al suffisso random."""
        output_path = tmp_path / "sweep1.xlsx"
        snapshots = [
            ger.save_report_with_history(_wb_with_marker(f"run{i}"), output_path)
            for i in range(20)
        ]
        assert len(set(snapshots)) == 20, "tutti gli snapshot devono avere path distinti"
        for i, snap in enumerate(snapshots):
            assert load_workbook(snap).active["A1"].value == f"run{i}"

    def test_creates_output_and_history_dirs_if_missing(self, tmp_path):
        output_path = tmp_path / "nested" / "sweepX.xlsx"
        snapshot_path = ger.save_report_with_history(_wb_with_marker("v1"), output_path)
        assert output_path.exists()
        assert snapshot_path.exists()
        assert snapshot_path.parent == tmp_path / "nested" / "history"
