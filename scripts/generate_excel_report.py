#!/usr/bin/env python3
"""
scripts/generate_excel_report.py
ChargeShield-FL — Genera report Excel da tutti i risultati in experiments/

Legge tutti i file experiment_*.json e produce:
  Sheet 1  "Raw Data"          — ogni esperimento su una riga (Yeom + Shadow + LiRA)
  Sheet 2  "Heat Map"          — matrice AUC-ROC: righe=round, colonne=epsilon
  Sheet 3  "Per Rounds"        — AUC-ROC medio, min, max per numero di round
  Sheet 4  "Per Epsilon"       — AUC-ROC medio, min, max per valore di epsilon
  Sheet 5  "Comparison"        — confronto diretto fra configurazioni
  Sheet 6  "AUC Progression"   — AUC Yeom per round (convergenza FL)
  Sheet 7  "Attack Comparison" — Yeom vs Shadow vs LiRA sintetico per esperimento
  Sheet 8  "Yeom Per Round"    — AUC Yeom round-by-round per ogni esperimento
  Sheet 9  "Shadow Per Round"  — AUC Shadow round-by-round per ogni esperimento
  Sheet 10 "LiRA Per Round"    — AUC LiRA round-by-round per ogni esperimento (★ PRIMARY)

Usage:
  python scripts/generate_excel_report.py
  python scripts/generate_excel_report.py --output experiments/my_report.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl non trovato. Installa con: pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).parent.parent
EXPERIMENTS_DIR = PROJECT_ROOT / "experiments"

# ── Palette colori ChargeShield ────────────────────────────────────────────────
COLOR_HEADER_BG  = "1F4E79"   # blu scuro
COLOR_HEADER_FG  = "FFFFFF"   # bianco
COLOR_SUBHDR_BG  = "2E75B6"   # blu medio
COLOR_ROW_ALT    = "D6E4F0"   # azzurro chiaro (righe alternate)
COLOR_ROW_PLAIN  = "FFFFFF"
COLOR_ACCENT     = "FF6B35"   # arancione accent (AUC-ROC sopra 0.5)
COLOR_GOOD       = "70AD47"   # verde (LOW risk)
COLOR_WARN       = "FFC000"   # giallo (MEDIUM risk)
COLOR_BAD        = "FF0000"   # rosso (HIGH risk)
COLOR_RANDOM     = "BDD7EE"   # celeste (≈ random, AUC ≈ 0.5)
COLOR_ANOMALY    = "7030A0"   # viola (AUC anomalo — probabile attacco rotto, non "DP sicura")

# Soglia sotto la quale un AUC-ROC va trattato come anomalia da investigare,
# non come "privacy sicura". Fix 2026-07-21 (review): prima di questo fix,
# QUALSIASI AUC <= 0.52 veniva colorato di verde/"LOW risk" indistintamente —
# un lira_auc_roc di 0.14-0.32 (visto in experiments/exp3 prima dei fix a/b/c
# a run_lira()) veniva quindi presentato come "sicuro", mentre in realtà
# segnalava un attacco sistematicamente invertito (bug), non DP efficace.
# Un vero AUC ≈ 0.5 (DP che funziona) resta nel range [ANOMALY_LOW_AUC, 0.52];
# sotto ANOMALY_LOW_AUC è quasi certamente un problema di implementazione.
ANOMALY_LOW_AUC  = 0.40

FONT_NAME = "Arial"

# ── Helpers stile ──────────────────────────────────────────────────────────────

def _font(bold=False, color="000000", size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)


def _auc_risk_color(val: float | None) -> str:
    """
    Classifica un AUC-ROC in un colore di rischio, DISTINGUENDO un AUC vicino
    a 0.5 (DP efficace / attacco al livello del caso) da un AUC anomalmente
    BASSO (probabile bug nell'attacco, non privacy sicura — vedi ANOMALY_LOW_AUC).

    Bande:
        val is None        → nero (nessun dato)
        val > 0.60          → BAD (rosso)      — MIA efficace, rischio alto
        val > 0.52          → WARN (giallo)    — segnale debole ma presente
        val >= ANOMALY_LOW_AUC → GOOD (verde)  — ≈ random, DP verosimilmente efficace
        val <  ANOMALY_LOW_AUC → ANOMALY (viola) — sospetto bug, non "sicuro"
    """
    if val is None:
        return "000000"
    if val > 0.60:
        return COLOR_BAD
    if val > 0.52:
        return COLOR_WARN
    if val < ANOMALY_LOW_AUC:
        return COLOR_ANOMALY
    return COLOR_GOOD


def _risk_label_color(risk: str | None) -> str:
    """Colora la stringa privacy_risk ("HIGH"/"MEDIUM"/"LOW"/"ANOMALY") coerentemente
    con _auc_risk_color(). "ANOMALY" ha priorità visiva (viola) — non va MAI
    confuso con "LOW" (verde), anche se compaiono entrambi come "basso rischio"."""
    if risk == "ANOMALY":
        return COLOR_ANOMALY
    if risk == "HIGH":
        return COLOR_BAD
    if risk == "MEDIUM":
        return COLOR_WARN
    return COLOR_GOOD

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _right():
    return Alignment(horizontal="right", vertical="center")

def _header_cell(cell, text, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, size=10):
    cell.value = text
    cell.font = _font(bold=True, color=fg, size=size)
    cell.fill = _fill(bg)
    cell.alignment = _center()
    cell.border = _border()

def _data_cell(cell, value, fmt=None, alt_row=False, bold=False):
    cell.value = value
    cell.font = _font(bold=bold)
    cell.fill = _fill(COLOR_ROW_ALT if alt_row else COLOR_ROW_PLAIN)
    cell.border = _border()
    cell.alignment = _center()
    if fmt:
        cell.number_format = fmt

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ── Carica dati ────────────────────────────────────────────────────────────────

def load_experiments(experiments_dir: Path | None = None) -> list[dict]:
    """Legge tutti i file experiment_*.json in experiments/ e li ordina."""
    src = experiments_dir if experiments_dir is not None else EXPERIMENTS_DIR
    records = []
    for path in sorted(src.glob("experiment_*.json")):
        try:
            with open(path) as f:
                data = json.load(f)
            cfg   = data.get("config", {})
            summ  = data.get("summary", {})
            per_round_raw = data.get("per_round", {})
            records.append({
                "timestamp":    data.get("timestamp", path.stem.replace("experiment_", "")),
                "file":         path.name,
                "rounds":       int(cfg.get("fl_rounds", 0)),
                "epsilon":      float(cfg.get("epsilon", 0)),
                # Fix 2026-07-22 (review indipendente fresh-pass): campo
                # aggiunto a save_results() lo stesso giorno (epsilon × fl_rounds,
                # None sotto no_dp) ma mai letto qui — spariva silenziosamente
                # dal report Excel. È il numero di composizione naive/worst-case
                # che il paper userà per il claim di privacy (vedi
                # gradient_manager.py::_compute_sigma() e docs/CaseStudies.md
                # §2.4.3) — non va perso.
                "epsilon_cumulative_naive": (
                    float(cfg["epsilon_cumulative_naive"])
                    if cfg.get("epsilon_cumulative_naive") is not None else None
                ),
                "delta":        float(cfg.get("delta", 1e-5)),
                "proximal_mu":  float(cfg.get("proximal_mu", 0)),
                "no_dp":        bool(cfg.get("no_dp", False)),
                # dp_mode (2026-07-22): "dp-fedavg" (default)/"central"/"local" —
                # irrilevante quando no_dp=True. Vedi docs/CaseStudies.md §2.4.3.
                "dp_mode":      cfg.get("dp_mode", "dp-fedavg"),
                # Yeom 2018 — loss-based MIA sul modello globale (baseline debole)
                "auc_roc":      float(summ["mean_auc_roc"]) if summ.get("mean_auc_roc") is not None else None,
                "auc_max":      float(summ["max_auc_roc"])  if summ.get("max_auc_roc")  is not None else None,
                "auc_min":      float(summ["min_auc_roc"])  if summ.get("min_auc_roc")  is not None else None,
                # Shadow MIA calibrated — Carlini 2022 style, modello globale
                "shadow_auc":   float(summ["mean_shadow_auc_roc"]) if summ.get("mean_shadow_auc_roc") is not None else None,
                "shadow_max":   float(summ["max_shadow_auc_roc"])  if summ.get("max_shadow_auc_roc")  is not None else None,
                # LiRA — Carlini 2022, server-side su raw_updates PRE-aggregazione
                "lira_auc":     float(summ["mean_lira_auc_roc"]) if summ.get("mean_lira_auc_roc") is not None else None,
                "lira_max":     float(summ["max_lira_auc_roc"])  if summ.get("max_lira_auc_roc")  is not None else None,
                "lira_min":     float(summ["min_lira_auc_roc"])  if summ.get("min_lira_auc_roc")  is not None else None,
                # Seed: necessario per multi-seed aggregation (mean±std) — fix M1
                "seed":         int(cfg.get("seed", 42)),
                # Metrica primaria: LiRA > Shadow > Yeom (attacco più forte)
                "primary_attack": summ.get("primary_attack", "Yeom"),
                "privacy_risk": summ.get("privacy_risk", ""),
                # IDS: conta gli alert totali nei round
                "total_alerts": sum(
                    len(v.get("ids", {}).get("alerts", []))
                    for v in per_round_raw.values()
                ),
                "byzantine_rounds": sum(
                    1 for v in per_round_raw.values()
                    if v.get("ids", {}).get("byzantine_detected")
                ),
                # AUC-ROC per round (per progression chart) — Yeom
                "per_round_auc": {
                    int(k): v["mia"]["auc_roc"]
                    for k, v in per_round_raw.items()
                    if v.get("mia", {}).get("auc_roc") is not None
                },
                # Shadow AUC per round
                "per_round_shadow_auc": {
                    int(k): v["mia"]["shadow_auc_roc"]
                    for k, v in per_round_raw.items()
                    if v.get("mia", {}).get("shadow_auc_roc") is not None
                },
                # LiRA AUC per round
                "per_round_lira_auc": {
                    int(k): v["mia"]["lira_auc_roc"]
                    for k, v in per_round_raw.items()
                    if v.get("mia", {}).get("lira_auc_roc") is not None
                },
                # FL training loss per round (andamento convergenza)
                "per_round_loss": {
                    int(k): v["fl"]["mean_loss"]
                    for k, v in per_round_raw.items()
                    if v.get("fl", {}).get("mean_loss") is not None
                },
            })
        except Exception as e:
            print(f"WARN: impossibile leggere {path.name}: {e}")
    return records


# ── Sheet 1: Raw Data ──────────────────────────────────────────────────────────

def build_raw_data(ws, records: list[dict]) -> None:
    ws.title = "Raw Data"

    # Titolo
    # Fix 2026-07-22 (review indipendente fresh-pass): merge/header/indici di
    # colonna estesi da 15 (A:O) a 16 (A:P) per inserire "ε cumulativo (naive)"
    # subito dopo "Epsilon (ε)" — vedi commento su epsilon_cumulative_naive in
    # load_experiments(). Tutti gli indici di colonna da qui in poi sono
    # shiftati di 1 rispetto alla versione precedente di questa funzione.
    ws.merge_cells("A1:P1")
    title = ws["A1"]
    title.value = "ChargeShield-FL — Experiment Results (Full Sweep)"
    title.font = _font(bold=True, color=COLOR_HEADER_FG, size=12)
    title.fill = _fill(COLOR_HEADER_BG)
    title.alignment = _center()

    headers = [
        "Timestamp", "FL Rounds", "Epsilon (ε)", "ε cumulativo (naive, T×ε)",
        "Delta (δ)", "Proximal μ", "No-DP",
        # Yeom
        "Yeom AUC (mean)", "Yeom AUC (max)", "Yeom AUC (min)",
        # Shadow
        "Shadow AUC (mean)", "Shadow AUC (max)",
        # LiRA (primary)
        "LiRA AUC (mean)", "LiRA AUC (max)",
        # Sintesi
        "Privacy Risk", "IDS Alerts",
    ]
    for col, h in enumerate(headers, 1):
        _header_cell(ws.cell(2, col), h, bg=COLOR_SUBHDR_BG)
    # Colora i gruppi di header per distinguere gli attacchi
    for col in [8, 9, 10]:   # Yeom
        _header_cell(ws.cell(2, col), headers[col - 1], bg="4472C4")
    for col in [11, 12]:     # Shadow
        _header_cell(ws.cell(2, col), headers[col - 1], bg="ED7D31")
    for col in [13, 14]:     # LiRA (primary)
        _header_cell(ws.cell(2, col), headers[col - 1], bg="70AD47")

    def _auc_cell(cell, val, alt_row):
        _data_cell(cell, val, fmt="0.0000", alt_row=alt_row)
        if val is not None:
            cell.font = _font(bold=True, color=_auc_risk_color(val))

    for row_idx, rec in enumerate(records, 3):
        alt = (row_idx % 2 == 0)
        _data_cell(ws.cell(row_idx, 1),  rec["timestamp"],           alt_row=alt)
        _data_cell(ws.cell(row_idx, 2),  rec["rounds"],              alt_row=alt)
        _data_cell(ws.cell(row_idx, 3),  rec["epsilon"],             fmt="0.0#", alt_row=alt)
        _data_cell(ws.cell(row_idx, 4),  rec.get("epsilon_cumulative_naive"), fmt="0.0#", alt_row=alt)
        _data_cell(ws.cell(row_idx, 5),  rec["delta"],               fmt="0.00E+00", alt_row=alt)
        _data_cell(ws.cell(row_idx, 6),  rec["proximal_mu"],         fmt="0.00", alt_row=alt)
        # dp_mode (2026-07-22): mostrato tra parentesi solo quando DP è attivo e
        # non è la modalità storica di default ("dp-fedavg") — evita di dover
        # inserire un'ulteriore colonna e shiftare di nuovo tutti gli indici.
        _dp_mode = rec.get("dp_mode", "dp-fedavg")
        _no_dp_label = (
            "YES" if rec.get("no_dp")
            else "no" if _dp_mode == "dp-fedavg"
            else f"no ({_dp_mode})"
        )
        _data_cell(ws.cell(row_idx, 7),  _no_dp_label, alt_row=alt)
        # Yeom
        _auc_cell(ws.cell(row_idx, 8),  rec.get("auc_roc"),    alt)
        _auc_cell(ws.cell(row_idx, 9),  rec.get("auc_max"),    alt)
        _auc_cell(ws.cell(row_idx, 10), rec.get("auc_min"),    alt)
        # Shadow
        _auc_cell(ws.cell(row_idx, 11), rec.get("shadow_auc"), alt)
        _auc_cell(ws.cell(row_idx, 12), rec.get("shadow_max"), alt)
        # LiRA
        _auc_cell(ws.cell(row_idx, 13), rec.get("lira_auc"),   alt)
        _auc_cell(ws.cell(row_idx, 14), rec.get("lira_max"),   alt)
        # Privacy risk con colore
        risk_cell = ws.cell(row_idx, 15)
        risk = rec["privacy_risk"]
        _data_cell(risk_cell, risk, alt_row=alt, bold=True)
        risk_cell.font = _font(bold=True, color=_risk_label_color(risk))
        _data_cell(ws.cell(row_idx, 16), rec["total_alerts"], alt_row=alt)

    widths = [20, 10, 10, 16, 12, 10, 7, 16, 14, 14, 18, 16, 16, 14, 14, 12]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, get_column_letter(i), w)

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18


# ── Sheet 2: Heat Map ──────────────────────────────────────────────────────────

def build_heat_map(ws, records: list[dict]) -> None:
    ws.title = "Heat Map"

    # Fix 2026-07-22 (review A2): questa griglia è indicizzata solo su
    # (rounds, epsilon) — se mescolasse dp_mode diversi (dp-fedavg/central/local)
    # allo stesso (rounds, epsilon), un run "central" e uno "dp-fedavg" con lo
    # stesso ε si sovrascriverebbero silenziosamente a vicenda (vince l'ultimo
    # per timestamp). Restringere esplicitamente a dp-fedavg (il meccanismo che
    # questo foglio è nato per rappresentare) invece di ristrutturare la griglia
    # per un terzo asse — central/local DP si confrontano nel foglio Comparison.
    records = [r for r in records if r.get("dp_mode", "dp-fedavg") == "dp-fedavg"]

    # Raccoglie valori unici ordinati
    rounds_list  = sorted(set(r["rounds"] for r in records))
    epsilon_list = sorted(set(r["epsilon"] for r in records))

    # Indice (rounds, epsilon) → auc_roc
    data_map: dict[tuple, float | None] = {}
    for rec in records:
        key = (rec["rounds"], rec["epsilon"])
        # Se duplicato, tieni il più recente (già ordinati per timestamp)
        data_map[key] = rec["auc_roc"]

    # Titolo
    n_cols = len(epsilon_list) + 2
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    title = ws["A1"]
    title.value = "AUC-ROC Heat Map — FedMIA vs ε (Differential Privacy Budget)"
    title.font = _font(bold=True, color=COLOR_HEADER_FG, size=12)
    title.fill = _fill(COLOR_HEADER_BG)
    title.alignment = _center()

    # Sottotitolo
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    sub = ws["A2"]
    sub.value = (
        "AUC-ROC ≈ 0.50 → MIA non migliore del random (DP efficace)  |  "
        "AUC-ROC > 0.55 → MIA parzialmente efficace  |  "
        # Fix 2026-07-22 (review indipendente fresh-pass): etichetta stale,
        # residuo dell'epoca pre-Sprint-10 (singolo dataset mislabeled "JPL",
        # in realtà Caltech — vedi correzione in README) e pre-3-siti-reali.
        # I dati ora coprono 3 siti reali (Caltech/JPL/Office1) × più anni —
        # rimossa la cifra fissa, fuorviante e non più veritiera.
        "Dataset: ACN-Data — 3 siti reali (Caltech/JPL/Office1)  |  "
        "Solo dp_mode=dp-fedavg (central/local DP: vedi foglio Comparison)"
    )
    sub.font = _font(bold=False, color="404040", size=9)
    sub.fill = _fill("EBF3FB")
    sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Header riga 3: etichetta + epsilon columns
    _header_cell(ws.cell(3, 1), "FL Rounds \\ ε →", bg=COLOR_HEADER_BG)
    for col, eps in enumerate(epsilon_list, 2):
        _header_cell(ws.cell(3, col), f"ε = {eps}", bg=COLOR_SUBHDR_BG)
    _header_cell(ws.cell(3, len(epsilon_list) + 2), "Row Avg", bg=COLOR_SUBHDR_BG)

    # Righe dati
    for row_idx, rnd in enumerate(rounds_list, 4):
        alt = (row_idx % 2 == 0)
        _header_cell(ws.cell(row_idx, 1), f"{rnd} rounds", bg=COLOR_SUBHDR_BG)

        row_values = []
        for col_idx, eps in enumerate(epsilon_list, 2):
            val = data_map.get((rnd, eps))
            c = ws.cell(row_idx, col_idx)
            c.value = val
            c.font = _font(bold=val is not None and val > 0.51)
            c.fill = _fill(COLOR_ROW_ALT if alt else COLOR_ROW_PLAIN)
            c.border = _border()
            c.alignment = _center()
            c.number_format = "0.0000"
            if val is not None:
                row_values.append(val)
                # Colora AUC: viola (anomalo) → verde → giallo → rosso
                if val < ANOMALY_LOW_AUC:
                    c.fill = _fill("E1D5E7")   # viola chiaro — probabile bug, non DP efficace
                    c.font = _font(bold=True, color=COLOR_ANOMALY)
                elif val > 0.60:
                    c.fill = _fill("FFCCCC")   # rosso chiaro
                    c.font = _font(bold=True, color=COLOR_BAD)
                elif val > 0.52:
                    c.fill = _fill("FFE5B4")   # arancione chiaro
                    c.font = _font(bold=True, color="8B4513")
                else:
                    c.fill = _fill("D5E8D4")   # verde chiaro (DP efficace)
                    c.font = _font(bold=False, color="2D6A2D")

        # Media di riga
        avg_col = len(epsilon_list) + 2
        avg_cell = ws.cell(row_idx, avg_col)
        if row_values:
            avg = sum(row_values) / len(row_values)
            avg_cell.value = avg
            avg_cell.number_format = "0.0000"
            avg_cell.font = _font(bold=True)
        else:
            avg_cell.value = "N/A"
        avg_cell.fill = _fill("E2EFDA")
        avg_cell.border = _border()
        avg_cell.alignment = _center()

    # Riga media di colonna
    avg_row = len(rounds_list) + 4
    _header_cell(ws.cell(avg_row, 1), "Col Avg", bg=COLOR_SUBHDR_BG)
    for col_idx, eps in enumerate(epsilon_list, 2):
        col_vals = [
            data_map.get((rnd, eps))
            for rnd in rounds_list
            if data_map.get((rnd, eps)) is not None
        ]
        c = ws.cell(avg_row, col_idx)
        if col_vals:
            avg = sum(col_vals) / len(col_vals)
            c.value = avg
            c.number_format = "0.0000"
            c.font = _font(bold=True)
        else:
            c.value = "N/A"
        c.fill = _fill("E2EFDA")
        c.border = _border()
        c.alignment = _center()

    # Legenda
    legend_row = avg_row + 2
    ws.cell(legend_row, 1).value = "Legenda:"
    ws.cell(legend_row, 1).font = _font(bold=True)
    legends = [
        ("E1D5E7", COLOR_ANOMALY, f"AUC < {ANOMALY_LOW_AUC} — ANOMALO: probabile bug nell'attacco, non DP efficace"),
        ("D5E8D4", "2D6A2D", f"{ANOMALY_LOW_AUC} ≤ AUC ≤ 0.52 — DP efficace, MIA ≈ random"),
        ("FFE5B4", "8B4513", "0.52 < AUC ≤ 0.60 — leakage parziale"),
        ("FFCCCC", COLOR_BAD, "AUC > 0.60 — MIA efficace, rischio HIGH"),
    ]
    for i, (bg, fg, label) in enumerate(legends):
        c = ws.cell(legend_row + 1 + i, 2)
        c.value = label
        c.fill = _fill(bg)
        c.font = _font(color=fg, bold=True)
        c.border = _border()
        c.alignment = Alignment(horizontal="left")
        ws.merge_cells(
            start_row=legend_row + 1 + i, start_column=2,
            end_row=legend_row + 1 + i, end_column=5
        )

    # Larghezze
    _set_col_width(ws, "A", 16)
    for col_idx in range(2, len(epsilon_list) + 3):
        _set_col_width(ws, get_column_letter(col_idx), 14)

    ws.freeze_panes = "B4"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 18


# ── Sheet 3: Per Rounds ────────────────────────────────────────────────────────

def build_per_rounds(ws, records: list[dict]) -> None:
    ws.title = "Per Rounds"

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "AUC-ROC Statistics by Number of FL Rounds"
    t.font = _font(bold=True, color=COLOR_HEADER_FG, size=12)
    t.fill = _fill(COLOR_HEADER_BG)
    t.alignment = _center()

    headers = ["FL Rounds", "N Experiments", "AUC-ROC Mean", "AUC-ROC Min", "AUC-ROC Max", "Std Dev"]
    for col, h in enumerate(headers, 1):
        _header_cell(ws.cell(2, col), h, bg=COLOR_SUBHDR_BG)

    from statistics import mean, stdev

    rounds_list = sorted(set(r["rounds"] for r in records))
    for row_idx, rnd in enumerate(rounds_list, 3):
        alt = (row_idx % 2 == 0)
        group = [r["auc_roc"] for r in records if r["rounds"] == rnd and r["auc_roc"] is not None]
        _data_cell(ws.cell(row_idx, 1), rnd,           alt_row=alt, bold=True)
        _data_cell(ws.cell(row_idx, 2), len(group),    alt_row=alt)
        _data_cell(ws.cell(row_idx, 3), mean(group) if group else None, fmt="0.0000", alt_row=alt)
        _data_cell(ws.cell(row_idx, 4), min(group)  if group else None, fmt="0.0000", alt_row=alt)
        _data_cell(ws.cell(row_idx, 5), max(group)  if group else None, fmt="0.0000", alt_row=alt)
        _data_cell(ws.cell(row_idx, 6), stdev(group) if len(group) > 1 else None, fmt="0.0000", alt_row=alt)

    for col, w in zip("ABCDEF", [14, 16, 16, 14, 14, 12]):
        _set_col_width(ws, col, w)


# ── Sheet 4: Per Epsilon ───────────────────────────────────────────────────────

def build_per_epsilon(ws, records: list[dict]) -> None:
    ws.title = "Per Epsilon"

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "AUC-ROC Statistics by DP Budget (ε) — Privacy/Utility Trade-off"
    t.font = _font(bold=True, color=COLOR_HEADER_FG, size=12)
    t.fill = _fill(COLOR_HEADER_BG)
    t.alignment = _center()

    headers = ["Epsilon (ε)", "N Experiments", "AUC-ROC Mean", "AUC-ROC Min", "AUC-ROC Max", "Interpretation"]
    for col, h in enumerate(headers, 1):
        _header_cell(ws.cell(2, col), h, bg=COLOR_SUBHDR_BG)

    from statistics import mean

    epsilon_list = sorted(set(r["epsilon"] for r in records))
    interpretations = {
        0.1: "Strong DP — high noise, MIA likely ineffective",
        0.5: "Moderate-strong DP",
        1.0: "Standard DP budget",
        2.0: "Moderate DP — lower noise",
        5.0: "Weak DP — low noise, higher MIA risk",
    }
    for row_idx, eps in enumerate(epsilon_list, 3):
        alt = (row_idx % 2 == 0)
        group = [r["auc_roc"] for r in records if r["epsilon"] == eps and r["auc_roc"] is not None]
        avg = mean(group) if group else None

        _data_cell(ws.cell(row_idx, 1), eps,            fmt="0.0#",  alt_row=alt, bold=True)
        _data_cell(ws.cell(row_idx, 2), len(group),     alt_row=alt)
        auc_cell = ws.cell(row_idx, 3)
        _data_cell(auc_cell, avg, fmt="0.0000", alt_row=alt)
        # Fix 2026-07-21: un AUC medio anomalmente basso (< ANOMALY_LOW_AUC) segnala
        # un probabile bug nell'attacco (score invertito), non "privacy sicura" —
        # va colorato in modo distinto da un vero AUC≈0.5 (DP efficace).
        if avg is not None and avg < ANOMALY_LOW_AUC:
            auc_cell.font = _font(bold=True, color=COLOR_ANOMALY)
        elif avg is not None and avg > 0.55:
            auc_cell.font = _font(bold=True, color=COLOR_BAD)
        elif avg is not None and avg > 0.51:
            auc_cell.font = _font(bold=True, color=COLOR_WARN)
        else:
            auc_cell.font = _font(bold=True, color=COLOR_GOOD)

        _data_cell(ws.cell(row_idx, 4), min(group) if group else None, fmt="0.0000", alt_row=alt)
        _data_cell(ws.cell(row_idx, 5), max(group) if group else None, fmt="0.0000", alt_row=alt)

        interp_cell = ws.cell(row_idx, 6)
        interp_cell.value = interpretations.get(eps, "")
        interp_cell.font = _font()
        interp_cell.fill = _fill(COLOR_ROW_ALT if alt else COLOR_ROW_PLAIN)
        interp_cell.border = _border()
        interp_cell.alignment = Alignment(horizontal="left", vertical="center")

    for col, w in zip("ABCDEF", [14, 16, 16, 14, 14, 42]):
        _set_col_width(ws, col, w)


# ── Sheet 5: Comparison ────────────────────────────────────────────────────────

def build_comparison(ws, records: list[dict]) -> None:
    """Tabella di confronto: una colonna per esperimento, metriche in riga."""
    ws.title = "Comparison"

    n_exp = len(records)
    n_cols = n_exp + 1

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = "Multi-Experiment Comparison — FedMIA Attack vs DP (FedProx)"
    t.font = _font(bold=True, color=COLOR_HEADER_FG, size=12)
    t.fill = _fill(COLOR_HEADER_BG)
    t.alignment = _center()

    # Sub-header: nome esperimento per colonna
    _header_cell(ws.cell(2, 1), "Metric", bg=COLOR_HEADER_BG)
    for col, rec in enumerate(records, 2):
        label = f"{rec['rounds']} rounds\nε={rec['epsilon']}"
        _header_cell(ws.cell(2, col), label, bg=COLOR_SUBHDR_BG)
        ws.row_dimensions[2].height = 28

    # Righe metriche
    metrics = [
        ("FL Rounds",              "rounds",          None),
        ("Epsilon (ε)",            "epsilon",          "0.0#"),
        ("No-DP baseline",         "no_dp",            None),
        # dp_mode (2026-07-22): dp-fedavg (default) / central / local — vedi
        # docs/CaseStudies.md §2.4.3. Irrilevante se No-DP baseline = YES.
        ("DP Mode",                "dp_mode",          None),
        # ── Yeom 2018 (baseline debole) ──
        ("Yeom AUC (mean)",        "auc_roc",          "0.0000"),
        ("Yeom AUC (max)",         "auc_max",          "0.0000"),
        ("Yeom AUC (min)",         "auc_min",          "0.0000"),
        # ── Shadow calibrated (Carlini 2022, modello globale) ──
        ("Shadow AUC (mean)",      "shadow_auc",       "0.0000"),
        ("Shadow AUC (max)",       "shadow_max",       "0.0000"),
        # ── LiRA (Carlini 2022, raw_updates, PRIMARY) ──
        ("LiRA AUC (mean) ★",     "lira_auc",         "0.0000"),
        ("LiRA AUC (max)",         "lira_max",         "0.0000"),
        # ── Sintesi ──
        ("Primary Attack",         "primary_attack",   None),
        ("Privacy Risk",           "privacy_risk",     None),
        ("IDS Alerts (total)",     "total_alerts",     None),
        ("Byzantine Rounds",       "byzantine_rounds", None),
    ]

    for row_idx, (label, key, fmt) in enumerate(metrics, 3):
        alt = row_idx % 2 == 0
        _header_cell(ws.cell(row_idx, 1), label, bg=COLOR_SUBHDR_BG)
        for col, rec in enumerate(records, 2):
            val = rec.get(key)
            c = ws.cell(row_idx, col)
            _data_cell(c, val, fmt=fmt, alt_row=alt)
            # Colora AUC-ROC (Yeom, Shadow, LiRA) e Privacy Risk.
            # Fix 2026-07-22: mancavano shadow_auc/shadow_max/lira_auc/lira_max —
            # solo le righe Yeom (auc_roc/auc_max/auc_min) venivano colorate, quindi
            # un LiRA AUC anomalo (invertito o ≈0.14, l'esatto sintomo dei fix
            # 2026-07-21a-e) sarebbe rimasto testo nero non evidenziato proprio nel
            # foglio "Comparison" — trovato da una review indipendente.
            if key in (
                "auc_roc", "auc_max", "auc_min",
                "shadow_auc", "shadow_max",
                "lira_auc", "lira_max",
            ) and val is not None:
                c.font = _font(bold=True, color=_auc_risk_color(val))
            elif key == "privacy_risk":
                c.font = _font(bold=True, color=_risk_label_color(val))

    _set_col_width(ws, "A", 22)
    for col_idx in range(2, n_cols + 1):
        _set_col_width(ws, get_column_letter(col_idx), 18)
    ws.freeze_panes = "B3"


# ── Sheet 6: AUC Progression ───────────────────────────────────────────────────

def build_auc_progression(ws, records: list[dict]) -> None:
    """AUC-ROC round per round per ogni esperimento — mostra convergenza e drift."""
    ws.title = "AUC Progression"

    n_exp = len(records)
    n_cols = n_exp * 2 + 1  # round num + (AUC, score_mean) per esperimento

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = "AUC-ROC per Round — Andamento FedMIA Attack (convergenza modello FL)"
    t.font = _font(bold=True, color=COLOR_HEADER_FG, size=12)
    t.fill = _fill(COLOR_HEADER_BG)
    t.alignment = _center()

    sub = ws["A2"]
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    sub.value = (
        "AUC ≈ 0.50 su tutti i round → DP efficace.  "
        "Salita progressiva → memorizzazione crescente.  "
        "Ogni esperimento ha la propria sequenza di round."
    )
    sub.font = _font(size=9, color="404040")
    sub.fill = _fill("EBF3FB")
    sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24

    # Header riga 3
    _header_cell(ws.cell(3, 1), "Round #", bg=COLOR_HEADER_BG)
    for i, rec in enumerate(records):
        base_col = i * 2 + 2
        label = f"{rec['rounds']} rounds / ε={rec['epsilon']}"
        ws.merge_cells(
            start_row=3, start_column=base_col,
            end_row=3,   end_column=base_col + 1,
        )
        _header_cell(ws.cell(3, base_col), label, bg=COLOR_SUBHDR_BG)

    # Sub-header riga 4
    _header_cell(ws.cell(4, 1), "", bg=COLOR_HEADER_BG)
    for i in range(n_exp):
        base_col = i * 2 + 2
        _header_cell(ws.cell(4, base_col),     "AUC-ROC", bg=COLOR_SUBHDR_BG)
        _header_cell(ws.cell(4, base_col + 1), "FL Loss", bg=COLOR_SUBHDR_BG)

    # Raccoglie tutti i round number unici, ordinati
    all_rounds: set[int] = set()
    for rec in records:
        all_rounds.update(rec["per_round_auc"].keys())
    sorted_rounds = sorted(all_rounds)

    for row_idx, rnd in enumerate(sorted_rounds, 5):
        alt = row_idx % 2 == 0
        _data_cell(ws.cell(row_idx, 1), rnd, alt_row=alt, bold=True)
        for i, rec in enumerate(records):
            base_col = i * 2 + 2
            auc_val  = rec["per_round_auc"].get(rnd)
            loss_val = rec["per_round_loss"].get(rnd)

            auc_cell = ws.cell(row_idx, base_col)
            _data_cell(auc_cell, auc_val, fmt="0.0000", alt_row=alt)
            if auc_val is not None:
                if auc_val > 0.60:
                    auc_cell.fill = _fill("FFCCCC")
                    auc_cell.font = _font(bold=True, color=COLOR_BAD)
                elif auc_val > 0.52:
                    auc_cell.fill = _fill("FFE5B4")
                    auc_cell.font = _font(color="8B4513")
                elif auc_val < ANOMALY_LOW_AUC:
                    auc_cell.fill = _fill("E1D5E7")
                    auc_cell.font = _font(bold=True, color=COLOR_ANOMALY)
                else:
                    auc_cell.fill = _fill("D5E8D4")
                    auc_cell.font = _font(color="2D6A2D")

            _data_cell(ws.cell(row_idx, base_col + 1), loss_val, fmt="0.0000", alt_row=alt)

    _set_col_width(ws, "A", 10)
    for i in range(n_exp):
        base_col = i * 2 + 2
        _set_col_width(ws, get_column_letter(base_col),     12)
        _set_col_width(ws, get_column_letter(base_col + 1), 14)

    ws.freeze_panes = "B5"
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16


# ── Sheet 7: Attack Comparison ────────────────────────────────────────────────

def build_attack_comparison(ws, records: list[dict]) -> None:
    """
    Tabella di confronto diretto Yeom vs Shadow vs LiRA per ogni esperimento.
    Scopo: capire immediatamente quale attacco funziona meglio e quanto DP aiuta.
    Layout: ogni riga = un esperimento; colonne = metriche per ciascun attacco.
    """
    ws.title = "Attack Comparison"

    n_exp = len(records)
    n_cols = 12

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = "Attack Comparison — Yeom (2018) vs Shadow (Carlini) vs LiRA (Carlini 2022, PRIMARY)"
    t.font = _font(bold=True, color=COLOR_HEADER_FG, size=12)
    t.fill = _fill(COLOR_HEADER_BG)
    t.alignment = _center()

    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    sub = ws["A2"]
    sub.value = (
        "★ LiRA = attacco primario del paper (server intercetta raw_updates pre-FedProx).  "
        "Yeom = baseline debole (modello globale).  "
        "Shadow = calibrated (modello globale).  "
        "AUC > 0.55 → membership signal presente  |  AUC ≈ 0.50 → DP efficace."
    )
    sub.font = _font(size=9, color="404040")
    sub.fill = _fill("EBF3FB")
    sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Gruppi di header: identificazione + 3 colonne per attacco
    # Row 3: label gruppi
    _header_cell(ws.cell(3, 1), "Experiment", bg=COLOR_HEADER_BG)
    _header_cell(ws.cell(3, 2), "ε",          bg=COLOR_HEADER_BG)
    _header_cell(ws.cell(3, 3), "Rounds",     bg=COLOR_HEADER_BG)
    _header_cell(ws.cell(3, 4), "No-DP",      bg=COLOR_HEADER_BG)

    # Yeom group (blu)
    ws.merge_cells("E3:F3")
    _header_cell(ws.cell(3, 5), "Yeom 2018 (debole)", bg="4472C4")
    # Shadow group (arancione)
    ws.merge_cells("G3:H3")
    _header_cell(ws.cell(3, 7), "Shadow MIA (medio)", bg="ED7D31")
    # LiRA group (verde) — primary
    ws.merge_cells("I3:J3")
    _header_cell(ws.cell(3, 9), "★ LiRA (primario)", bg="375623")
    # Delta columns
    ws.merge_cells("K3:L3")
    _header_cell(ws.cell(3, 11), "Δ (LiRA − Yeom)", bg="7030A0")

    # Row 4: sub-header
    for col in [1, 2, 3, 4]:
        _header_cell(ws.cell(4, col), "", bg=COLOR_HEADER_BG)
    for col, label in zip([5, 6, 7, 8, 9, 10, 11, 12],
                          ["Mean", "Max", "Mean", "Max", "Mean", "Max", "Δ mean", "Δ max"]):
        bgs = {5: "4472C4", 6: "4472C4", 7: "ED7D31", 8: "ED7D31",
               9: "70AD47", 10: "70AD47", 11: "7030A0", 12: "7030A0"}
        _header_cell(ws.cell(4, col), label, bg=bgs[col])

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 16

    def _auc_col(cell, val, alt_row, bold=False):
        _data_cell(cell, val, fmt="0.0000", alt_row=alt_row, bold=bold)
        if val is None:
            cell.value = "—"
            return
        if val < ANOMALY_LOW_AUC:
            # Fix 2026-07-21: AUC anomalmente basso — probabile attacco rotto,
            # non "DP efficace". Va distinto visivamente dal verde "sicuro".
            cell.fill = _fill("E1D5E7")
            cell.font = _font(bold=True, color=COLOR_ANOMALY)
        elif val > 0.60:
            cell.fill = _fill("FFCCCC")
            cell.font = _font(bold=True, color=COLOR_BAD)
        elif val > 0.55:
            cell.fill = _fill("FFE5B4")
            cell.font = _font(bold=True, color="8B4513")
        elif val > 0.52:
            cell.fill = _fill("FFF2CC")
            cell.font = _font(color="8B4513")
        else:
            cell.fill = _fill("D5E8D4")
            cell.font = _font(color="2D6A2D")

    def _delta_col(cell, val, alt_row):
        _data_cell(cell, val, fmt="+0.0000;-0.0000;0.0000", alt_row=alt_row)
        if val is None:
            cell.value = "—"
            return
        if val > 0.05:
            cell.fill = _fill("FFCCCC")
            cell.font = _font(bold=True, color=COLOR_BAD)
        elif val > 0.01:
            cell.fill = _fill("FFE5B4")
            cell.font = _font(color="8B4513")
        else:
            cell.fill = _fill("D5E8D4")
            cell.font = _font(color="2D6A2D")

    for row_idx, rec in enumerate(records, 5):
        alt = (row_idx % 2 == 0)
        _data_cell(ws.cell(row_idx, 1), rec["timestamp"][:15],    alt_row=alt)
        _data_cell(ws.cell(row_idx, 2), rec["epsilon"], fmt="0.0#", alt_row=alt)
        _data_cell(ws.cell(row_idx, 3), rec["rounds"],             alt_row=alt)
        _data_cell(ws.cell(row_idx, 4), "YES" if rec.get("no_dp") else "no", alt_row=alt)

        yeom_mean  = rec.get("auc_roc")
        yeom_max   = rec.get("auc_max")
        shadow_mean = rec.get("shadow_auc")
        shadow_max  = rec.get("shadow_max")
        lira_mean   = rec.get("lira_auc")
        lira_max    = rec.get("lira_max")

        _auc_col(ws.cell(row_idx, 5),  yeom_mean,  alt)
        _auc_col(ws.cell(row_idx, 6),  yeom_max,   alt)
        _auc_col(ws.cell(row_idx, 7),  shadow_mean, alt)
        _auc_col(ws.cell(row_idx, 8),  shadow_max,  alt)
        _auc_col(ws.cell(row_idx, 9),  lira_mean,  alt, bold=True)
        _auc_col(ws.cell(row_idx, 10), lira_max,   alt)

        # Δ = LiRA − Yeom: positivo = LiRA vede più segnale (attacco più forte)
        delta_mean = (lira_mean - yeom_mean) if lira_mean is not None and yeom_mean is not None else None
        delta_max  = (lira_max  - yeom_max)  if lira_max  is not None and yeom_max  is not None else None
        _delta_col(ws.cell(row_idx, 11), delta_mean, alt)
        _delta_col(ws.cell(row_idx, 12), delta_max,  alt)

    # Widths
    widths = [18, 8, 8, 8, 12, 12, 12, 12, 14, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, get_column_letter(i), w)

    # Legenda
    leg_row = n_exp + 7
    ws.cell(leg_row, 1).value = "Legenda:"
    ws.cell(leg_row, 1).font = _font(bold=True)
    for i, (bg, fg, label) in enumerate([
        ("E1D5E7", COLOR_ANOMALY, f"AUC < {ANOMALY_LOW_AUC} — ANOMALO: probabile bug, non DP efficace"),
        ("D5E8D4", "2D6A2D", f"{ANOMALY_LOW_AUC} ≤ AUC ≤ 0.52 — DP efficace"),
        ("FFF2CC", "8B4513", "0.52 < AUC ≤ 0.55 — leakage debole"),
        ("FFE5B4", "8B4513", "0.55 < AUC ≤ 0.60 — leakage moderato"),
        ("FFCCCC", COLOR_BAD, "AUC > 0.60 — MIA efficace, HIGH risk"),
    ]):
        c = ws.cell(leg_row + 1 + i, 2)
        c.value = label
        c.fill = _fill(bg)
        c.font = _font(color=fg, bold=True)
        c.border = _border()
        ws.merge_cells(
            start_row=leg_row + 1 + i, start_column=2,
            end_row=leg_row + 1 + i, end_column=6,
        )

    ws.freeze_panes = "A5"


# ── Sheets 8-10: Per-Attack Round-by-Round ─────────────────────────────────────

def _build_per_round_sheet(
    ws,
    records: list[dict],
    data_key: str,
    attack_name: str,
    accent_bg: str,
    description: str,
) -> None:
    """
    Sheet generico per un singolo attacco: righe=round, colonne=esperimento.
    Mostra come l'AUC di quell'attacco evolve round-by-round al variare delle config.
    Usato per Yeom, Shadow MIA e LiRA — tre sheet separati, stesso layout.

    Args:
        data_key:    chiave nel record con il dict {round → AUC}
        attack_name: nome visualizzato nel titolo e nell'intestazione
        accent_bg:   colore hex degli header di colonna (diverso per attacco)
        description: testo del sottotitolo
    """
    ws.title = f"{attack_name} Per Round"
    n_exp  = len(records)
    n_cols = n_exp + 1  # col1=Round #, poi una colonna AUC per esperimento

    # ── Titolo ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = f"AUC-ROC per Round — {attack_name}"
    t.font  = _font(bold=True, color=COLOR_HEADER_FG, size=12)
    t.fill  = _fill(COLOR_HEADER_BG)
    t.alignment = _center()

    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    sub = ws["A2"]
    sub.value = description
    sub.font  = _font(size=9, color="404040")
    sub.fill  = _fill("EBF3FB")
    sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # ── Header riga 3 ──────────────────────────────────────────────────────────
    _header_cell(ws.cell(3, 1), "Round #", bg=COLOR_HEADER_BG)
    for i, rec in enumerate(records):
        col    = i + 2
        no_dp  = rec.get("no_dp", False)
        eps_str = "no-DP" if no_dp else f"ε={rec['epsilon']}"
        label  = f"{rec['rounds']}r / {eps_str}"
        _header_cell(ws.cell(3, col), label, bg=accent_bg)

    # ── Dati ───────────────────────────────────────────────────────────────────
    all_rounds: set[int] = set()
    for rec in records:
        all_rounds.update(rec.get(data_key, {}).keys())
    sorted_rounds = sorted(all_rounds)

    for row_idx, rnd in enumerate(sorted_rounds, 4):
        alt = row_idx % 2 == 0
        _data_cell(ws.cell(row_idx, 1), rnd, alt_row=alt, bold=True)
        for i, rec in enumerate(records):
            col     = i + 2
            auc_val = rec.get(data_key, {}).get(rnd)
            cell    = ws.cell(row_idx, col)
            _data_cell(cell, auc_val, fmt="0.0000", alt_row=alt)
            if auc_val is not None:
                if auc_val < ANOMALY_LOW_AUC:
                    cell.fill = _fill("E1D5E7")
                    cell.font = _font(bold=True, color=COLOR_ANOMALY)
                elif auc_val > 0.60:
                    cell.fill = _fill("FFCCCC")
                    cell.font = _font(bold=True, color=COLOR_BAD)
                elif auc_val > 0.55:
                    cell.fill = _fill("FFE5B4")
                    cell.font = _font(bold=True, color="8B4513")
                elif auc_val > 0.52:
                    cell.fill = _fill("FFF2CC")
                    cell.font = _font(color="8B4513")
                else:
                    cell.fill = _fill("D5E8D4")
                    cell.font = _font(color="2D6A2D")

    # ── Colonne ────────────────────────────────────────────────────────────────
    _set_col_width(ws, "A", 10)
    for i in range(n_exp):
        _set_col_width(ws, get_column_letter(i + 2), 16)

    ws.freeze_panes = "B4"

    # ── Legenda ────────────────────────────────────────────────────────────────
    leg_row = len(sorted_rounds) + 6
    ws.cell(leg_row, 1).value = "Legenda:"
    ws.cell(leg_row, 1).font  = _font(bold=True)
    merge_end = min(n_cols, 6)
    for j, (bg, fg, label) in enumerate([
        ("E1D5E7", COLOR_ANOMALY, f"AUC < {ANOMALY_LOW_AUC} — ANOMALO: probabile bug, non DP efficace"),
        ("D5E8D4", "2D6A2D", f"{ANOMALY_LOW_AUC} ≤ AUC ≤ 0.52 — DP efficace / no leakage"),
        ("FFF2CC", "8B4513", "0.52 < AUC ≤ 0.55 — leakage debole"),
        ("FFE5B4", "8B4513", "0.55 < AUC ≤ 0.60 — leakage moderato"),
        ("FFCCCC", COLOR_BAD, "AUC > 0.60 — MIA efficace, HIGH risk"),
    ]):
        c = ws.cell(leg_row + 1 + j, 2)
        c.value  = label
        c.fill   = _fill(bg)
        c.font   = _font(color=fg, bold=True)
        c.border = _border()
        if merge_end >= 3:
            ws.merge_cells(
                start_row=leg_row + 1 + j, start_column=2,
                end_row=leg_row + 1 + j,   end_column=merge_end,
            )


def build_yeom_per_round(ws, records: list[dict]) -> None:
    """Sheet 8 — Yeom 2018 AUC per round, per ogni esperimento."""
    _build_per_round_sheet(
        ws, records,
        data_key    = "per_round_auc",
        attack_name = "Yeom 2018",
        accent_bg   = "1F4E79",   # blu ChargeShield
        description = (
            "Yeom 2018 — loss-based MIA sul modello globale (baseline debole). "
            "Confonde 'campioni facili' con non-membri. "
            "AUC > 0.55 → leakage rilevabile; AUC ≈ 0.50 → DP efficace."
        ),
    )


def build_shadow_per_round(ws, records: list[dict]) -> None:
    """Sheet 9 — Shadow MIA AUC per round, per ogni esperimento."""
    _build_per_round_sheet(
        ws, records,
        data_key    = "per_round_shadow_auc",
        attack_name = "Shadow MIA",
        accent_bg   = "375623",   # verde scuro
        description = (
            "Shadow MIA — calibrated attack sul modello globale (baseline medio). "
            "Score = MSE(shadow) − MSE(target): corregge il bias dei campioni facili di Yeom. "
            "Più preciso di Yeom, ma comunque limitato al modello globale post-FedProx."
        ),
    )


def build_lira_per_round(ws, records: list[dict]) -> None:
    """Sheet 10 — LiRA AUC per round, per ogni esperimento (attacco primario ★)."""
    _build_per_round_sheet(
        ws, records,
        data_key    = "per_round_lira_auc",
        attack_name = "LiRA (★ PRIMARY)",
        accent_bg   = "7B2C2C",   # rosso scuro
        description = (
            "LiRA (Carlini et al. 2022) — attacco PRIMARIO ★. "
            "Server-side: intercetta raw_updates PRE-aggregazione FedProx. "
            "Score = log P(loss|IN) − log P(loss|OUT) via shadow models locali. "
            "AUC > 0.55 → DP NON protegge; AUC ≈ 0.50 → DP efficace → claim DSN 2027 validato."
        ),
    )


# ── Sheet 11: Seed Aggregation — mean±std per gruppo (rounds, ε, no_dp) ────────

def build_seed_aggregation(ws, records: list[dict]) -> None:
    """
    Sheet 11 — Aggrega tutti i run con lo stesso (rounds, epsilon, no_dp) e
    calcola mean ± std per Yeom / Shadow / LiRA.

    Scopo: riportare i risultati del paper DSN 2027 con errore statistico.
    Il foglio è vuoto se ci sono solo 1 seed per configurazione — in quel caso
    usa prima 'make experiment-nodp-sweep' o 'make experiment-dp-sweep'.

    Raggruppa per (fl_rounds, epsilon, no_dp, dp_mode).  Ogni gruppo diventa
    una riga. Ordine: prima no-DP, poi DP crescente per ε (a parità di ε,
    dp-fedavg/central/local restano gruppi separati — fix 2026-07-22, review
    A2: senza dp_mode nella chiave, seed di modalità DP diverse allo stesso
    ε finirebbero mediati insieme in una sola riga senza alcun avviso).
    """
    import statistics

    ws.title = "Seed Aggregation"

    n_cols = 13  # vedi headers sotto

    # ── Titolo ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = "Seed Aggregation — mean ± std AUC-ROC per (rounds, ε, no-DP)  [DSN 2027]"
    t.font  = _font(bold=True, color=COLOR_HEADER_FG, size=12)
    t.fill  = _fill(COLOR_HEADER_BG)
    t.alignment = _center()

    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    sub = ws["A2"]
    sub.value = (
        "Ogni riga = media su N seed con la stessa configurazione (rounds, ε, no-DP).  "
        "N ≥ 5 per riportare mean±std nel paper (minimo DSN 2027).  "
        "Std = None con N=1 (aumenta seed con experiment-nodp-sweep / experiment-dp-sweep)."
    )
    sub.font  = _font(size=9, color="404040")
    sub.fill  = _fill("EBF3FB")
    sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # ── Headers ────────────────────────────────────────────────────────────────
    headers = [
        ("Rounds",       "1F4E79"),
        ("ε / no-DP",    "1F4E79"),
        ("N Seed",       "1F4E79"),
        ("Seeds usati",  "1F4E79"),
        ("Yeom mean",    "2E75B6"),
        ("Yeom std",     "2E75B6"),
        ("Shadow mean",  "375623"),
        ("Shadow std",   "375623"),
        ("LiRA mean ★",  "7B2C2C"),
        ("LiRA std",     "7B2C2C"),
        ("LiRA min",     "7B2C2C"),
        ("LiRA max",     "7B2C2C"),
        ("Δ LiRA−Yeom",  "4A4A4A"),
    ]
    for col_idx, (label, bg) in enumerate(headers, 1):
        _header_cell(ws.cell(3, col_idx), label, bg=bg)

    # ── Raggruppa records ──────────────────────────────────────────────────────
    groups: dict[tuple, list[dict]] = {}
    for rec in records:
        key = (
            rec["rounds"], rec["epsilon"], rec.get("no_dp", False),
            rec.get("dp_mode", "dp-fedavg"),
        )
        groups.setdefault(key, []).append(rec)

    # Ordine: no-DP prima, poi DP crescente per ε e per dp_mode; round crescenti per parità
    def _sort_key(k):
        rounds, eps, no_dp, dp_mode = k
        return (0 if no_dp else 1, rounds, eps, dp_mode)

    sorted_keys = sorted(groups.keys(), key=_sort_key)

    def _mean_or_none(vals):
        vals = [v for v in vals if v is not None]
        return statistics.mean(vals) if vals else None

    def _std_or_none(vals):
        vals = [v for v in vals if v is not None]
        if len(vals) < 2:
            return None
        return statistics.stdev(vals)

    def _auc_fill(cell, val, bold=False):
        """Colora la cella AUC per livello di rischio."""
        if val is None:
            cell.value = "—"
            return
        cell.value       = val
        cell.number_format = "0.0000"
        cell.font        = _font(bold=bold)
        cell.border      = _border()
        cell.alignment   = _center()
        if val < ANOMALY_LOW_AUC:
            # Fix 2026-07-21: questa e' la stessa riga che prima mostrava un
            # lira_auc_roc di 0.14-0.32 (experiments/exp3, pre-fix run_lira())
            # in VERDE come "LOW risk" — era un attacco rotto, non DP efficace.
            cell.fill = _fill("E1D5E7"); cell.font = _font(bold=True, color=COLOR_ANOMALY)
        elif val > 0.60:
            cell.fill = _fill("FFCCCC"); cell.font = _font(bold=True, color=COLOR_BAD)
        elif val > 0.55:
            cell.fill = _fill("FFE5B4"); cell.font = _font(bold=True, color="8B4513")
        elif val > 0.52:
            cell.fill = _fill("FFF2CC"); cell.font = _font(color="8B4513")
        else:
            cell.fill = _fill("D5E8D4"); cell.font = _font(color="2D6A2D", bold=bold)

    for row_idx, key in enumerate(sorted_keys, 4):
        rounds, eps, no_dp, dp_mode = key
        recs = groups[key]
        alt  = row_idx % 2 == 0

        seeds_used = sorted(rec.get("seed", 42) for rec in recs)
        n_seeds    = len(recs)
        if no_dp:
            eps_label = "no-DP" if dp_mode == "dp-fedavg" else f"no-DP ({dp_mode})"
        else:
            eps_label = str(eps) if dp_mode == "dp-fedavg" else f"{eps} ({dp_mode})"

        yeom_vals   = [r.get("auc_roc")    for r in recs]
        shadow_vals = [r.get("shadow_auc") for r in recs]
        lira_vals   = [r.get("lira_auc")   for r in recs]
        lira_min_v  = [r.get("lira_min")   for r in recs]
        lira_max_v  = [r.get("lira_max")   for r in recs]

        yeom_mean   = _mean_or_none(yeom_vals)
        yeom_std    = _std_or_none(yeom_vals)
        shadow_mean = _mean_or_none(shadow_vals)
        shadow_std  = _std_or_none(shadow_vals)
        lira_mean   = _mean_or_none(lira_vals)
        lira_std    = _std_or_none(lira_vals)
        lira_min_m  = _mean_or_none(lira_min_v)
        lira_max_m  = _mean_or_none(lira_max_v)
        delta       = (lira_mean - yeom_mean) if (lira_mean is not None and yeom_mean is not None) else None

        _data_cell(ws.cell(row_idx, 1), rounds,    alt_row=alt, bold=True)
        _data_cell(ws.cell(row_idx, 2), eps_label, alt_row=alt)
        _data_cell(ws.cell(row_idx, 3), n_seeds,   alt_row=alt, bold=(n_seeds >= 5))
        _data_cell(ws.cell(row_idx, 4), ", ".join(str(s) for s in seeds_used), alt_row=alt)

        _auc_fill(ws.cell(row_idx, 5),  yeom_mean)
        if yeom_std is not None:
            _data_cell(ws.cell(row_idx, 6), yeom_std, fmt="0.0000", alt_row=alt)
        else:
            _data_cell(ws.cell(row_idx, 6), "—", alt_row=alt)

        _auc_fill(ws.cell(row_idx, 7),  shadow_mean)
        if shadow_std is not None:
            _data_cell(ws.cell(row_idx, 8), shadow_std, fmt="0.0000", alt_row=alt)
        else:
            _data_cell(ws.cell(row_idx, 8), "—", alt_row=alt)

        _auc_fill(ws.cell(row_idx, 9),  lira_mean, bold=True)
        if lira_std is not None:
            _data_cell(ws.cell(row_idx, 10), lira_std, fmt="0.0000", alt_row=alt)
        else:
            _data_cell(ws.cell(row_idx, 10), "—", alt_row=alt)
        _data_cell(ws.cell(row_idx, 11), lira_min_m, fmt="0.0000", alt_row=alt)
        _data_cell(ws.cell(row_idx, 12), lira_max_m, fmt="0.0000", alt_row=alt)

        # Δ = LiRA_mean − Yeom_mean
        d_cell = ws.cell(row_idx, 13)
        _data_cell(d_cell, delta, fmt="+0.0000;-0.0000;0.0000", alt_row=alt)
        if delta is not None:
            if delta > 0.05:
                d_cell.fill = _fill("FFCCCC"); d_cell.font = _font(bold=True, color=COLOR_BAD)
            elif delta > 0.01:
                d_cell.fill = _fill("FFE5B4"); d_cell.font = _font(color="8B4513")
            else:
                d_cell.fill = _fill("D5E8D4"); d_cell.font = _font(color="2D6A2D")

        # N seed < 5 → cella gialla (warning: non sufficiente per paper)
        n_cell = ws.cell(row_idx, 3)
        if n_seeds < 5:
            n_cell.fill = _fill("FFF2CC")
            n_cell.font = _font(color="8B4513", bold=True)

    # ── Widths ─────────────────────────────────────────────────────────────────
    widths = [8, 10, 8, 22, 12, 10, 12, 10, 14, 10, 10, 10, 14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, get_column_letter(i), w)
    ws.freeze_panes = "A4"

    # ── Legenda N-seed ─────────────────────────────────────────────────────────
    leg_row = len(sorted_keys) + 6
    ws.cell(leg_row, 1).value = "Legenda N Seed:"
    ws.cell(leg_row, 1).font  = _font(bold=True)
    for j, (bg, fg, label) in enumerate([
        ("FFF2CC", "8B4513", "N < 5  — non sufficiente per paper (aumenta con experiment-nodp-sweep)"),
        (COLOR_ROW_PLAIN, "000000", "N ≥ 5  — minimo DSN 2027 (mean±std)"),
        (COLOR_ROW_PLAIN, "000000", "N ≥ 10 — sufficiente per Wilcoxon significance test"),
    ]):
        c = ws.cell(leg_row + 1 + j, 2)
        c.value  = label
        c.fill   = _fill(bg)
        c.font   = _font(color=fg)
        c.border = _border()
        ws.merge_cells(
            start_row=leg_row + 1 + j, start_column=2,
            end_row=leg_row + 1 + j,   end_column=8,
        )


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Genera report Excel sweep ChargeShield-FL")
    parser.add_argument(
        "--output", type=Path,
        default=EXPERIMENTS_DIR / "ChargeShield_FL_Results.xlsx",
    )
    args = parser.parse_args()

    records = load_experiments()
    if not records:
        print("ERROR: Nessun file experiment_*.json trovato in experiments/")
        sys.exit(1)

    print(f"Caricati {len(records)} esperimenti.")

    wb = Workbook()
    wb.remove(wb.active)  # rimuovi sheet vuoto di default

    ws_raw    = wb.create_sheet("Raw Data")
    ws_heat   = wb.create_sheet("Heat Map")
    ws_round  = wb.create_sheet("Per Rounds")
    ws_eps    = wb.create_sheet("Per Epsilon")
    ws_comp   = wb.create_sheet("Comparison")
    ws_prog   = wb.create_sheet("AUC Progression")
    ws_atk    = wb.create_sheet("Attack Comparison")
    ws_yeom   = wb.create_sheet("Yeom Per Round")
    ws_shadow = wb.create_sheet("Shadow Per Round")
    ws_lira   = wb.create_sheet("LiRA Per Round")
    ws_sagg   = wb.create_sheet("Seed Aggregation")

    build_raw_data(ws_raw, records)
    build_heat_map(ws_heat, records)
    build_per_rounds(ws_round, records)
    build_per_epsilon(ws_eps, records)
    build_comparison(ws_comp, records)
    build_auc_progression(ws_prog, records)
    build_attack_comparison(ws_atk, records)
    build_yeom_per_round(ws_yeom, records)
    build_shadow_per_round(ws_shadow, records)
    build_lira_per_round(ws_lira, records)
    build_seed_aggregation(ws_sagg, records)

    # Proprietà workbook
    wb.properties.title   = "ChargeShield-FL Experiment Results"
    wb.properties.subject = "FedMIA vs Differential Privacy — DSN 2027"
    wb.properties.creator = "ChargeShield-FL Framework"

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Report salvato: {args.output}")


if __name__ == "__main__":
    main()
