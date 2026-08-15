#!/usr/bin/env python3
# scripts/run_experiments.py
# ChargeShield-FL — Sprint 5: Experiment Runner
#
# Esegue il ciclo completo:
#   1. Carica ACN-Data JPL 2019 + 2020
#   2. Esegue FL rounds via ML Plane (AutoencoderTrainer + GradientManager + FedAvgAggregator)
#   3. Lancia FedMIA attack per ogni round
#   4. Valuta IDS come baseline defense
#   5. Misura AUC-ROC e privacy/utility trade-off (epsilon vs AUC-ROC)
#   6. Salva risultati in experiments/
#
# Usage:
#   python scripts/run_experiments.py --config config/experiment.yaml
#   python scripts/run_experiments.py --epsilon 0.5 --rounds 10
#   python scripts/run_experiments.py --dry-run

from __future__ import annotations

import argparse
import json
import logging
import random
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import torch
import yaml

# ── Path setup ─────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from adapters.acn_dataset import ACNDataset
from auditor.privacy_auditor import PrivacyAuditor
from core.autoencoder import Autoencoder
from ids.charging_ids import ChargingIDS
from ml.autoencoder_trainer import AutoencoderTrainer
from ml.fedavg_aggregator import FedAvgAggregator
from ml.gradient_manager import GradientManager
from ml.ml_plane import FLArtifactCollector, MLPlane
from ml.base_ml import MLPlaneEvent

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("run_experiment")


# ── Config ─────────────────────────────────────────────────────────────────────

def load_config(config_path: Path | None, overrides: dict) -> dict:
    """Carica config da YAML e applica override da CLI."""
    if config_path and config_path.exists():
        with open(config_path) as f:
            cfg = yaml.safe_load(f)
    else:
        raise FileNotFoundError(
            f"Config non trovata: {config_path}. "
            "Specifica --config o crea config/experiment.yaml"
        )
    if overrides.get("epsilon") is not None:
        cfg["experiment"]["epsilon"] = overrides["epsilon"]
    if overrides.get("rounds") is not None:
        cfg["experiment"]["fl_rounds"] = overrides["rounds"]
    return cfg


# ── Dataset ────────────────────────────────────────────────────────────────────

# Mappa siteID ACN-Data (campo "site_id" già estratto da ACNDataset.get_sample(),
# vedi src/adapters/acn_dataset.py) → nome sito leggibile usato come cluster_id
# nell'FL. Verificata empiricamente 2026-07-22 confrontando il numero di stazioni
# uniche per siteID coi valori ufficiali pubblicati su https://ev.caltech.edu/dataset
# (Caltech 54 EVSE, JPL 50 EVSE, Office 1 8 EVSE): siteID="0002"→54 stazioni→Caltech,
# siteID="0001"→52 stazioni (≈50 ufficiali)→JPL, siteID="0019"→8 stazioni→Office 1.
# FIX 2026-07-22 (scoperta importante): il progetto ha SEMPRE chiamato "jpl" il
# dataset con siteID="0002" (54 stazioni) — che in realtà è Caltech, non JPL (50
# EVSE). Ogni esperimento precedente (nodp-sweep1/dp-sweep1 inclusi) ha quindi
# usato dati Caltech mal etichettati come "jpl" — vedi README/CaseStudies.md per
# la correzione della documentazione storica. Questa mappa ora usa l'identità
# corretta, verificata via conteggio stazioni, non il nome storico del file.
_SITE_ID_TO_NAME = {
    "0002": "caltech",
    "0001": "jpl",
    "0019": "office1",
}


def load_sessions(cfg: dict) -> list[dict[str, Any]]:
    """
    Carica sessioni EV da ACN-Data, da TUTTI i siti/anni elencati in
    cfg["sites"] (2026-07-22: sostituisce il precedente cfg["datasets"] — un
    solo dataset condiviso affettato arbitrariamente in 4 "cluster" fittizi).
    cfg["sites"] è {nome_sito: [path_anno1, path_anno2, ...]} — ogni sito reale
    combina tutti gli anni disponibili in un unico pool (stessa logica già
    usata per jpl_2019+jpl_2020 prima di oggi, ora per sito invece che globale).

    Restituisce una lista PIATTA (stesso contratto di sempre — compute_feature_stats/
    normalize_sessions/il train-holdout split lavorano su questa lista intera).
    L'appartenenza al sito reale di ciascuna sessione resta comunque disponibile
    nel campo "site_id" (estratto da ACNDataset) — vedi group_sessions_by_site()
    per il raggruppamento usato da run_fl_rounds()/run_lira().
    """
    sessions: list[dict[str, Any]] = []
    sites_cfg = cfg.get("sites") or cfg.get("datasets") or {}
    for site_name, paths in sites_cfg.items():
        path_list = paths if isinstance(paths, list) else [paths]
        site_count = 0
        for path_str in path_list:
            p = PROJECT_ROOT / path_str
            if not p.exists():
                logger.warning(f"Dataset non trovato: {p} — skip")
                continue
            dataset = ACNDataset()
            dataset.load(str(p))
            loaded = [dataset.get_sample(i) for i in range(len(dataset))]
            sessions.extend(loaded)
            site_count += len(loaded)
        logger.info(f"{site_name}: {site_count} sessioni caricate (tutti gli anni)")
    if not sessions:
        raise FileNotFoundError(
            "Nessun dataset trovato. Scarica ACN-Data da "
            "https://ev.caltech.edu/dataset e posizionalo in datasets/acn/<sito>/"
        )
    logger.info(f"Totale sessioni (tutti i siti): {len(sessions)}")
    return sessions


def group_sessions_by_site(sessions: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """
    Raggruppa le sessioni per SITO REALE, usando il campo "site_id" già presente
    in ogni sessione (estratto da ACNDataset dal siteID di ACN-Data) — non per
    slicing posizionale/contiguo come in precedenza.

    FIX 2026-07-22 (review indipendente su scripts/run_nvflare_mia.py, stesso
    principio applicato qui): ricostruire l'appartenenza a un cluster affettando
    una lista per indice (es. "le prime N/4 sessioni sono il cluster A") è
    fragile — dipende dall'ordine/dalla lunghezza esatta della lista a monte, e
    un mismatch tra chi produce la lista e chi la riaffetta produce risultati
    sbagliati SENZA errori visibili (esattamente il bug trovato oggi). Usare il
    campo site_id proprio di ogni sessione elimina questa intera classe di bug:
    il raggruppamento è corretto qualunque sia l'ordine/la provenienza della
    lista di sessioni.

    Sessioni con site_id sconosciuto (non in _SITE_ID_TO_NAME) sono raggruppate
    sotto il loro site_id grezzo invece di essere scartate — permette di
    scoprire nuovi siti aggiunti in futuro senza modificare questa funzione.
    """
    groups: dict[str, list[dict[str, Any]]] = {}
    for s in sessions:
        site_id = s.get("site_id", "")
        name = _SITE_ID_TO_NAME.get(site_id, site_id or "unknown")
        groups.setdefault(name, []).append(s)
    return groups


def group_indices_by_site(sessions: list[dict[str, Any]]) -> dict[str, list[int]]:
    """
    Come group_sessions_by_site(), ma restituisce gli INDICI GLOBALI in
    `sessions` invece delle sessioni stesse — usato da run_lira() (fase MIA),
    che deve campionare gli shadow model da un pool di indici nello stesso
    spazio di train_sessions (per costruire, dato un round, i sotto-tensori
    IN/OUT), non dalle sessioni stesse. Vedi cluster_membership in run_lira().
    """
    groups: dict[str, list[int]] = {}
    for i, s in enumerate(sessions):
        site_id = s.get("site_id", "")
        name = _SITE_ID_TO_NAME.get(site_id, site_id or "unknown")
        groups.setdefault(name, []).append(i)
    return groups


def inject_synthetic_client_indices(
    real_index_groups: dict[str, list[int]],
    n_synthetic: int = 2,
    seed: int = 42,
) -> dict[str, list[int]]:
    """
    Aggiunge n_synthetic client FITTIZI a real_index_groups (l'output di
    group_indices_by_site()) — SOLO per lo sweep IDS/Byzantine (config/
    experiment.yaml: byzantine_attack.enabled=True). MAI usato per l'esperimento
    privacy/FedMIA/LiRA principale, che deve vedere SOLO i 3 client reali
    (caltech/jpl/office1) — questa funzione va chiamata unicamente sul percorso
    codice dello sweep Byzantine, mai su quello di default.

    Perché servono client fittizi: Krum (usato da ChargingIDS per il rilevamento
    Byzantine) garantisce di rilevare f nodi Byzantine solo con n≥2f+3 nodi
    totali. Con f=1 (un solo attaccante, unico scenario oggi supportato)
    servono n≥5 client totali — i soli 3 siti reali non bastano su basi
    teoriche solide. Nessun 4°/5° sito REALE esiste in ACN-Data (solo Caltech/
    JPL/Office1, verificato su https://ev.caltech.edu/dataset, sezione "Sites")
    — un vero 4°/5° sito richiederebbe un dataset EV completamente diverso.

    Come sono costruiti: pool = concatenazione di TUTTI gli indici reali (3
    siti, nell'ordine di iterazione di real_index_groups), shuffle con seed
    fisso, poi affettato in n_synthetic parti aggiuntive. Le sessioni
    referenziate possono sovrapporsi con quelle già assegnate ai client reali —
    accettabile SOLO per lo scopo di validazione IDS (Krum verifica un
    comportamento geometrico locale — un gradiente scalato artificialmente —
    non richiede popolazioni disgiunte come farebbe invece un esperimento di
    privacy/generalizzazione). Per questo motivo l'uso di questi client
    fittizi in FedMIA/LiRA sarebbe metodologicamente invalido e va evitato.
    """
    import random as _random

    pooled_indices: list[int] = [i for idxs in real_index_groups.values() for i in idxs]
    rng = _random.Random(seed + 271828)  # offset arbitrario, separato dal seed sperimentale
    shuffled = pooled_indices[:]
    rng.shuffle(shuffled)

    expanded: dict[str, list[int]] = dict(real_index_groups)
    chunk_size = max(1, len(shuffled) // n_synthetic)
    for i in range(n_synthetic):
        start = i * chunk_size
        end = len(shuffled) if i == n_synthetic - 1 else start + chunk_size
        expanded[f"synthetic_{i + 1}"] = shuffled[start:end]
    return expanded

# ── Session enrichment ─────────────────────────────────────────────────────────
def enrich_sessions(sessions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Aggiunge feature derivate dai timestamp ACN-Data.
    - hour_of_day: ora di connessione LOCALE al sito (0–23), pattern comportamentale
    - duration_hours: durata sessione in ore, correlata all'energia

    Fix 2026-07-22 (review indipendente fresh-pass, bug reale confermato
    empiricamente): start_time/end_time (da ACNDataset) sono UTC nonostante il
    suffisso "GMT" nel dato grezzo ACN-Data sia fuorviante — vedi commento su
    "timezone" in src/adapters/acn_dataset.py per la verifica empirica (picco
    orario grezzo di office1 implausibile per charging da ufficio, plausibile
    dopo conversione a America/Los_Angeles). hour_of_day va quindi calcolato
    sull'ora LOCALE del sito, non sui pesi grezzi di start.hour (che erano UTC,
    sistematicamente sfasati di 7-8h da quanto la feature dichiara di
    rappresentare — "ora di connessione, pattern comportamentale"). Non
    tocchiamo start_time/end_time stessi (restano UTC — usati altrove per
    year-splitting/holdout, che non deve essere influenzato da questo fix) né
    duration_hours (invariante a un offset UTC uniforme, entrambi i lati
    dell'intervallo si spostano della stessa quantità).
    """
    from zoneinfo import ZoneInfo

    enriched = []
    for s in sessions:
        try:
            start = datetime.fromisoformat(s["start_time"])
            end   = datetime.fromisoformat(s["end_time"])

            tz_name = s.get("timezone")
            if tz_name:
                try:
                    local_start = start.replace(tzinfo=ZoneInfo("UTC")).astimezone(
                        ZoneInfo(tz_name)
                    )
                    hour_of_day = float(local_start.hour)
                except Exception:
                    # Timezone IANA sconosciuta/malformata — fallback all'ora
                    # grezza (comportamento pre-fix) invece di scartare la
                    # sessione: un hour_of_day leggermente sfasato è preferibile
                    # a perdere il campione.
                    hour_of_day = float(start.hour)
            else:
                hour_of_day = float(start.hour)  # nessun timezone noto — fallback

            s["hour_of_day"]    = hour_of_day
            s["duration_hours"] = max(0.0, (end - start).total_seconds() / 3600.0)
            enriched.append(s)
        except (KeyError, ValueError):
            pass  # scarta sessioni con timestamp malformati
    return enriched


# ── Feature Normalization ──────────────────────────────────────────────────────

def compute_feature_stats(
    sessions: list[dict[str, Any]],
    features: list[str],
) -> dict[str, tuple[float, float]]:
    """
    Calcola min e max per ogni feature dalle sessioni di training.
    Chiamare SOLO su train_sessions per evitare data leakage dal hold-out.
    """
    stats: dict[str, tuple[float, float]] = {}
    for feat in features:
        vals = []
        for s in sessions:
            raw = s.get(feat)
            if raw is None:
                continue
            try:
                vals.append(float(raw))
            except (ValueError, TypeError):
                continue  # salta valori non numerici (corrotti o stringhe errate)
        if not vals:
            stats[feat] = (0.0, 1.0)
            continue
        fmin, fmax = min(vals), max(vals)
        stats[feat] = (fmin, fmax if fmax != fmin else fmin + 1.0)
    return stats


def normalize_sessions(
    sessions: list[dict[str, Any]],
    stats: dict[str, tuple[float, float]],
    features: list[str],
) -> list[dict[str, Any]]:
    """
    Applica min-max scaling [0,1] alle feature continue.
    stats deve provenire da compute_feature_stats(train_sessions, ...).
    """
    normalized = []
    for s in sessions:
        s = dict(s)  # shallow copy — non modificare l'originale
        for feat in features:
            val = s.get(feat)
            if val is None:
                continue
            fmin, fmax = stats[feat]
            s[feat] = (float(val) - fmin) / (fmax - fmin)
        normalized.append(s)
    return normalized

# ── FL Experiment ──────────────────────────────────────────────────────────────

def run_fl_rounds(
    cfg: dict,
    sessions: list[dict[str, Any]],
    no_dp: bool = False,
    dp_mode: str = "dp-fedavg",
    cluster_sessions: dict[str, list[dict[str, Any]]] | None = None,
) -> dict[int, dict[str, Any]]:
    """
    Esegue FL rounds via ML Plane.
    Ogni round: train locale → [DP opzionale] → FedAvg → global model.
    Restituisce gradient history per round.

    Args:
        cfg:     configurazione esperimento
        sessions: sessioni di training (usato SOLO come fallback se
                 cluster_sessions non è fornito — vedi sotto).
        cluster_sessions: sessioni GIÀ raggruppate per cluster/sito reale
                 (2026-07-22, vedi group_sessions_by_site()) — {cluster_id:
                 [sessioni]}. Se fornito, `sessions` viene ignorato e i cluster
                 sono esattamente quelli passati (es. 3 siti reali: caltech/
                 jpl/office1, o 5 per lo sweep IDS/Byzantine — vedi
                 inject_synthetic_clients()). Se None (default, retrocompatibile
                 coi test esistenti che usano dati sintetici senza site_id
                 reale), affetta `sessions` in 4 parti uguali fittizie — stesso
                 comportamento storico pre-2026-07-22.
        no_dp:   se True, salta il rumore DP (σ=0) — usato per baseline experiment.
                 Permette di distinguere:
                   Scenario A: DP funziona → AUC > 0.5 senza DP, ≈0.5 con DP
                   Scenario B: modello non memorizza → AUC ≈ 0.5 in entrambi i casi
        dp_mode: quale placement DP usare quando no_dp=False (2026-07-22, vedi
                 docs/CaseStudies.md §2.4.3 per la tassonomia completa):
                   "dp-fedavg" (default, comportamento storico) — il server
                     riceve l'update raw di ogni client e lo clippa+rumorizza
                     PRIMA di aggregarlo. Questo placement (rumore per-client
                     prima dell'aggregazione) è una variante più restrittiva e
                     non-standard, non descritta letteralmente nell'Algoritmo 1
                     di McMahan et al. 2018 — vedi "central" sotto per il mode
                     a cui quel paper corrisponde davvero. Il server/IDS vede
                     transitoriamente l'update raw (usato oggi da run_ids()).
                   "central" [McMahan et al. 2018 — meccanismo esatto del loro
                     Algoritmo 1 (DP-FedAvg): clip lato client, un solo draw di
                     rumore lato server sull'aggregato] — ogni client CLIPPA il
                     proprio update (bound la sensitività) ma NON lo rumorizza;
                     il server (trusted) aggrega gli update puliti-ma-clippati
                     con FedAvg, poi aggiunge UN SOLO rumore Gaussiano
                     all'aggregato (scalato 1/n_participants). I singoli update
                     non sono mai rumorizzati: un attacco sul singolo update
                     (LiRA) non dovrebbe mostrare alcuna soppressione — è il
                     risultato atteso, non un bug.
                   "local" — stesso meccanismo per-client di "dp-fedavg", ma il
                     server/IDS non deve MAI vedere l'update raw, nemmeno
                     transitoriamente: run_ids() userà `updates` (rumorizzati)
                     invece di `raw_updates` (che in questa modalità non viene
                     salvato in fl_results — vedi sotto).
                 Ignorato se no_dp=True (nessun rumore in nessun caso).
    """
    exp_cfg  = cfg["experiment"]
    ml_cfg   = cfg["ml"]
    fl_rounds = exp_cfg["fl_rounds"]

    if cluster_sessions is not None:
        # Sessioni già raggruppate per sito/cluster reale (2026-07-22) — vedi
        # group_sessions_by_site()/inject_synthetic_clients(). `sessions` (il
        # parametro posizionale) viene ignorato in questo ramo.
        cluster_ids = list(cluster_sessions.keys())
    else:
        # Fallback storico (pre-2026-07-22, retrocompatibile coi test esistenti
        # che usano dati sintetici senza site_id reale): affetta `sessions` in
        # 4 parti uguali fittizie, come sempre fatto finora.
        cluster_ids = ["highway", "urban", "residential", "corporate"]
        cluster_size = max(1, len(sessions) // len(cluster_ids))
        cluster_sessions = {}
        for i, cid in enumerate(cluster_ids):
            start = i * cluster_size
            end   = None if i == len(cluster_ids) - 1 else start + cluster_size
            cluster_sessions[cid] = sessions[start:end]

    # Inizializza trainer per ogni cluster
    #
    # LIMITE NOTO, NON RISOLTO (osservazione review indipendente round 4,
    # 2026-07-24): a differenza di run_lira()/run_fedmia_shadow() (che
    # chiamano torch.manual_seed() esplicitamente subito prima di ogni
    # Autoencoder(...)), i 3 trainer reali qui sotto vengono costruiti uno
    # dopo l'altro senza un manual_seed() dedicato per ciascuno — l'init dei
    # pesi dipende dalla sequenza corrente del generatore RNG globale (seedato
    # una volta sola in main(), vedi torch.manual_seed(seed) più sopra nel
    # flusso). Funziona correttamente OGGI (verificato: i pesi del round 1
    # sono bit-identici tra experiment_20260724_111109.json [ε=1.0] e
    # experiment_20260724_144952.json [ε=0.1], stesso seed=42, stesso ordine
    # di operazioni prima di questo punto) ma è fragile: qualunque futura
    # modifica che inserisca un draw casuale in più prima di questo loop
    # desincronizzerebbe silenziosamente i pesi iniziali dei 3 siti, senza
    # sollevare errori. Non corretto qui (aggiungere un manual_seed() per
    # cluster cambierebbe i pesi iniziali rispetto a OGNI esperimento già
    # pubblicato, invalidando silenziosamente la comparabilità storica, senza
    # possibilità di verificarne l'effetto per esecuzione reale in questo
    # sandbox) — lasciato come nota per un refactor deliberato, non un
    # blind-fix.
    trainers: dict[str, AutoencoderTrainer] = {}
    for cid in cluster_ids:
        # Propaga il seed sperimentale nella config ml così AutoencoderTrainer
        # lo usa per il DataLoader generator → shuffle deterministico per seed.
        trainer_cfg = {**ml_cfg, "seed": exp_cfg.get("seed", 42)}
        trainers[cid] = AutoencoderTrainer(
            config=trainer_cfg,
            node_id=f"{cid}-01",
            cluster_id=cid,
        )
        logger.info(f"Cluster {cid}: {len(cluster_sessions[cid])} sessioni")

    gm = GradientManager({
        "epsilon":       exp_cfg["epsilon"],
        "delta":         exp_cfg["delta"],
        "max_grad_norm": exp_cfg["max_grad_norm"],
    })

    agg = FedAvgAggregator({"min_participants": len(cluster_ids)})

    # ML Plane (2026-07-22, richiesta esplicita — vedi README "Relation to
    # Prior Work" e src/ml/ml_plane.py per il contesto completo): PRIMA di
    # questo fix, AutoencoderTrainer/GradientManager/FedAvgAggregator
    # emettevano già eventi ML Plane reali (`emit_event()`), ma nessuno
    # chiamava mai `subscribe()` nella pipeline — gli eventi finivano nel
    # vuoto e `results[round_num]` sotto veniva costruito leggendo le
    # variabili Python locali `raw_updates`/`round_updates`/`aggregated`
    # direttamente, mai attraverso il ML Plane. Ora un singolo `MLPlane`
    # collega tutti e tre i componenti (`wire()`) a un `FLArtifactCollector`
    # reale, che raccoglie gli stessi oggetti (stessa identità, non copie)
    # nel punto in cui il paper QRS 2026 li colloca: "client updates are
    # temporarily available... before the execution of FedAvg". Da qui in
    # poi, `results[round_num]` viene popolato leggendo dal collector, non
    # dalle variabili locali — il ML Plane è quindi il meccanismo realmente
    # usato per raccogliere il flusso di artefatti FL, non un'osservazione
    # parallela mai consultata.
    mlplane   = MLPlane()
    collector = FLArtifactCollector()
    mlplane.subscribe(collector)
    mlplane.wire(*trainers.values(), gm, agg)

    results: dict[int, dict[str, Any]] = {}

    # ── Byzantine attack config ────────────────────────────────────────────────
    # Legge la sezione byzantine_attack dal config. Se assente o disabled, nessun attacco.
    _byz_cfg      = cfg.get("byzantine_attack", {})
    _byz_enabled  = _byz_cfg.get("enabled", False)
    # Default aggiornato 2026-07-22 (3 siti reali + 2 client sintetici per lo
    # sweep IDS): "highway" non esiste più come cluster_id — l'attaccante di
    # default è ora uno dei client sintetici (mai un sito reale, per non
    # implicare che un sito reale specifico sia "malevolo" nella narrazione).
    _byz_node     = _byz_cfg.get("byzantine_node", "synthetic_1")   # cluster attaccante
    _byz_type     = _byz_cfg.get("attack_type", "gradient_scaling")
    _byz_scale    = float(_byz_cfg.get("scale_factor", 10.0))

    if _byz_enabled and _byz_node not in cluster_ids:
        logger.error(
            f"[BYZANTINE ATTACK] byzantine_node={_byz_node!r} non è tra i cluster "
            f"attivi {cluster_ids} — l'attacco NON verrà applicato a nessun client "
            "(nessun errore verrà sollevato più avanti, il confronto cid==_byz_node "
            "semplicemente non scatterà mai mai). Verifica byzantine_node nel config."
        )

    if _byz_enabled:
        logger.warning(
            f"[BYZANTINE ATTACK] abilitato — nodo={_byz_node}, "
            f"tipo={_byz_type}, scale={_byz_scale}"
        )

    if no_dp:
        logger.warning(
            "[NO-DP BASELINE] Rumore Differential Privacy DISABILITATO (σ=0). "
            "Il modello si addestra senza privacy noise. "
            "Confronta AUC con esperimento DP per disambiguare: "
            "AUC>0.5 → Scenario A (DP sopprime MIA); "
            "AUC≈0.5 → Scenario B (modello non memorizza)."
        )

    # Baseline per IDS al round 1: pesi del modello inizializzato (prima del training).
    # Consente di calcolare il delta round 1 = post_training - init_model invece di
    # usare pesi assoluti (che causano GRADIENT_EXPLOSION falso per L2 >> max_grad_norm).
    _init_weights = trainers[cluster_ids[0]].get_weights() if cluster_ids else None
    results[0] = {"raw_global_weights": _init_weights}

    for round_num in range(1, fl_rounds + 1):
        logger.info(f"=== FL Round {round_num}/{fl_rounds} ===")

        for cid, trainer in trainers.items():
            # Fix 2026-07-22 (review B1): cattura i pesi del modello PRIMA del
            # training locale di questo round — sono il "modello ricevuto" da
            # usare come riferimento per il clipping del DELTA (non del vettore
            # assoluto) in privatize()/clip_only() più sotto. Per il round 1
            # coincide con l'init casuale del trainer (nessun modello globale
            # ancora applicato); per i round successivi coincide col modello
            # applicato da apply_global_model() alla fine del round precedente.
            pre_round_weights = trainer.get_weights()

            # Training locale
            update = trainer.train_local(cluster_sessions[cid], round_num)

            # ── Gradient scaling attack ──────────────────────────────────────
            # Se questo cluster è il nodo Byzantine e l'attacco è abilitato,
            # moltiplica tutti i pesi per scale_factor.
            # Effetto: l'update è geometricamente ~scale_factor× più distante
            # dagli altri nodi → Krum score >> 1.5 → alert reale.
            # L'attacco avviene sui pesi raw (pre-DP), che è ciò che l'IDS analizza;
            # poi passa anche nella privatizzazione → distorce FedAvg.
            if _byz_enabled and cid == _byz_node and _byz_type == "gradient_scaling":
                scaled_weights = [
                    (w if isinstance(w, torch.Tensor) else torch.tensor(float(w))) * _byz_scale
                    for w in (update.weights or [])
                ]
                from ml.base_ml import GradientUpdate as _GU
                update = _GU(
                    node_id=update.node_id,
                    cluster_id=update.cluster_id,
                    round_num=update.round_num,
                    weights=scaled_weights,
                    gradients=update.gradients,
                    loss=update.loss,
                    n_samples=update.n_samples,
                    metadata={**update.metadata, "byzantine_attack": True,
                               "scale_factor": _byz_scale},
                )
                logger.warning(
                    f"[BYZANTINE] Round {round_num}: {cid} — "
                    f"gradient scaling ×{_byz_scale} applicato"
                )
                # Ri-emissione ML Plane (2026-07-22, wiring reale — vedi
                # src/ml/ml_plane.py): train_local() ha già emesso un evento
                # purdue_level=1 con l'update ORIGINALE (pre-scaling) qualche
                # riga fa. Senza questa ri-emissione, FLArtifactCollector
                # continuerebbe a restituire da raw_updates() il peso non
                # scalato per questo nodo — sbagliato, perché ciò che
                # "attraversa davvero il confine" verso l'aggregatore (e verso
                # l'IDS) è la versione scalata. Semantica "ultimo vince" per
                # (round, node_id) nel collector gestisce correttamente questa
                # sovrascrittura.
                trainers[cid].emit_event(MLPlaneEvent(
                    event_type="gradient_upload",
                    purdue_level=1,
                    payload=update,
                    round_num=round_num,
                    metadata={**update.metadata, "byzantine_reemit": True},
                ))

            # Nota (2026-07-22, wiring ML Plane): non serve più appendere
            # `update` a una lista locale — train_local() (e, per il nodo
            # Byzantine, la ri-emissione sopra) ha già emesso l'evento
            # purdue_level=1 attraverso il ML Plane; sarà `collector.raw_updates()`,
            # letto sotto dopo il loop, a fornire la stessa informazione — ma
            # realmente "raccolta" dal ML Plane, non passata a mano.
            # Applica DP — passa le chiavi per escludere buffer BatchNorm dal rumore.
            # Con --no-dp, usa l'update raw direttamente (σ=0, nessun rumore).
            weight_keys = trainer.get_weight_keys()
            if no_dp:
                private_update = update  # baseline: nessuna privacy noise
            elif dp_mode == "central":
                # Central DP (2026-07-22): clip-only per client, NESSUN rumore
                # individuale — il rumore va sull'aggregato (vedi sotto, dopo
                # agg.aggregate()). Un attacco sul singolo update (LiRA) vedrà
                # quindi l'update clippato ma pulito — atteso, non un bug.
                # reference_weights=pre_round_weights (fix 2026-07-22, review
                # B1): clippa il DELTA rispetto al modello ricevuto, non il
                # vettore assoluto — vedi GradientManager._clip_weights().
                private_update = gm.clip_only(
                    update, weight_keys=weight_keys, reference_weights=pre_round_weights
                )
            else:
                # "dp-fedavg" (default) e "local" condividono lo stesso meccanismo
                # per-client (clip+noise prima dell'aggregazione) — la differenza
                # tra i due è SOLO nella visibilità di raw_updates per l'IDS
                # (vedi sotto, dopo il loop dei client).
                private_update = gm.privatize(
                    update, weight_keys=weight_keys, reference_weights=pre_round_weights
                )
            agg.collect(private_update)
            # Nota (2026-07-22): idem — `private_update` è già stato emesso da
            # gm.privatize()/clip_only() (o è l'update raw stesso, se no_dp)
            # attraverso il ML Plane; niente più lista locale `round_updates`.

        # ── Lettura dal ML Plane (2026-07-22, wiring reale) ──────────────────
        # Questi due elenchi sono ORA la vera fonte di raw_updates/updates per
        # `results[round_num]` sotto — non più le liste locali `raw_updates`/
        # `round_updates` costruite a mano durante il loop qui sopra (rimosse).
        # `collector` (src/ml/ml_plane.py) le ha popolate osservando gli eventi
        # realmente emessi da train_local() (+ ri-emissione Byzantine sopra) e
        # da privatize()/clip_only(), con semantica "ultimo vince" per
        # (round, node_id) — quindi riflettono correttamente l'eventuale
        # scalatura Byzantine senza bisogno di logica speciale qui.
        _collected_raw = collector.raw_updates(round_num)
        _collected_updates = collector.privatized_updates(round_num)

        # Calcola raw_global_weights: media semplice dei pesi pre-DP.
        # Usato da IDS come riferimento per i delta (evita che il rumore DP
        # del global aggregato inquini l'analisi degli update locali).
        raw_global_weights: list[Any] | None = None
        if _collected_raw and _collected_raw[0].weights:
            n_w    = len(_collected_raw[0].weights)
            total  = sum(u.n_samples for u in _collected_raw) or len(_collected_raw)
            raw_global_weights = []
            for i in range(n_w):
                wavg = sum(
                    (u.weights[i] if isinstance(u.weights[i], torch.Tensor)
                     else torch.tensor(float(u.weights[i])))
                    * (u.n_samples / total)
                    for u in _collected_raw
                )
                raw_global_weights.append(wavg)

        # FedAvg
        aggregated = agg.aggregate(round_num)

        if aggregated is None:
            logger.warning(f"Round {round_num} saltato — partecipanti insufficienti")
            continue

        # Central DP (2026-07-22): UN SOLO rumore Gaussiano sull'aggregato pulito,
        # dopo FedAvg — non per client (vedi clip_only() sopra). Il modello
        # rumorizzato è quello effettivamente distribuito ai trainer per il
        # prossimo round E quello salvato in results (usato da LiRA per il
        # warm-start degli shadow — deve vedere lo stesso rumore del target).
        if not no_dp and dp_mode == "central" and aggregated.global_weights:
            _agg_weight_keys = (
                trainers[cluster_ids[0]].get_weight_keys() if cluster_ids else None
            )
            aggregated.global_weights = gm.privatize_aggregate(
                aggregated.global_weights,
                weight_keys=_agg_weight_keys,
                n_participants=aggregated.n_participants or len(cluster_ids),
            )
            # Ri-emissione ML Plane (2026-07-22, wiring reale): agg.aggregate()
            # sopra ha già emesso un evento "aggregation" con l'aggregato
            # PULITO (pre-rumore). privatize_aggregate() modifica `aggregated`
            # in-place aggiungendo il rumore centrale — senza questa
            # ri-emissione, FLArtifactCollector.aggregation(round_num)
            # restituirebbe l'aggregato sbagliato (senza rumore), mentre è
            # quello rumorizzato che viene davvero distribuito ai trainer e
            # salvato in results per il warm-start degli shadow di LiRA.
            agg.emit_event(MLPlaneEvent(
                event_type="aggregation",
                purdue_level=3,
                payload=aggregated,
                round_num=round_num,
            ))

        # Distribuisci modello globale ai trainer
        for trainer in trainers.values():
            trainer.apply_global_model(aggregated)

        loss_str = f"{aggregated.mean_loss:.6f}" if aggregated.mean_loss is not None else "N/A"
        logger.info(f"Round {round_num} — loss globale: {loss_str}")

        # Local DP (2026-07-22): il server/IDS non deve MAI vedere l'update raw,
        # nemmeno transitoriamente — a differenza di dp-fedavg/central, dove un
        # server "honest-but-curious"/trusted vede comunque il raw update prima
        # di clippare+rumorizzare (dp-fedavg) o prima di aggregare (central).
        # Non salviamo raw_updates in questa modalità: run_ids() (vedi sotto,
        # "Preferisci raw_updates... Fallback su updates per retrocompatibilità")
        # userà automaticamente `updates` (già rumorizzati) — modellando
        # correttamente il fatto che sotto local DP non esiste nessun momento in
        # cui il server osserva il valore pulito.
        _store_raw = _collected_raw if not (dp_mode == "local" and not no_dp) else None
        _store_raw_global = (
            raw_global_weights if not (dp_mode == "local" and not no_dp) else None
        )

        results[round_num] = {
            "mean_loss":         aggregated.mean_loss,
            "n_participants":    aggregated.n_participants,
            "updates":           _collected_updates,  # privatized — usati da FedMIA — dal ML Plane
            "raw_updates":       _store_raw,          # pre-DP — usati da IDS (None sotto local DP) — dal ML Plane
            "raw_global_weights": _store_raw_global,  # media raw — riferimento IDS (idem)
            "global_weights":    aggregated.global_weights,
        }

    return results

# ── FedMIA Attack ──────────────────────────────────────────────────────────────

# Feature ACN usate per FedMIA — allineate con AutoencoderTrainer
_MIA_FEATURES = [
    "total_energy_kwh", "max_power_kw", "kwh_requested",
    "minutes_available", "hour_of_day", "duration_hours",
]


def run_fedmia(
    cfg: dict,
    members: list[dict[str, Any]],
    non_members: list[dict[str, Any]],
    fl_results: dict[int, dict[str, Any]],
) -> dict[int, dict[str, Any]]:
    """
    Loss-based Membership Inference Attack per FL con autoencoder.

    Per ogni round FL carica i pesi globali aggregati in un Autoencoder
    locale e misura l'errore di ricostruzione su membri e non-membri.
    Principio (Yeom et al., 2018): il modello FL produce errore basso sui
    campioni visti nel training (membri) e alto sui non visti (non-membri).
    La DP riduce questa gap → AUC-ROC → 0.5 (attacco non migliore del random).

    IMPORTANTE: members deve contenere sessioni effettivamente usate per il
    training FL; non_members deve essere un hold-out set mai visto da nessun
    FL node. Usare lo stesso pool per entrambi invalida la misura di AUC-ROC.

    L'AUC varia per round: round iniziali → modello non converge → AUC ≈ 0.5;
    round finali → modello memorizza → AUC cresce se DP insufficiente.

    Args:
        cfg:         configurazione esperimento
        members:     sessioni usate per FL training (vere member)
        non_members: sessioni hold-out mai viste durante training (vere non-member)
        fl_results:  dict round → {"global_weights": [...], ...}

    Returns:
        {round_num: {"auc_roc": float, "member_score_mean": float, ...}}
    """
    from sklearn.metrics import roc_auc_score

    logger.info(f"FedMIA — members: {len(members)}, non-members: {len(non_members)}")

    # ── Bilanciamento pool MIA ──────────────────────────────────────────────────
    # ACN-Data: 10,458 members vs 2,615 non-members (split 80/20).
    # Un pool sbilanciato 4:1 non invalida l'AUC-ROC (che è rank-based) ma produce
    # soglie di classificazione asimmetriche. Si usa un subsample fisso dei members
    # per garantire pool identici in dimensione e confrontabilità tra esperimenti.
    _pool_rng        = random.Random(cfg.get("experiment", {}).get("seed", 42))
    _n_bal           = min(len(members), len(non_members))
    members_balanced = _pool_rng.sample(members, _n_bal)
    logger.info(
        f"FedMIA pool bilanciato — members: {len(members_balanced)}, "
        f"non-members: {len(non_members)} "
        f"(members originali: {len(members)}, campionati con seed fisso)"
    )

    input_dim = cfg["ml"]["input_dim"]

    def _score_batch(model: Autoencoder, sess_list: list[dict]) -> list[float]:
        """Calcola membership scores (-MSE) sulle sessioni con feature complete.
        Coerente con _sessions_to_tensor: scarta sessioni con None nelle feature."""
        rows: list[list[float]] = []
        for s in sess_list:
            try:
                row = [float(s[f]) for f in _MIA_FEATURES]
                rows.append(row)
            except (KeyError, TypeError, ValueError):
                continue
        if not rows:
            return []
        tensor = torch.tensor(rows, dtype=torch.float32)
        results_: list[float] = []
        with torch.no_grad():
            for i in range(0, len(tensor), 256):
                batch = tensor[i : i + 256]
                recon  = model(batch)
                errors = torch.mean((recon - batch) ** 2, dim=1)
                # Score = -errore: basso errore → membro → score alto
                results_.extend(-e.item() for e in errors)
        return results_

    mia_results: dict[int, dict[str, Any]] = {}

    for round_num, round_data in sorted(fl_results.items()):
        global_weights = round_data.get("global_weights")
        if global_weights is None:
            logger.warning(f"Round {round_num}: global_weights assenti — skip FedMIA")
            continue

        # Carica pesi globali FL in un autoencoder locale (inference only).
        # load_state_dict trasferisce anche i buffer BatchNorm (running_mean/var).
        # global_weights è una lista con lo stesso ordine di state_dict().values():
        # sia AutoencoderTrainer.get_weights() che questo zip usano state_dict()
        # sulla stessa architettura Autoencoder, quindi l'ordine è garantito.
        model = Autoencoder(input_dim=input_dim)
        orig_state = model.state_dict()
        keys = list(orig_state.keys())
        if len(global_weights) != len(keys):
            logger.error(
                f"Round {round_num}: global_weights ha {len(global_weights)} elementi, "
                f"state_dict ne richiede {len(keys)} — skip FedMIA"
            )
            continue
        state = {
            k: (w if isinstance(w, torch.Tensor) else torch.tensor(w)).to(orig_state[k].dtype)
            for k, w in zip(keys, global_weights)
        }
        model.load_state_dict(state, strict=True)
        # Clamp BatchNorm running_var a valori positivi: il rumore DP con σ grande
        # (es. σ=48 per ε=0.1) può rendere running_var negativa, causando NaN in
        # sqrt(running_var + eps) durante la forward pass in eval mode.
        # Questo guard è difensivo; con il fix in GradientManager._add_noise() i
        # buffer BN non ricevono più rumore, quindi running_var sarà già positiva.
        for buf_name, buf in model.named_buffers():
            if "running_var" in buf_name:
                buf.clamp_(min=1e-8)
        model.eval()

        member_scores     = _score_batch(model, members_balanced)
        non_member_scores = _score_batch(model, non_members)

        if not member_scores or not non_member_scores:
            logger.warning(f"Round {round_num}: score batch vuoto — skip AUC")
            continue

        labels = [1] * len(member_scores) + [0] * len(non_member_scores)
        scores = member_scores + non_member_scores

        # Filtra NaN residui (belt-and-suspenders: senza il BN fix potrebbe ancora
        # verificarsi NaN per epsilon molto piccoli con sigma >> 1)
        scores_arr = np.array(scores)
        labels_arr = np.array(labels)
        valid_mask = ~np.isnan(scores_arr) & ~np.isinf(scores_arr)
        if valid_mask.sum() < 10:
            logger.warning(
                f"Round {round_num}: troppi score NaN/Inf "
                f"({(~valid_mask).sum()}/{len(scores_arr)}) — probabile corruzione pesi DP "
                f"(σ grande >> norma pesi). AUC impostato a 0.5 (baseline random)."
            )
            mia_results[round_num] = {
                "auc_roc":               0.5,
                "member_score_mean":     float("nan"),
                "non_member_score_mean": float("nan"),
                "nan_fraction":          float((~valid_mask).sum()) / len(scores_arr),
            }
            continue
        if not valid_mask.all():
            logger.warning(
                f"Round {round_num}: {(~valid_mask).sum()} score NaN/Inf filtrati "
                f"su {len(scores_arr)} totali"
            )
        labels_arr = labels_arr[valid_mask]
        scores_arr = scores_arr[valid_mask]
        auc = roc_auc_score(labels_arr, scores_arr)
        logger.info(f"Round {round_num} — FedMIA AUC-ROC: {auc:.4f}")

        mia_results[round_num] = {
            "auc_roc":               auc,
            "member_score_mean":     float(np.nanmean(member_scores)),
            "non_member_score_mean": float(np.nanmean(non_member_scores)),
        }

    return mia_results


# ── Shadow Model MIA Attack ────────────────────────────────────────────────────

def run_fedmia_shadow(
    cfg: dict,
    train_sessions: list[dict[str, Any]],
    holdout_sessions: list[dict[str, Any]],
    fl_results: dict[int, dict[str, Any]],
) -> dict[int, dict[str, Any]]:
    """
    Calibrated Shadow-Model MIA Attack (ispirato a LiRA, Carlini et al. 2022).

    Motivazione:
        L'attacco Yeom (loss-based) fallisce quando il modello FL generalizza bene:
        la loss su membro e non-membro è simile → AUC ≈ 0.5. Il shadow attack
        controlla per la "difficoltà intrinseca" di ogni campione confrontando
        il modello FL target con un modello shadow addestrato su dati simili ma
        diversi: il segnale di membership emerge dalla *differenza* di loss, non
        dal valore assoluto.

    Setup:
        - shadow_train (50% di train_sessions) → addestra shadow model (no FL, no DP)
        - eval_members (50% di train_sessions) → valutazione membro:
            * FL ha visto questi campioni → loss_target bassa
            * shadow NON li ha visti     → loss_shadow alta
            * score = loss_shadow - loss_target > 0  ✓
        - holdout_sessions → valutazione non-membro:
            * né FL né shadow li hanno visti
            * score = loss_shadow - loss_target ≈ 0  ✓

    Score di membership:
        score(x) = MSE(shadow_model, x) − MSE(target_model, x)
        Maggiore score → più probabile che x sia stato nel training FL.

    Riferimenti:
        Carlini et al., "Membership Inference Attacks From First Principles",
        IEEE S&P 2022. https://arxiv.org/abs/2112.03570

    Args:
        cfg:              configurazione esperimento
        train_sessions:   sessioni usate nel FL training (tutti i membri)
        holdout_sessions: sessioni mai viste durante FL (non-membri)
        fl_results:       dict round → {"global_weights": [...], ...}

    Returns:
        {round_num: {
            "shadow_auc_roc": float,
            "shadow_member_score_mean": float,
            "shadow_non_member_score_mean": float,
            "shadow_score_gap": float,   # differenza media membro - non-membro
            "n_eval_members": int,
            "n_non_members": int,
        }}
    """
    from sklearn.metrics import roc_auc_score

    seed       = cfg.get("experiment", {}).get("seed", 42)
    input_dim  = cfg["ml"]["input_dim"]
    ml_cfg     = cfg["ml"]
    local_epochs = ml_cfg.get("epochs", 3)
    total_rounds = cfg.get("experiment", {}).get("fl_rounds", 100)

    # ── Step 1: split train → shadow_train (50%) + eval_members (50%) ──────────
    rng = random.Random(seed + 999)       # seed diverso da quello del pool Yeom
    shuffled = list(train_sessions)
    rng.shuffle(shuffled)
    mid           = max(1, len(shuffled) // 2)
    shadow_train  = shuffled[:mid]
    eval_members  = shuffled[mid:]

    logger.info(
        f"Shadow MIA — shadow_train: {len(shadow_train)}, "
        f"eval_members: {len(eval_members)}, non-members: {len(holdout_sessions)}"
    )

    # ── Step 2: addestra il shadow model ──────────────────────────────────────
    # Autoencoder locale (non FL, no DP) addestrato sul shadow_train.
    # Epoche totali = local_epochs × total_rounds (equivalente al training FL),
    # capped a 500 per non rallentare troppo il sweep.
    shadow_epochs = min(local_epochs * total_rounds, 500)

    def _build_tensor(sess_list: list[dict]) -> torch.Tensor | None:
        rows = []
        for s in sess_list:
            try:
                rows.append([float(s[f]) for f in _MIA_FEATURES])
            except (KeyError, TypeError, ValueError):
                continue
        return torch.tensor(rows, dtype=torch.float32) if rows else None

    shadow_tensor = _build_tensor(shadow_train)
    if shadow_tensor is None or len(shadow_tensor) == 0:
        logger.warning("Shadow MIA: shadow_train vuoto dopo feature extraction — skip")
        return {}

    # Fix (2026-07-24, review indipendente — stessa classe di bug già corretta in
    # run_lira() il 2026-07-21d): seed PRIMA di istanziare il modello. Autoencoder()
    # pesca l'init casuale dei pesi dall'RNG globale di torch — se il seed viene
    # fissato solo qui sotto (dopo la costruzione), l'init dipende da qualunque cosa
    # abbia consumato l'RNG globale prima (draw di rumore DP, init di altri modelli,
    # ecc.), rendendo shadow_auc_roc/shadow_score_gap non riproducibili "a parità di
    # seed" e — più grave — confondendo il confronto no-DP vs DP: il path no-DP non
    # fa mai draw di rumore prima di questo punto, il path DP sì, quindi i due rami
    # partivano da pesi iniziali diversi anche a seed identico.
    torch.manual_seed(seed + 999)
    shadow_model = Autoencoder(input_dim=input_dim)
    shadow_optimizer = torch.optim.Adam(shadow_model.parameters(), lr=ml_cfg.get("lr", 1e-3))
    shadow_criterion = torch.nn.MSELoss()
    batch_size = ml_cfg.get("batch_size", 32)

    if len(shadow_tensor) < batch_size:
        logger.warning(
            f"Shadow MIA: shadow_train ({len(shadow_tensor)} sessioni) < "
            f"batch_size ({batch_size}) — shadow model non addestrato con drop_last=True. "
            "shadow_auc_roc non significativo. Aumentare il dataset o ridurre batch_size."
        )
        return {}

    shadow_model.train()
    torch.manual_seed(seed + 999)
    shadow_ds = torch.utils.data.TensorDataset(shadow_tensor)
    shadow_loader_gen = torch.Generator()
    shadow_loader_gen.manual_seed(seed + 999)
    shadow_loader = torch.utils.data.DataLoader(
        shadow_ds, batch_size=batch_size, shuffle=True,
        drop_last=True, generator=shadow_loader_gen,
    )

    for epoch in range(shadow_epochs):
        for (batch,) in shadow_loader:
            shadow_optimizer.zero_grad()
            recon = shadow_model(batch)
            loss  = shadow_criterion(recon, batch)
            loss.backward()
            shadow_optimizer.step()

    shadow_model.eval()
    logger.info(f"Shadow model addestrato — {shadow_epochs} epoche su {len(shadow_train)} sessioni")

    # ── Step 3: compute scores per-sample con entrambi i modelli ───────────────

    def _mse_batch(model: Autoencoder, sess_list: list[dict]) -> list[float]:
        """Calcola MSE per sessione (non negato: valore grezzo per calibrazione)."""
        tensor = _build_tensor(sess_list)
        if tensor is None:
            return []
        scores: list[float] = []
        with torch.no_grad():
            for i in range(0, len(tensor), 256):
                batch = tensor[i : i + 256]
                recon  = model(batch)
                errors = torch.mean((recon - batch) ** 2, dim=1)
                scores.extend(e.item() for e in errors)
        return scores

    # Pre-computa shadow scores una volta sola (non dipende dal round FL)
    shadow_scores_members     = _mse_batch(shadow_model, eval_members)
    shadow_scores_nonmembers  = _mse_batch(shadow_model, holdout_sessions)

    # Bilanciamento: stessa dimensione per eval_members e non-members
    _bal_rng       = random.Random(seed + 999)
    _n_bal         = min(len(shadow_scores_members), len(shadow_scores_nonmembers))
    shadow_scores_members    = _bal_rng.sample(shadow_scores_members, _n_bal)
    shadow_scores_nonmembers = _bal_rng.sample(shadow_scores_nonmembers, _n_bal)

    # ── Step 4: per ogni round FL, calcola score calibrato ─────────────────────
    shadow_results: dict[int, dict[str, Any]] = {}

    for round_num, round_data in sorted(fl_results.items()):
        global_weights = round_data.get("global_weights")
        if global_weights is None:
            continue

        # Carica pesi globali FL nel target model
        target_model = Autoencoder(input_dim=input_dim)
        orig_state   = target_model.state_dict()
        keys         = list(orig_state.keys())
        if len(global_weights) != len(keys):
            logger.error(
                f"Shadow MIA round {round_num}: global_weights {len(global_weights)} "
                f"!= state_dict {len(keys)} — skip"
            )
            continue
        state = {
            k: (w if isinstance(w, torch.Tensor) else torch.tensor(w)).to(orig_state[k].dtype)
            for k, w in zip(keys, global_weights)
        }
        target_model.load_state_dict(state, strict=True)
        for buf_name, buf in target_model.named_buffers():
            if "running_var" in buf_name:
                buf.clamp_(min=1e-8)
        target_model.eval()

        # Scores target model
        target_scores_members    = _mse_batch(target_model, eval_members)
        target_scores_nonmembers = _mse_batch(target_model, holdout_sessions)

        # Bilancia anche i target scores allo stesso indice del shadow
        _bal_rng2 = random.Random(seed + 999)
        target_scores_members    = _bal_rng2.sample(target_scores_members, _n_bal)
        target_scores_nonmembers = _bal_rng2.sample(target_scores_nonmembers, _n_bal)

        # score calibrato = loss_shadow − loss_target
        # Positivo → target conosce il campione meglio del shadow → membro
        calibrated_members    = [
            s - t for s, t in zip(shadow_scores_members,    target_scores_members)
        ]
        calibrated_nonmembers = [
            s - t for s, t in zip(shadow_scores_nonmembers, target_scores_nonmembers)
        ]

        labels = [1] * len(calibrated_members) + [0] * len(calibrated_nonmembers)
        scores = calibrated_members + calibrated_nonmembers

        scores_arr = np.array(scores)
        labels_arr = np.array(labels)
        valid_mask = ~np.isnan(scores_arr) & ~np.isinf(scores_arr)
        if valid_mask.sum() < 10:
            logger.warning(f"Shadow MIA round {round_num}: troppi NaN — skip")
            continue
        scores_arr = scores_arr[valid_mask]
        labels_arr = labels_arr[valid_mask]

        try:
            auc = roc_auc_score(labels_arr, scores_arr)
        except ValueError:
            auc = 0.5

        score_gap = float(np.nanmean(calibrated_members) - np.nanmean(calibrated_nonmembers))
        logger.info(
            f"Round {round_num} — Shadow MIA AUC: {auc:.4f} "
            f"(gap={score_gap:.6f})"
        )

        shadow_results[round_num] = {
            "shadow_auc_roc":               round(auc, 6),
            "shadow_member_score_mean":     round(float(np.nanmean(calibrated_members)), 6),
            "shadow_non_member_score_mean": round(float(np.nanmean(calibrated_nonmembers)), 6),
            "shadow_score_gap":             round(score_gap, 6),
            "n_eval_members":               _n_bal,
            "n_non_members":                _n_bal,
        }

    return shadow_results


# ── LiRA Attack (Carlini et al. 2022) ──────────────────────────────────────────

def run_lira(
    cfg: dict,
    train_sessions: list[dict[str, Any]],
    holdout_sessions: list[dict[str, Any]],
    fl_results: dict[int, dict[str, Any]],
    n_shadow: int = 8,
    shadow_epochs_cap: int | None = None,
    no_dp: bool = False,
    dp_mode: str = "dp-fedavg",
    cluster_membership: dict[str, list[int]] | None = None,
    composed_output: dict[str, Any] | None = None,
) -> dict[int, dict[str, Any]]:
    """
    LiRA — Likelihood Ratio Attack, server-side, on each client's per-round update
    as actually submitted for aggregation (Carlini et al., IEEE S&P 2022).

    Threat model: semi-honest aggregator receives each client's local model update
    and runs MIA before aggregating them. This is stronger than attacking the
    global model, because FedAvg averaging destroys per-cluster memorisation that
    is still present in the individual client updates.

    Attack flow:
        1. Reconstruct per-cluster session assignment (deterministic — must match
           run_fl_rounds() to identify which sessions each client trained on).
        2. Fix, once, the IN/OUT sample assignment for n_shadow PER-CLUSTER shadow
           models (attacker's fixed auxiliary knowledge: which subset of that
           cluster's sessions each shadow "would have seen").
        3. For EACH ROUND, retrain every shadow model warm-started from that
           round's real global weights (the same starting point real clients use)
           for local_epochs epochs on its fixed IN subset, then — if DP is enabled
           — apply the SAME clip+noise privatisation a real client's update
           receives (see "Fix — DP must be observable" below).
        4. For each round and each client's submitted update:
           a. Load client's update weights.
           b. Compute target_loss = MSE(client_model, x) for every eval sample.
           c. Split THAT CLIENT'S CLUSTER shadow losses (this round) into IN
              (shadow saw x) and OUT (didn't see x).
           d. score(x) = log P(loss | IN dist) − log P(loss | OUT dist)  [Gaussian log-LR]
        5. Pool scores across all clients per round → AUC-ROC.

    Fix — shadow/target distribution mismatch (2026-07-21a):
        Shadow models were trained on random 50% subsets drawn from ALL
        train_sessions (all 4 clusters mixed), while the attacked model is a
        per-cluster specialist trained only on its own cluster's ~2600 sessions.
        This violates LiRA's core assumption (shadow ≈ target's training
        distribution) and produced a systematically INVERTED score (lira_auc_roc
        as low as 0.14–0.32, lira_non_member_score_mean saturating near +20).
        Fix: one shadow ensemble PER CLUSTER, sampled only from that cluster's
        own index range.

    Fix — shadow/target TRAINING PROCEDURE mismatch (2026-07-21b):
        The per-cluster fix alone was NOT sufficient: round 1 improved
        (0.558→0.75) but rounds 2–10 stayed broken and flat (~0.26), confirmed on
        10/10 runs across nodp-sweep1 + dp-sweep1. Root cause: from round 2
        onward, a real client does NOT train from scratch — it starts from the
        previous round's shared GLOBAL weights and does only local_epochs (50)
        epochs of local fine-tuning. The shadow ensemble, however, was trained
        ONCE, from random init, for a fixed 250-epoch budget — a completely
        different trajectory. Round 1 (no shared init yet) had no such mismatch,
        which is why it alone looked healthy.
        Fix: shadows are now retrained EVERY ROUND, warm-started from
        fl_results[round-1]["global_weights"] (round 1 uses random init, matching
        what real clients do), then fine-tuned for exactly local_epochs epochs —
        mirroring the real per-round client procedure. The IN/OUT sample
        assignment per shadow stays fixed across rounds (fixed auxiliary
        knowledge); only the model weights are retrained each round.

    Fix — DP must be observable by the attack it's supposed to defend against
    (2026-07-21c):
        LiRA previously attacked `raw_updates`, captured in run_fl_rounds()
        BEFORE gm.privatize() is called. Since privatize() always returns a new
        object (never mutates its input), raw_updates NEVER carries DP noise,
        regardless of --no-dp / --epsilon. Verified empirically: lira_auc_roc at
        round 1 was bit-for-bit identical between nodp-sweep1 and dp-sweep1 for
        matching seeds (e.g. seed=42 → 0.750451 in both). By construction, DP
        could never suppress this attack — which defeats the purpose of the
        no-DP vs DP comparison (the goal is to measure how much DP degrades each
        attack's strength).
        Fix: LiRA now attacks `updates` (the actual per-client update submitted
        for aggregation — post-privatize when DP is enabled, identical to raw
        when --no-dp). Shadows are privatised with the same GradientManager
        clip+noise procedure when DP is enabled, so the IN/OUT calibration
        reflects the same noise regime the target went through — otherwise a
        noisy target would be miscalibrated against clean shadows, which is a
        mismatch of its own kind.

    Non-member handling:
        Non-members are NEVER in any shadow model's training set → in_losses is
        always empty. We fall back to the PER-CLUSTER, PER-ROUND global IN
        distribution (pooled from that cluster's members' IN shadow losses this
        round) as the IN reference. Non-members' target_loss is typically high
        (model never saw them), so log_p_in << log_p_out → negative score → correct.

    Fix — asymmetric IN/OUT variance estimation for non-members (2026-07-21e):
        Discovered from real sweep data (nodp-sweep2, seeds 42/123, fixes a-d
        already applied): lira_non_member_score_mean spiked to +17.9 at round 2
        (clip ceiling is ±20) then decayed slowly toward +1.4 by round 10, while
        lira_member_score_mean stayed flat near 0 throughout — an AUC inversion
        (0.15-0.27 at rounds 2-4) that is NOT explained by fixes a-d.
        Root cause: for a genuine non-member, out_losses is the per-sample,
        cross-shadow spread over just n_shadow (~8) values on that ONE point.
        Right after a shared warm-start (round >= 2), all shadows for a cluster
        start from the identical global_weights and have only briefly diverged,
        so on a point none of them trained on they produce nearly identical
        reconstructions → σ_out collapses toward the μ×0.05 floor. Meanwhile the
        IN side for a non-member uses the pooled, cluster-wide
        global_in_stats_per_cluster fallback (hundreds of samples, naturally
        much wider σ). Comparing a near-collapsed per-point σ_out against a wide
        pooled σ_in makes the Gaussian log-likelihood-ratio dominated by the
        1/σ² term rather than by genuine membership signal. Member scoring does
        NOT show this because both its IN and OUT sides are already
        per-sample/small-N (symmetric, so the collapse — when it happens —
        affects both terms similarly and partly cancels).
        Fix: compute an analogous pooled, per-cluster/per-round GLOBAL OUT
        distribution (global_out_stats_per_cluster — same construction as
        global_in_stats_per_cluster, pooling the complementary OUT
        observations) and use it as a floor for σ_out — and, symmetrically, for
        σ_in — so neither side's variance can collapse below what is typically
        observed across the whole cluster this round. μ_in/μ_out (the actual
        discriminative signal) are untouched by this fix; only the
        variance floor changes.
        Contributing factor: n_shadow=8 ("fast demo" per config/experiment.yaml)
        makes any per-sample cross-shadow variance estimate inherently noisy;
        the config itself recommends 16-32 for "paper quality" — increasing
        n_shadow would also reduce reliance on this floor.

    Why this differs from run_fedmia / run_fedmia_shadow:
        Both previous attacks use the GLOBAL aggregated model, which is itself a
        cross-cluster blend — so a cross-cluster, one-shot shadow ensemble is the
        correct reference for those two (no mismatch there; these fixes only
        apply to LiRA). LiRA uses each CLIENT's individual update — the signal
        before FedAvg averaging — which is why its shadows must mirror both the
        per-cluster data AND the per-round training procedure of that client.

    Args:
        cfg:              experiment configuration dict
        train_sessions:    sessions used in FL training (members)
        holdout_sessions:  hold-out sessions never seen by FL (non-members)
        fl_results:        per-round FL data, must contain "updates" and
                           "global_weights" (previous round used as warm-start)
        n_shadow:          number of shadow models PER CLUSTER (8 = fast demo;
                           ≥32 = paper quality)
        shadow_epochs_cap: override for shadow training epochs (default:
                           local_epochs, matching the real client). Use a small
                           value (e.g. 20) only for smoke tests.
        no_dp:             must match the flag used for this experiment — controls
                           whether shadows are privatised like real clients.
        dp_mode:           must match the dp_mode used in run_fl_rounds() for this
                           experiment (2026-07-22) — controls HOW shadows are
                           privatised to mirror the target's actual noise regime:
                             "dp-fedavg"/"local" → gm.privatize() per shadow (clip
                               + noise), matching what a real client submitted.
                             "central" → gm.clip_only() per shadow (clip, NO
                               noise) — matching that under central DP, individual
                               client updates are never noised, only the aggregate
                               is. LiRA on individual updates should therefore show
                               NO suppression under central DP, at any ε — this is
                               the expected empirical result the CS4 comparison
                               (docs/CaseStudies.md §2.4.3) is designed to surface,
                               not a bug to fix.
        cluster_membership: {cluster_id: [indici GLOBALI in train_sessions]}
                           (2026-07-22) — deve rispecchiare ESATTAMENTE i cluster
                           usati da run_fl_rounds() per lo stesso esperimento
                           (es. via group_sessions_by_site() per i 3 siti reali,
                           o inject_synthetic_clients() per i 5 client dello
                           sweep IDS). Se None (default, retrocompatibile coi
                           test con dati sintetici senza site_id reale), affetta
                           train_sessions in 4 parti contigue uguali — stesso
                           comportamento storico pre-2026-07-22.

    Returns:
        {round_num: {
            "lira_auc_roc":               float,
            "lira_member_score_mean":     float,
            "lira_non_member_score_mean": float,
            "lira_score_gap":             float,
            "n_shadow":                   int,
        }}
    """
    from sklearn.metrics import roc_auc_score

    from ml.base_ml import GradientUpdate as _GU

    seed         = cfg.get("experiment", {}).get("seed", 42)
    input_dim    = cfg["ml"]["input_dim"]
    ml_cfg       = cfg["ml"]
    exp_cfg      = cfg.get("experiment", {})
    lr           = ml_cfg.get("lr", 1e-3)
    batch_size   = ml_cfg.get("batch_size", 32)
    local_epochs = ml_cfg.get("epochs", 50)
    # Epoche di training per gli shadow ad ogni round: di default = local_epochs,
    # esattamente come i client reali (fix 2026-07-21b). shadow_epochs_cap permette
    # di ridurle per gli smoke test (es. 20) senza alterare il regime dei run reali.
    shadow_epochs = shadow_epochs_cap if shadow_epochs_cap is not None else local_epochs

    # GradientManager per privatizzare gli shadow ESATTAMENTE come i client reali
    # (stesso clipping + stesso meccanismo di rumore) — fix 2026-07-21c: senza
    # questo, un target rumoroso (DP on) verrebbe calibrato contro shadow puliti,
    # un mismatch che si aggiungerebbe a quelli già corretti sopra.
    gm = GradientManager({
        "epsilon":       exp_cfg.get("epsilon", 1.0),
        "delta":         exp_cfg.get("delta", 1e-5),
        "max_grad_norm": exp_cfg.get("max_grad_norm", 1.0),
    })

    # ── Step 1: Reconstruct per-cluster membership — must match run_fl_rounds() ─
    # FIX 2026-07-22 (coerente col fix di run_nvflare_mia.py/group_sessions_by_site()):
    # cluster_idx_pools ora è {cluster_id: [indici GLOBALI in train_sessions]} —
    # se cluster_membership è fornito (3 siti reali o 5 per lo sweep IDS), viene
    # usato direttamente; altrimenti fallback storico (4 fette contigue fittizie,
    # per compatibilità coi test con dati sintetici senza site_id).
    if cluster_membership is not None:
        _CLUSTER_IDS = list(cluster_membership.keys())
        cluster_idx_pools: dict[str, list[int]] = cluster_membership
    else:
        _CLUSTER_IDS = ["highway", "urban", "residential", "corporate"]
        cluster_size = max(1, len(train_sessions) // len(_CLUSTER_IDS))
        cluster_idx_pools = {}
        for i, cid in enumerate(_CLUSTER_IDS):
            start = i * cluster_size
            end   = len(train_sessions) if i == len(_CLUSTER_IDS) - 1 else start + cluster_size
            cluster_idx_pools[cid] = list(range(start, end))

    cluster_members: dict[str, list[dict[str, Any]]] = {
        cid: [train_sessions[i] for i in idxs] for cid, idxs in cluster_idx_pools.items()
    }

    # Reverse map: sample Python id → cluster (for correct client-member matching).
    # Members must be evaluated ONLY against their home cluster's client.
    # Cross-cluster evaluation (highway sample vs urban client) gives high target_loss
    # → looks like a non-member → contaminates member pool with false negatives
    # → AUC drops below 0.5 when 3/4 of evaluations are cross-cluster (4 clusters).
    _sample_to_cluster: dict[int, str] = {
        id(s): cid
        for cid, sessions in cluster_members.items()
        for s in sessions
    }

    # Fix 2026-08-11: simmetrico a _sample_to_cluster ma per i NON-member
    # (holdout_sessions). Trovato investigando un'anomalia reale: il nodp-sweep1
    # in corso mostrava lira_score_gap CRESCENTE (separazione IN/OUT reale e
    # forte) ma AUC-ROC pooled piatto vicino al caso — la firma di un pool
    # eterogeneo, non di segnale assente. Causa: il guard cross-cluster sopra
    # (poche righe più in basso, "if is_member: ... skip mismatched pairs") è
    # applicato SOLO ai member — ogni non-member viene invece valutato contro
    # TUTTI i client/cluster senza restrizione, quindi entra fino a 3 volte in
    # round_nonmember_scores, mescolando client con scale di loss potenzialmente
    # diverse. Sotto DP con clipping per-client (central/dp-fedavg/local) tutti
    # i client sono forzati in una scala di delta comparabile (max_grad_norm),
    # quindi la mescolanza è quasi inerte — ma sotto --no-dp, senza alcun vincolo
    # di norma, il client più piccolo (office1, ~1341 sessioni contro
    # ~25-27k di caltech/jpl) può divergere di scala in modo scorrelato
    # round-per-round, e quel rumore extra nel pool diluisce un ranking che a
    # livello per-cluster è in realtà ben calibrato. Fix: ogni non-member ha
    # comunque un site_id reale proprio (proviene da una site FL prima dello
    # split, non "non è di nessun sito") — usiamo la stessa risoluzione
    # _SITE_ID_TO_NAME di group_indices_by_site() per restringere anche i
    # non-member al proprio cluster di origine, esattamente come i member.
    # Guard applicato SOLO quando cluster_membership è reale (3+ siti reali/
    # sweep IDS n=5) — quando è None (fallback storico a 4 fette fittizie,
    # usato dai test con dati sintetici senza site_id reale), holdout_sessions
    # non è mai stato partizionato in quelle 4 fette fittizie: risolvere
    # site_id qui darebbe "unknown" per ogni non-member, mai uguale a una delle
    # 4 fette fittizie lato client → guard sempre vero → round_nonmember_scores
    # vuoto per ogni round → stesso genere di regressione silenziosa che questo
    # fix vuole eliminare, ma sui test invece che sui dati reali. Dict vuoto in
    # quel caso preserva esattamente il comportamento pre-fix (nessuna restrizione
    # sui non-member), identico a prima per tutta la suite di test esistente.
    _holdout_sample_to_cluster: dict[int, str] = {
        id(s): _SITE_ID_TO_NAME.get(s.get("site_id", ""), s.get("site_id", "") or "unknown")
        for s in holdout_sessions
    } if cluster_membership is not None else {}

    # Balanced eval pool: subsample members to match hold-out size
    _pool_rng      = random.Random(seed + 31415)
    _n_bal         = min(len(train_sessions), len(holdout_sessions))
    members_bal    = _pool_rng.sample(train_sessions, _n_bal)
    nonmembers_bal = _pool_rng.sample(holdout_sessions, min(_n_bal, len(holdout_sessions)))

    logger.info(
        f"LiRA — n_shadow={n_shadow}/cluster, shadow_epochs={shadow_epochs}/round, "
        f"eval pool: {len(members_bal)} members, {len(nonmembers_bal)} non-members"
    )

    # Concatenate into a single ordered list: members first, then non-members.
    # Index j < len(members_bal) → member; j >= len(members_bal) → non-member.
    eval_samples = members_bal + nonmembers_bal
    n_eval       = len(eval_samples)

    # Accumulatore opzionale per l'attacco "composto" multi-round (2026-08-12,
    # vedi ComposedLiRAAttack in src/plugins/attacks/lira.py): somma il
    # log-likelihood-ratio dello STESSO campione su tutti i round invece di
    # mediare 10 AUC-ROC indipendenti — combinazione Neyman-Pearson ottima per
    # evidenza indipendente ripetuta, pensata per sfruttare il vero budget ε
    # COMPOSTO (vedi "epsilon_cumulative_naive" in config) invece del singolo
    # ε per round che ogni round-AUC indipendente vede. Piggyback sulla STESSA
    # retraining degli shadow già fatta da questa funzione, per non raddoppiare
    # il costo computazionale con una funzione separata. Attivo SOLO se il
    # chiamante passa composed_output (non None) — default None, quindi zero
    # impatto sul comportamento/output esistente per ogni chiamante attuale.
    _cumulative_scores: dict[int, float] = {}
    _sample_is_member:  dict[int, bool]  = {id(s): True for s in members_bal}
    _sample_is_member.update({id(s): False for s in nonmembers_bal})

    # Map each eval member to its index in train_sessions (for IN/OUT tracking).
    # Members were sampled from train_sessions without copying → id() is valid.
    _train_idx: dict[int, int] = {id(s): i for i, s in enumerate(train_sessions)}

    def _build_tensor(sess_list: list[dict]) -> torch.Tensor | None:
        rows = []
        for s in sess_list:
            try:
                rows.append([float(s[f]) for f in _MIA_FEATURES])
            except (KeyError, TypeError, ValueError):
                continue
        return torch.tensor(rows, dtype=torch.float32) if rows else None

    def _load_weights_into(model: Autoencoder, weights: list) -> bool:
        """Carica una lista di pesi (stesso ordine di state_dict().values()) in
        `model`. Restituisce False se le shape non combaciano (nessuna eccezione)."""
        orig_state = model.state_dict()
        keys = list(orig_state.keys())
        if len(weights) != len(keys):
            return False
        state = {
            k: (w if isinstance(w, torch.Tensor) else torch.tensor(w)).to(orig_state[k].dtype)
            for k, w in zip(keys, weights)
        }
        model.load_state_dict(state, strict=True)
        for buf_name, buf in model.named_buffers():
            if "running_var" in buf_name:
                buf.clamp_(min=1e-8)
        return True

    # ── Step 2: IN/OUT sample assignment per shadow — FISSO tra i round ────────
    # Rappresenta la conoscenza ausiliaria fissa dell'attaccante (stesso subset di
    # sessioni per ogni cluster/shadow). Solo i PESI dello shadow vengono
    # riaddestrati ogni round (Step 3), non il subset di campioni.
    # Offset grande e primo per cluster, per garantire seed distinti tra cluster senza
    # usare hash(str) (non deterministico tra processi Python — PYTHONHASHSEED random).
    _CLUSTER_SEED_OFFSET = 104729
    shadow_in_idx_sets_per_cluster: dict[str, list[set[int]]] = {}
    shadow_tensors_per_cluster: dict[str, list[torch.Tensor | None]] = {}

    for cluster_idx, cid in enumerate(_CLUSTER_IDS):
        cluster_idx_pool = cluster_idx_pools[cid]
        cluster_in_idx_sets: list[set[int]] = []
        cluster_tensors: list[torch.Tensor | None] = []

        for shadow_idx in range(n_shadow):
            _s = seed + cluster_idx * _CLUSTER_SEED_OFFSET + shadow_idx * 31337
            shadow_rng = random.Random(_s)
            n_in       = max(batch_size + 1, len(cluster_idx_pool) // 2)
            n_in       = min(n_in, len(cluster_idx_pool))
            in_indices = set(shadow_rng.sample(cluster_idx_pool, n_in))
            cluster_in_idx_sets.append(in_indices)

            in_sessions   = [train_sessions[i] for i in sorted(in_indices)]
            shadow_tensor = _build_tensor(in_sessions)
            if shadow_tensor is None or len(shadow_tensor) < batch_size:
                logger.warning(
                    f"LiRA[{cid}] shadow {shadow_idx}: {len(in_sessions)} sessioni < "
                    f"batch_size={batch_size} — shadow skippato in ogni round"
                )
            cluster_tensors.append(shadow_tensor)

        shadow_in_idx_sets_per_cluster[cid] = cluster_in_idx_sets
        shadow_tensors_per_cluster[cid]     = cluster_tensors

    # ── Step 3 & 4: per round, riaddestra gli shadow (warm-start) e valuta i client ─
    lira_results: dict[int, dict[str, Any]] = {}

    for round_num, round_data in sorted(
        (item for item in fl_results.items() if item[0] > 0), key=lambda x: x[0]
    ):
        # Fix 2026-07-21c: attacca "updates" (post-privatize, ciò che viene davvero
        # sottoposto ad aggregazione) invece di "raw_updates" (pre-DP per costruzione).
        client_updates = round_data.get("updates", [])
        if not client_updates:
            logger.warning(f"LiRA round {round_num}: nessun update — skip")
            continue

        # Warm-start per gli shadow di QUESTO round: stesso punto di partenza usato
        # dai client reali per il training locale del round (fix 2026-07-21b).
        # Round 1 → init casuale (nessun round precedente, come i client reali).
        _warm_start = (
            fl_results.get(round_num - 1, {}).get("global_weights")
            if round_num > 1 else None
        )

        shadow_mse_matrix_per_cluster: dict[str, list[list[float | None]]] = {}

        for cluster_idx, cid in enumerate(_CLUSTER_IDS):
            in_sets  = shadow_in_idx_sets_per_cluster[cid]
            tensors  = shadow_tensors_per_cluster[cid]
            cluster_mse_matrix: list[list[float | None]] = []

            for shadow_idx, (in_indices, shadow_tensor) in enumerate(zip(in_sets, tensors)):
                if shadow_tensor is None:
                    cluster_mse_matrix.append([None] * n_eval)
                    continue

                _s = (
                    seed + round_num * 1_000_003
                    + cluster_idx * _CLUSTER_SEED_OFFSET + shadow_idx * 31337
                )

                # Fix (review indipendente 2026-07-21d): seed PRIMA di istanziare il
                # modello — altrimenti, quando _warm_start è None (round 1), l'init
                # casuale di Autoencoder() dipende dallo stato ambientale del RNG
                # globale di torch invece che da _s, rendendo il round 1 non
                # riproducibile in isolamento.
                torch.manual_seed(_s)
                shadow_model = Autoencoder(input_dim=input_dim)
                if _warm_start is not None and not _load_weights_into(shadow_model, _warm_start):
                    logger.warning(
                        f"LiRA[{cid}] round {round_num} shadow {shadow_idx}: "
                        "warm-start non applicabile (shape mismatch) — init casuale"
                    )

                # Fix 2026-07-22 (review B1): cattura i pesi PRIMA del training
                # dello shadow — riferimento per il clip del DELTA più sotto,
                # simmetrico a pre_round_weights in run_fl_rounds(). Coincide
                # col warm-start se applicato con successo, altrimenti con
                # l'init casuale appena fatto (round 1 / shape mismatch).
                _shadow_pretrain_weights = [
                    w.detach().clone() for w in shadow_model.state_dict().values()
                ]

                shadow_opt  = torch.optim.Adam(shadow_model.parameters(), lr=lr)
                shadow_crit = torch.nn.MSELoss()
                shadow_ds  = torch.utils.data.TensorDataset(shadow_tensor)
                shadow_gen = torch.Generator()
                shadow_gen.manual_seed(_s)
                shadow_loader = torch.utils.data.DataLoader(
                    shadow_ds, batch_size=batch_size, shuffle=True,
                    drop_last=True, generator=shadow_gen,
                )

                shadow_model.train()
                for _ in range(shadow_epochs):
                    for (batch,) in shadow_loader:
                        shadow_opt.zero_grad()
                        recon = shadow_model(batch)
                        loss  = shadow_crit(recon, batch)
                        loss.backward()
                        shadow_opt.step()
                shadow_model.eval()

                # Fix 2026-07-21c: privatizza lo shadow ESATTAMENTE come un client
                # reale, cosi' la calibrazione IN/OUT riflette il vero effetto della
                # DP quando abilitata (altrimenti un target rumoroso verrebbe
                # confrontato con shadow puliti — mismatch aggiuntivo).
                # Fix 2026-07-22: sotto "central" DP, il client NON rumorizza il
                # proprio update (solo clip — il rumore va sull'aggregato, mai
                # osservabile da LiRA che attacca il singolo update) — lo shadow
                # deve rispecchiare la STESSA cosa, non gm.privatize() completo.
                if not no_dp:
                    _shadow_keys = list(shadow_model.state_dict().keys())
                    _shadow_update = _GU(
                        node_id=f"shadow-{cid}-{shadow_idx}",
                        cluster_id=cid,
                        round_num=round_num,
                        weights=[w.detach().clone() for w in shadow_model.state_dict().values()],
                        gradients=None,
                        loss=None,
                        n_samples=len(shadow_tensor),
                        metadata={},
                    )
                    # reference_weights=_shadow_pretrain_weights (fix 2026-07-22,
                    # review B1): clippa il DELTA dello shadow rispetto al proprio
                    # punto di partenza, come per i client reali in run_fl_rounds().
                    if dp_mode == "central":
                        _privatized = gm.clip_only(
                            _shadow_update, weight_keys=_shadow_keys,
                            reference_weights=_shadow_pretrain_weights,
                        )
                    else:
                        _privatized = gm.privatize(
                            _shadow_update, weight_keys=_shadow_keys,
                            reference_weights=_shadow_pretrain_weights,
                        )
                    _load_weights_into(shadow_model, _privatized.weights)
                    shadow_model.eval()

                mse_row: list[float | None] = []
                for sample in eval_samples:
                    try:
                        row    = [float(sample[f]) for f in _MIA_FEATURES]
                        tensor = torch.tensor([row], dtype=torch.float32)
                        with torch.no_grad():
                            recon = shadow_model(tensor)
                            mse_row.append(float(torch.mean((recon - tensor) ** 2).item()))
                    except (KeyError, TypeError, ValueError):
                        mse_row.append(None)
                cluster_mse_matrix.append(mse_row)

            shadow_mse_matrix_per_cluster[cid] = cluster_mse_matrix

        logger.info(
            f"LiRA round {round_num}: {n_shadow}×{len(_CLUSTER_IDS)} shadow "
            f"riaddestrati ({shadow_epochs} epoche/shadow, "
            f"warm_start={'sì' if _warm_start is not None else 'no (init casuale)'}, "
            f"dp_su_shadow={'no (--no-dp)' if no_dp else f'sì (dp_mode={dp_mode})'})"
        )

        # Per-cluster, per-round global IN distribution — fallback per i non-membri
        # (mai IN in nessuno shadow). Ricalcolata ogni round perché gli shadow
        # sono stati appena riaddestrati.
        global_in_stats_per_cluster: dict[str, tuple[float, float]] = {}
        for cid in _CLUSTER_IDS:
            in_sets    = shadow_in_idx_sets_per_cluster[cid]
            mse_matrix = shadow_mse_matrix_per_cluster.get(cid, [])
            pooled: list[float] = []
            for j in range(len(members_bal)):
                train_idx = _train_idx.get(id(eval_samples[j]))
                if train_idx is None:
                    continue
                for si, in_set in enumerate(in_sets):
                    if train_idx in in_set and si < len(mse_matrix):
                        mse = mse_matrix[si][j]
                        if mse is not None:
                            pooled.append(mse)

            mu    = float(np.mean(pooled)) if pooled else 0.05
            sigma = max(float(np.std(pooled)), max(mu * 0.05, 1e-4)) if pooled else 0.01
            global_in_stats_per_cluster[cid] = (mu, sigma)

        # Fix — asymmetric IN/OUT variance for non-members (2026-07-21e):
        # For a genuine non-member, out_losses is built from n_shadow (~8) MSE
        # values on THAT ONE sample — a per-sample, cross-shadow spread. Right
        # after a shared warm-start (round ≥ 2), all shadows for a cluster begin
        # from the identical global_weights and have only briefly diverged, so
        # on a point none of them trained on they tend to produce nearly
        # identical reconstructions → σ_out collapses toward the μ×0.05 floor
        # (observed: as low as ~1e-4). Meanwhile the IN side for a non-member
        # uses the pooled, cluster-wide global_in_stats_per_cluster fallback
        # above (hundreds of samples, naturally much wider σ). Comparing a
        # near-collapsed per-point σ_out against a wide pooled σ_in makes the
        # Gaussian log-likelihood-ratio dominated by the 1/σ² term rather than
        # by genuine membership signal, producing non-member scores that spike
        # to the clip ceiling (empirically +17.9 at round 2, seed 42,
        # nodp-sweep2) while member scores stay flat near 0 (member scoring is
        # symmetric: both its IN and OUT sides are already per-sample/small-N).
        # Fix: compute an analogous pooled, per-cluster/per-round GLOBAL OUT
        # distribution (same construction as global_in_stats_per_cluster, just
        # pooling the complementary — OUT — observations) and use it as a
        # floor for σ_out (and, symmetrically, for σ_in) so neither side's
        # variance can collapse below what is typically observed across the
        # whole cluster this round. This does not touch μ_in/μ_out (the
        # discriminative signal itself) — only prevents an under-estimated
        # per-sample σ from artificially amplifying the ratio.
        global_out_stats_per_cluster: dict[str, tuple[float, float]] = {}
        for cid in _CLUSTER_IDS:
            in_sets    = shadow_in_idx_sets_per_cluster[cid]
            mse_matrix = shadow_mse_matrix_per_cluster.get(cid, [])
            pooled_out: list[float] = []
            for j in range(n_eval):
                is_mem = j < len(members_bal)
                _sample = eval_samples[j]
                # Fix 2026-08-12: il fix 2026-08-11 (guard cross-cluster
                # simmetrico) era applicato solo nel loop di scoring per-sample
                # (righe ~1731-1742), non qui — questo pool "floor" continuava a
                # mescolare member/non-member di QUALSIASI sito nella statistica
                # OUT di OGNI cluster, reintroducendo un livello più in profondità
                # la stessa contaminazione cross-cluster che il fix precedente
                # doveva eliminare (un member/non-member fuori sito valutato
                # contro shadow di un cluster che non ha mai visto quel tipo di
                # dato dà una loss innaturalmente alta, gonfiando σ_out_fb in modo
                # diverso per cluster in base a quanta contaminazione riceve —
                # inerte sotto clipping, che allinea le scale tra client, ma reale
                # sotto --no-dp). Stesso guard usato sotto, con la stessa mappa
                # _sample_to_cluster/_holdout_sample_to_cluster.
                _home_cluster = (
                    _sample_to_cluster.get(id(_sample)) if is_mem
                    else _holdout_sample_to_cluster.get(id(_sample))
                )
                if _home_cluster is not None and _home_cluster != cid:
                    continue
                train_idx = _train_idx.get(id(_sample)) if is_mem else None
                for si, in_set in enumerate(in_sets):
                    if si >= len(mse_matrix):
                        continue
                    mse = mse_matrix[si][j]
                    if mse is None:
                        continue
                    if is_mem and train_idx is not None and train_idx in in_set:
                        continue  # this (shadow, sample) pair is an IN observation, skip
                    pooled_out.append(mse)

            mu_o    = float(np.mean(pooled_out)) if pooled_out else 0.05
            sigma_o = max(float(np.std(pooled_out)), max(mu_o * 0.05, 1e-4)) if pooled_out else 0.01
            global_out_stats_per_cluster[cid] = (mu_o, sigma_o)

        round_member_scores:    list[float] = []
        round_nonmember_scores: list[float] = []

        for update in client_updates:
            if update is None or not update.weights:
                continue

            # Load client's submitted update (post-privatize when DP enabled).
            client_model = Autoencoder(input_dim=input_dim)
            if not _load_weights_into(client_model, update.weights):
                logger.warning(
                    f"LiRA round {round_num} {update.cluster_id}: "
                    f"weights shape mismatch — skip client"
                )
                continue
            client_model.eval()

            # Fix: usa l'ensemble shadow del cluster di QUESTO client — non un ensemble
            # cross-cluster globale — per calibrare IN/OUT sotto lo stesso regime di
            # training del modello attaccato (vedi docstring "shadow/target mismatch").
            _client_cluster_id = getattr(update, "cluster_id", None)
            _cluster_shadow_mse     = shadow_mse_matrix_per_cluster.get(_client_cluster_id, [])
            _cluster_shadow_in_sets = shadow_in_idx_sets_per_cluster.get(_client_cluster_id, [])
            _cluster_mu_in_fb, _cluster_sigma_in_fb = global_in_stats_per_cluster.get(
                _client_cluster_id, (0.05, 0.01)
            )
            # Fix 2026-07-21e: pooled OUT floor — see comment above global_out_stats_per_cluster.
            _cluster_mu_out_fb, _cluster_sigma_out_fb = global_out_stats_per_cluster.get(
                _client_cluster_id, (0.05, 0.01)
            )
            if not _cluster_shadow_mse:
                logger.warning(
                    f"LiRA round {round_num} {_client_cluster_id}: nessun ensemble shadow "
                    "per questo cluster — skip client"
                )
                continue

            for j, sample in enumerate(eval_samples):
                try:
                    row         = [float(sample[f]) for f in _MIA_FEATURES]
                    tensor      = torch.tensor([row], dtype=torch.float32)
                    with torch.no_grad():
                        recon       = client_model(tensor)
                        target_loss = float(torch.mean((recon - tensor) ** 2).item())
                except (KeyError, TypeError, ValueError):
                    continue

                is_member       = (j < len(members_bal))
                sample_train_idx = _train_idx.get(id(sample)) if is_member else None

                # Cross-cluster guard: only evaluate a member against its home cluster's client.
                # A highway sample vs an urban client always gives high target_loss (never seen)
                # → looks like a non-member → 3/4 of member evaluations are false negatives
                # → AUC < 0.5 even without DP. Skip mismatched (sample, client) pairs.
                if is_member:
                    _member_cluster = _sample_to_cluster.get(id(sample))
                    _client_cluster = getattr(update, "cluster_id", None)
                    if _member_cluster is not None and _member_cluster != _client_cluster:
                        continue
                else:
                    # Fix 2026-08-11: guard simmetrico — vedi commento su
                    # _holdout_sample_to_cluster sopra. Senza questo, ogni
                    # non-member entra nel pool fino a una volta per cluster
                    # (3x), mescolando client con scale di loss potenzialmente
                    # diverse — inerte sotto clipping per-client, ma diluisce
                    # l'AUC pooled sotto --no-dp dove non c'è alcun vincolo di
                    # norma a mantenere le scale dei client comparabili.
                    _nonmember_cluster = _holdout_sample_to_cluster.get(id(sample))
                    _client_cluster = getattr(update, "cluster_id", None)
                    if _nonmember_cluster is not None and _nonmember_cluster != _client_cluster:
                        continue

                # Split shadow losses: IN = shadows (di QUESTO cluster, QUESTO round)
                # che hanno visto il campione; OUT = resto.
                in_losses:  list[float] = []
                out_losses: list[float] = []
                for si, in_set in enumerate(_cluster_shadow_in_sets):
                    if si >= len(_cluster_shadow_mse):
                        continue
                    mse = _cluster_shadow_mse[si][j]
                    if mse is None:
                        continue
                    if is_member and sample_train_idx is not None and sample_train_idx in in_set:
                        in_losses.append(mse)
                    else:
                        out_losses.append(mse)  # non-members always go here

                if len(out_losses) < 2:
                    continue  # insufficient calibration data

                μ_out = float(np.mean(out_losses))
                # σ floor (fix 2026-07-21e): the per-sample cross-shadow std alone can
                # collapse to ~0 when all n_shadow (~8) shadows agree on this one point
                # (typically right after a shared warm-start, before they've diverged —
                # see comment on global_out_stats_per_cluster above). Floor it with the
                # larger of: the μ×0.05 scale-adaptive floor (original guard, prevents
                # 1e-8 collapse), and the pooled per-cluster/per-round OUT spread
                # (reflects the actual point-to-point + shadow-to-shadow variability
                # observed this round, not just this one sample's near-zero spread).
                σ_out = max(
                    float(np.std(out_losses)),
                    max(μ_out * 0.05, 1e-4),
                    _cluster_sigma_out_fb,
                )

                # Use per-sample IN distribution if available; fall back to this
                # client's cluster-specific global IN stats (not a cross-cluster global).
                # Same σ floor rationale as σ_out above — applied symmetrically so
                # neither side's variance can be artificially tighter than the other
                # purely due to small-N (n_shadow≈8) per-sample estimation noise.
                if len(in_losses) >= 2:
                    μ_in = float(np.mean(in_losses))
                    σ_in = max(
                        float(np.std(in_losses)),
                        max(μ_in * 0.05, 1e-4),
                        _cluster_sigma_in_fb,
                    )
                else:
                    μ_in = _cluster_mu_in_fb
                    σ_in = _cluster_sigma_in_fb

                # Gaussian log-likelihood ratio (Carlini 2022, Eq. 2):
                # score > 0 → loss matches IN distribution → member
                log_p_in  = (-0.5 * ((target_loss - μ_in)  / σ_in)  ** 2) - np.log(σ_in)
                log_p_out = (-0.5 * ((target_loss - μ_out) / σ_out) ** 2) - np.log(σ_out)
                # Clip to ±20: log-LR beyond this range has no practical discriminative
                # value and only amplifies numerical instabilities in edge cases.
                lira_score = float(np.clip(log_p_in - log_p_out, -20.0, 20.0))

                if np.isnan(lira_score) or np.isinf(lira_score):
                    continue

                if is_member:
                    round_member_scores.append(lira_score)
                else:
                    round_nonmember_scores.append(lira_score)

                if composed_output is not None:
                    _sid = id(sample)
                    _cumulative_scores[_sid] = _cumulative_scores.get(_sid, 0.0) + lira_score

        if not round_member_scores or not round_nonmember_scores:
            logger.warning(
                f"LiRA round {round_num}: score pool vuoto "
                f"(m={len(round_member_scores)}, nm={len(round_nonmember_scores)}) — skip"
            )
            continue

        labels = [1] * len(round_member_scores) + [0] * len(round_nonmember_scores)
        scores = round_member_scores + round_nonmember_scores

        try:
            auc = roc_auc_score(labels, scores)
        except ValueError:
            auc = 0.5

        score_gap = float(np.mean(round_member_scores) - np.mean(round_nonmember_scores))
        logger.info(
            f"Round {round_num} — LiRA AUC: {auc:.4f} "
            f"(gap={score_gap:.6f}, n_shadow={n_shadow})"
        )

        lira_results[round_num] = {
            "lira_auc_roc":               round(auc, 6),
            "lira_member_score_mean":     round(float(np.mean(round_member_scores)),    6),
            "lira_non_member_score_mean": round(float(np.mean(round_nonmember_scores)), 6),
            "lira_score_gap":             round(score_gap, 6),
            "n_shadow":                   n_shadow,
        }

    if composed_output is not None:
        if _cumulative_scores:
            _labels = [1 if _sample_is_member[_sid] else 0 for _sid in _cumulative_scores]
            _scores = list(_cumulative_scores.values())
            try:
                _composed_auc = roc_auc_score(_labels, _scores)
            except ValueError:
                _composed_auc = 0.5
            _member_vals    = [s for _sid, s in _cumulative_scores.items() if _sample_is_member[_sid]]
            _nonmember_vals = [s for _sid, s in _cumulative_scores.items() if not _sample_is_member[_sid]]
            composed_output["composed_lira_auc_roc"] = round(float(_composed_auc), 6)
            composed_output["composed_lira_score_gap"] = round(
                float(np.mean(_member_vals) - np.mean(_nonmember_vals))
                if _member_vals and _nonmember_vals else 0.0,
                6,
            )
            composed_output["n_samples_scored"]   = len(_cumulative_scores)
            composed_output["n_rounds_aggregated"] = len(lira_results)
            logger.info(
                f"LiRA composto (multi-round, evidenza sommata su "
                f"{len(lira_results)} round) — AUC-ROC: {_composed_auc:.4f} "
                f"(gap={composed_output['composed_lira_score_gap']:.6f}, "
                f"n_samples={len(_cumulative_scores)})"
            )
        else:
            composed_output["composed_lira_auc_roc"] = None
            logger.warning("LiRA composto: nessuno score accumulato — pool vuoto")

    return lira_results


# ── Dispatch pluggable degli attacchi (src/plugins/attacks/) ───────────────────

def run_registered_attacks(
    cfg: dict,
    train_sessions: list[dict[str, Any]],
    holdout_sessions: list[dict[str, Any]],
    fl_results: dict[int, dict[str, Any]],
    **attack_kwargs: Any,
) -> dict[int, dict[str, Any]]:
    """
    Esegue tutti gli attacchi in ATTACK_REGISTRY (src/plugins/attacks/) e fonde
    i risultati per round in un unico dict.

    Fix 2026-07-24 — implementazione reale della "pluggable attack interface"
    descritta in docs/DSN2027_Positioning.md e docs/DeveloperGuide.md (prima
    di questo fix quei documenti descrivevano un meccanismo di plugin/registro
    che non esisteva nel codice — src/plugins/attacks/ conteneva solo un file
    inutilizzato, senza classe base né registro; vedi le correzioni datate
    2026-07-24 in entrambi i documenti per la storia completa).

    Sostituisce due copie quasi-identiche della stessa logica di dispatch che
    esistevano prima: questa in main() sotto, e un'altra in
    scripts/run_nvflare_mia.py::main() (analisi post-hoc sui dump NVFLARE).
    Entrambe ora chiamano questa funzione — stesso comportamento garantito
    da un solo punto di verità, non due copie che potevano divergere.

    Aggiungere un nuovo attacco: registralo in
    src/plugins/attacks/__init__.py::ATTACK_REGISTRY — non serve toccare
    questa funzione né i suoi due call site.

    Comportamento preservato IDENTICO alla versione pre-refactor (chiamate
    dirette a run_fedmia()/run_fedmia_shadow()/run_lira()): stesso ordine di
    esecuzione (yeom, shadow, lira), stessa policy in caso di eccezione
    (logga e continua con gli altri, non solleva — un attacco fallito non
    deve impedire il salvataggio degli altri risultati), stessa logica di
    merge per round (round assente → assegna il dict; round presente →
    update in-place, i.e. shadow/lira aggiungono campi a un round già
    popolato da yeom senza sovrascriverlo). Zero cambio di comportamento
    numerico: ogni classe wrapper in src/plugins/attacks/ chiama la funzione
    esistente, invariata, con gli stessi argomenti (vedi src/core/base_attack.py
    per il razionale — queste funzioni hanno anni di fix empirici dietro,
    non riscritte qui).

    Args:
        attack_kwargs: passati a TUTTI gli attacchi registrati; ciascuna
            classe legge solo le chiavi che le servono (vedi
            src/plugins/attacks/lira.py per l'esempio — n_shadow,
            shadow_epochs_cap, no_dp, dp_mode, cluster_membership sono usati
            solo da LiRA, ignorati da Yeom/Shadow).
    """
    from plugins.attacks import ATTACK_REGISTRY

    # Fix 2026-07-24 (review indipendente, stesso giorno): questo loop iterava
    # una tupla hardcoded ("yeom", "shadow", "lira") invece di ATTACK_REGISTRY
    # stesso — la docstring sopra e il commento in
    # src/plugins/attacks/__init__.py promettono entrambi che registrare un
    # nuovo attacco lì basta, "non serve toccare questa funzione". Con la
    # tupla hardcoded era falso: un quarto attacco registrato non sarebbe mai
    # stato eseguito, in silenzio. Iterare ATTACK_REGISTRY.items() mantiene lo
    # stesso ordine di oggi (yeom, shadow, lira — un dict Python preserva
    # l'ordine di inserimento, e __init__.py li registra in quest'ordine) e
    # rende la promessa vera per davvero.
    mia_results: dict[int, dict[str, Any]] = {}
    for attack_name, attack_cls in ATTACK_REGISTRY.items():
        attack = attack_cls()
        try:
            attack_results = attack.run(
                cfg, train_sessions, holdout_sessions, fl_results, **attack_kwargs
            )
        except Exception as exc:  # noqa: BLE001
            logger.error(
                f"{attack_cls.__name__} ({attack_name}) fallito: {exc}. "
                "I risultati degli altri attacchi (e FL) vengono comunque salvati. "
                "Controllare il log per la causa (es. NaN negli score MIA).",
                exc_info=True,
            )
            continue
        for rnd, data in attack_results.items():
            if rnd in mia_results:
                mia_results[rnd].update(data)
            else:
                mia_results[rnd] = data
    return mia_results


# ── IDS Evaluation ─────────────────────────────────────────────────────────────

def run_ids(
    cfg: dict,
    fl_results: dict[int, dict[str, Any]],
    no_dp: bool = False,
) -> dict[int, dict[str, Any]]:
    """
    Valuta ChargingIDS su ogni round FL.

    Usa PrivacyAuditor per generare AuditReport reali con threats_detected
    popolato (GRADIENT_EXPLOSION, PRIVACY_BUDGET_EXHAUSTED, ecc.).
    Un singolo auditor persiste tra i round per tracciare l'epsilon cumulativo.

    Fix IDS (Sprint 9, 2026-07-16):
      1. GRADIENT_EXPLOSION — normalizzazione peer-relative:
         Con 50 epoch di training locale, la norma L2 delle differenze di peso
         supera sempre la soglia assoluta (max_grad_norm + 3σ ≈ 15.5).
         Fix: normalizza ogni delta per la norma mediana dei peer nello stesso round,
         portando la mediana = max_grad_norm. In questo modo solo gli outlier
         statistici (incl. Byzantine ×10) superano la soglia.
      2. Krum false positive — soglia calibrata per 50 epoch:
         Con 50 epoch, la varianza naturale inter-cluster produce score Krum fino a
         3.3 anche senza attacco. L'attacco Byzantine (scale_factor=10) dà score ≈4.0.
         Fix: soglia = 3.5 (Byzantine ≥ 4.0 > soglia > max FP osservato 3.27).
      3. Budget esaurito con no_dp=True — falso allarme:
         Quando DP è disabilitato (no_dp=True), il budget non viene consumato.
         Fix: passa epsilon=1000.0 all'auditor per eliminare i BUDGET_EXHAUSTED alert.

    Nota — degradazione attesa sotto dp_mode="local" (2026-07-22):
        Sotto vera local DP, run_fl_rounds() non salva raw_updates/
        raw_global_weights in fl_results (il server non deve mai vedere il
        valore pulito, nemmeno transitoriamente — vedi run_fl_rounds()
        docstring). Questa funzione usa già `round_data.get("raw_updates") or
        round_data.get("updates", [])`, quindi sotto local DP userà
        automaticamente gli update rumorizzati; ma senza un raw_global_weights
        di riferimento, il calcolo del delta peer-relative (fix 1 sopra) degrada
        a confronto sui pesi ASSOLUTI (rumorizzati) invece che su un delta
        round-su-round, il che può produrre più falsi GRADIENT_EXPLOSION. Questo
        NON è un bug da correggere: è la conseguenza reale e attesa del fatto che
        un server che non vede mai il gradiente pulito ha una difesa IDS
        strutturalmente più debole — un punto di discussione legittimo per il
        paper (motiva secure aggregation o central DP come alternative più
        IDS-compatibili quando serve sia privacy sia intrusion detection).
    """
    config_path = str(PROJECT_ROOT / "config" / "auditor.yaml")
    max_grad_norm = cfg["experiment"]["max_grad_norm"]

    # byzantine_tolerance/krum_threshold (2026-07-22, 3 siti reali + sweep IDS n=5):
    # PRIMA di questo fix, byzantine_tolerance era SEMPRE 0, anche durante lo
    # sweep dedicato con byzantine_attack.enabled=true (n=5 = 3 client reali +
    # 2 sintetici synthetic_1/synthetic_2, vedi inject_synthetic_client_indices()
    # in main()). Questo era sbagliato: la garanzia teorica di Krum di rilevare
    # f nodi Byzantine richiede n≥2f+3 — i 2 sintetici sono stati aggiunti
    # ESATTAMENTE per soddisfare n=5≥2·1+3 con f=1, quindi Krum va calcolato
    # con byzantine_tolerance=1 in quel caso (neighbors=n-f-2=2), non con f=0
    # (che userebbe neighbors=n-f-2=3 e non offrirebbe alcuna garanzia formale
    # per f=1). Nell'esperimento principale (n=3, solo client reali, mai
    # attaccato) byzantine_tolerance resta 0: con solo 3 nodi Krum funge da
    # trimmed-mean (esclude il più isolato) senza garanzia formale, il che è
    # accettabile perché quel run non inietta mai un attacco.
    _byz_cfg = cfg.get("byzantine_attack", {})
    _byz_enabled = _byz_cfg.get("enabled", False)
    _byz_tolerance = 1 if _byz_enabled else 0

    # FIX 2026-07-22 (review indipendente pre-push): byzantine_tolerance sopra
    # è derivato SOLO dal flag di config, non dal numero di client realmente
    # presenti in fl_results. Questo è un problema concreto per
    # scripts/run_nvflare_mia.py: i 2 client sintetici esistono SOLO nella
    # simulazione (mai portati su NVFLARE, per scelta esplicita — vedi
    # docs/NVFlareIntegration.md), quindi un dump NVFLARE reale ha SEMPRE e
    # SOLO 3 client anche se byzantine_attack.enabled=true è rimasto nel
    # config usato per l'analisi post-hoc. Con byzantine_tolerance=1 e solo 3
    # client, KrumDetector.compute_scores() (src/ids/charging_ids.py) applica
    # il guard n<2f+3 (3<5) e azzera silenziosamente TUTTI gli score (con un
    # solo warning nei log, facile da perdere) — non un crash, ma un IDS che
    # sembra funzionare e in realtà non rileva nulla. Cross-check: conta i
    # node_id distinti effettivamente presenti in fl_results e, se
    # byzantine_tolerance=1 non è supportato dal numero reale di client
    # (n<2·1+3=5), ricade su 0 con un warning esplicito — più sicuro che
    # lasciare che il guard di Krum lo faccia silenziosamente più a valle.
    if _byz_tolerance > 0:
        _observed_node_ids: set[str] = set()
        for _round_data in fl_results.values():
            for _upd in (_round_data or {}).get("raw_updates") or (_round_data or {}).get("updates", []):
                _nid = getattr(_upd, "node_id", None)
                if _nid is not None:
                    _observed_node_ids.add(_nid)
        _n_observed = len(_observed_node_ids)
        if _n_observed and _n_observed < 2 * _byz_tolerance + 3:
            logger.warning(
                f"[IDS] byzantine_attack.enabled=true ma solo {_n_observed} client "
                f"osservati in fl_results ({sorted(_observed_node_ids)}) — insufficienti "
                f"per la garanzia Krum n≥2f+3 con f={_byz_tolerance} (serve n≥{2*_byz_tolerance+3}). "
                "Probabile analisi di un run NVFLARE reale (dove i client sintetici non "
                "esistono, vedi docs/NVFlareIntegration.md) con un config di simulazione "
                "lasciato a byzantine_attack.enabled=true. Ricado su byzantine_tolerance=0 "
                "per evitare che Krum azzeri silenziosamente tutti gli score."
            )
            _byz_tolerance = 0

    # krum_threshold=3.5: calibrato empiricamente (2026-07-16) su n=4 "cluster"
    # fittizi — che in realtà erano 4 fette CONTIGUE dello STESSO singolo sito
    # reale (varianza inter-client dovuta solo a rumore di campionamento, MAI
    # a eterogeneità reale tra siti). Con 50 epoch, varianza naturale → score
    # Krum fino a ~3.3 (FP osservato). Attacco Byzantine ×10 → score ≈4.0.
    # Soglia 3.5: rileva Byzantine, non FP. Precedente 1.5 era calibrato per
    # 3 epoch (varianza bassa, score legittimi ≤1.1).
    #
    # ATTENZIONE — NON ANCORA RI-VALIDATA per n=5 (2026-07-22): lo sweep IDS
    # con n=5 usa 3 siti reali GENUINAMENTE eterogenei (Caltech/JPL/Office1 —
    # EVSE count, popolazione, pattern di ricarica diversi) più 2 client
    # sintetici che sono ri-affettature del pool COMBINATO (quindi vicini alla
    # "media" globale). Questo cambia la natura della varianza naturale rispetto
    # al vecchio caso (4 fette identiche dello stesso sito): un sito reale
    # potrebbe ora apparire geometricamente isolato rispetto ai 2 sintetici
    # anche SENZA alcun attacco, producendo falsi positivi Byzantine su un
    # client legittimo — oppure, al contrario, la soglia 3.5 potrebbe restare
    # valida se la varianza inter-sito è comunque dominata dal training (50
    # epoch). Non è stato possibile verificarlo in questa sessione (training
    # reale richiede torch/nvflare, non disponibili in questo sandbox).
    # PRIMA di fidarsi degli alert Krum di un run byzantine_attack.enabled=true,
    # eseguire lo sweep una volta con scale_factor=10 e ispezionare gli score
    # Krum reali in experiments/.../ids_audit_results (o l'export NVFLARE) per
    # i 3 client reali SENZA attacco attivo (o nei round prima che l'attaccante
    # venga scelto) — se il loro score naturale supera già ~3.0-3.5, alzare
    # krum_threshold di conseguenza prima di considerare i risultati attendibili.
    krum_threshold = 3.5

    ids = ChargingIDS(
        config_path=config_path,
        byzantine_tolerance=_byz_tolerance,
        cosine_threshold=0.3,
        krum_threshold=krum_threshold,
    )
    # Fix 3a: budget — se no_dp=True, epsilon enorme → budget_ratio resta ~0 (no BUDGET_EXHAUSTED).
    # Fix 3b: explosion threshold — con epsilon=1000, sigma≈0.005 → threshold≈1.015.
    #   Dopo peer-relative normalisation, median_norm=max_grad_norm=1.0 → ~50% client sopra 1.015
    #   → GRADIENT_EXPLOSION falsi in OGNI round della baseline no-DP.
    #   Con float("inf") il check è disabilitato: sensato perché senza DP non c'è sigma di rumore
    #   su cui basare la soglia; il Byzantine check rimane affidato a Krum.
    _auditor_epsilon    = 1000.0       if no_dp else cfg["experiment"]["epsilon"]
    _explosion_thresh   = float("inf") if no_dp else None   # None → formula Gaussian 3-sigma
    auditor = PrivacyAuditor(
        config_path=config_path,
        epsilon=_auditor_epsilon,
        explosion_threshold=_explosion_thresh,
    )

    ids_results: dict[int, dict[str, Any]] = {}

    # IDS usa pesi PRE-DP (raw_updates) e raw_global_weights come baseline.
    # Motivazione: con ε=0.1, σ ≈ 48×max_grad_norm. I pesi post-DP hanno
    # L2-norm >> max_grad_norm (rumore domina), causando GRADIENT_EXPLOSION
    # e BUDGET_EXHAUSTED falsi sistematici in ogni round.
    # In un sistema reale, il server/IDS vede gli update raw dai client PRIMA
    # che il rumore DP venga applicato → analisi corretta delle anomalie.
    # Delta = raw_local - raw_global_prev: rappresenta la deriva locale netta,
    # bounded dalla clipping norm × local epochs × lr (in pratica << max_grad_norm).

    # Inizializza prev_raw_global con i pesi del modello iniziale (round 0),
    # salvati in run_fl_rounds() prima dell'inizio del training loop.
    # Questo elimina il falso GRADIENT_EXPLOSION al round 1 dovuto ai pesi
    # assoluti del modello non ancora aggiornato.
    prev_raw_global: list[Any] | None = (fl_results.get(0) or {}).get("raw_global_weights")

    for round_num, round_data in sorted(
        (item for item in fl_results.items() if item[0] > 0), key=lambda x: x[0]
    ):
        # Preferisci raw_updates (pre-DP). Fallback su updates per retrocompatibilità.
        updates = round_data.get("raw_updates") or round_data.get("updates", [])
        # raw_global_weights di questo round (media raw, usato come prev al prossimo)
        current_raw_global = round_data.get("raw_global_weights")

        if not updates:
            prev_raw_global = current_raw_global
            ids_results[round_num] = {
                "alerts": [], "byzantine_detected": False, "drift_detected": False,
            }
            continue

        # ── Pass 1: calcola delta e norme L2 per tutti i client del round ────────
        # Necessario per la normalizzazione peer-relative (Fix 1 GRADIENT_EXPLOSION).
        _client_deltas: dict[str, list] = {}
        _client_norms:  dict[str, float] = {}

        for update in updates:
            if not update or not update.node_id:
                continue
            weights = update.weights or []

            # delta = raw_local_weights - raw_global_prev_round.
            if prev_raw_global is not None and len(prev_raw_global) == len(weights):
                delta_weights = [
                    (w.float() if isinstance(w, torch.Tensor) else torch.tensor(float(w)))
                    - (g.float() if isinstance(g, torch.Tensor) else torch.tensor(float(g)))
                    for w, g in zip(weights, prev_raw_global)
                ]
            else:
                delta_weights = [
                    (w.float() if isinstance(w, torch.Tensor) else torch.tensor(float(w)))
                    for w in weights
                ]

            # L2 norm del delta (somma di norme al quadrato di tutti i layer)
            l2_sq = sum(
                float(dw.float().norm() ** 2) if isinstance(dw, torch.Tensor)
                else float(dw) ** 2
                for dw in delta_weights
            )
            _client_deltas[update.node_id] = delta_weights
            _client_norms[update.node_id]  = float(np.sqrt(max(l2_sq, 1e-12)))

        # ── Normalizzazione peer-relative ─────────────────────────────────────
        # Fix 1: porta la norma mediana = max_grad_norm.
        # Risultato: nodi normali → norma ≈ max_grad_norm (no explosion);
        #            nodi Byzantine (×10) → norma >> max_grad_norm (explosion rilevata).
        # Perché mediana (non media): un nodo Byzantine estremo non sposta la mediana,
        # mentre sposta la media rendendo il riferimento instabile.
        if _client_norms:
            _sorted_norms = sorted(_client_norms.values())
            # Lower-middle per N pari (es. 4 client): evita che un Byzantine outlier
            # sposti la mediana upward, riducendo la sensibilità al rilevamento.
            _median_norm  = _sorted_norms[(len(_sorted_norms) - 1) // 2]
            # Guard: se la mediana è degenere (< 1e-4) tutti i client hanno delta ≈ 0
            # → nessun training significativo → skip normalizzazione (scale=1.0).
            # Senza questo guard, _scale = max_grad_norm/1e-8 = 1e8, che causa
            # GRADIENT_EXPLOSION falso su qualsiasi client con norma non-nulla.
            _scale = max_grad_norm / _median_norm if _median_norm >= 1e-4 else 1.0
        else:
            _scale = 1.0

        # ── Pass 2: audit con delta normalizzati + Krum analysis ─────────────
        reports:   dict[str, Any] = {}
        gradients: dict[str, dict[str, Any]] = {}

        for node_id, delta_weights in _client_deltas.items():
            # Normalizzazione peer-relative: scala in modo che la mediana = max_grad_norm
            model_update: dict[str, Any] = {
                f"layer_{i}": (
                    dw * _scale if isinstance(dw, torch.Tensor)
                    else torch.tensor(float(dw) * _scale)
                )
                for i, dw in enumerate(delta_weights)
            }

            # AuditReport: GRADIENT_EXPLOSION ora usa delta normalizzati → no FP sistematici
            reports[node_id] = auditor.audit(
                node_id=node_id,
                round_id=round_num,
                model_update=model_update,
            )

            # Gradient dict per Krum / cosine analysis (usa delta NON normalizzati:
            # Krum è già una misura geometrica relativa, non assoluta)
            gradients[node_id] = {
                f"layer_{i}": dw for i, dw in enumerate(delta_weights)
            }

        prev_raw_global = current_raw_global

        if not reports:
            ids_results[round_num] = {
                "alerts": [], "byzantine_detected": False, "drift_detected": False,
            }
            continue

        analysis = ids.analyze_round(
            round_id=round_num,
            reports=reports,
            gradients=gradients,
        )

        ids_results[round_num] = {
            "alerts": [
                {
                    "node_id":            a.node_id,
                    "severity":           a.severity,
                    "reasons":            a.reasons,
                    "recommended_action": a.recommended_action,
                }
                for a in (analysis.alerts if analysis else [])
            ],
            "byzantine_detected":   len(analysis.byzantine_nodes) > 0 if analysis else False,
            "drift_detected":       False,
            "low_similarity_nodes": analysis.low_similarity_nodes if analysis else [],
        }

    return ids_results


# ── Save Results ───────────────────────────────────────────────────────────────

def save_results(
    cfg: dict,
    mia_results: dict[int, dict[str, Any]],
    ids_results: dict[int, dict[str, Any]],
    fl_results: dict[int, dict[str, Any]] | None = None,
    sweep_dir: Path | None = None,
) -> Path:
    """
    Salva risultati in experiments/ (o sweep_dir) con timestamp.

    Se sweep_dir è fornita, i JSON vengono salvati in quella directory
    e l'Excel verrà nominato come la directory (es.
    experiments/nodp-sweep1/nodp-sweep1.xlsx).
    Questo garantisce che ogni sweep abbia il proprio file Excel separato.
    """
    if sweep_dir is not None:
        output_dir = sweep_dir
    else:
        output_dir = PROJECT_ROOT / cfg["output"]["experiments_dir"]
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_file = output_dir / f"experiment_{timestamp}.json"

    auc_values = [
        r["auc_roc"]
        for r in mia_results.values()
        if r.get("auc_roc") is not None
    ]
    shadow_auc_values = [
        r["shadow_auc_roc"]
        for r in mia_results.values()
        if r.get("shadow_auc_roc") is not None
    ]
    lira_auc_values = [
        r["lira_auc_roc"]
        for r in mia_results.values()
        if r.get("lira_auc_roc") is not None
    ]

    # Privacy risk basato sull'attacco più forte disponibile:
    # priority: LiRA (intercepts raw local models, strongest) > Shadow > Yeom.
    _primary_auc  = lira_auc_values if lira_auc_values else (
                    shadow_auc_values if shadow_auc_values else auc_values)
    _primary_mean = float(np.mean(_primary_auc)) if _primary_auc else None
    _primary_min  = float(np.min(_primary_auc))  if _primary_auc else None

    # Fix 2026-07-21 (review Excel/privacy_risk): un AUC-ROC molto SOTTO 0.5 non
    # è "privacy sicura" — è quasi sempre il sintomo di un attacco rotto/invertito
    # (visto in experiments/exp3 pre-fix: lira_auc_roc fino a 0.14, classificato
    # "LOW risk" dalla logica precedente). Controlliamo il MINIMO per round, non
    # solo la media: un'inversione isolata in pochi round può restare nascosta
    # nella media ma è comunque un segnale di bug da investigare, non da ignorare.
    _ANOMALY_LOW_AUC = 0.40
    _is_anomalous = _primary_min is not None and _primary_min < _ANOMALY_LOW_AUC

    summary = {
        "experiment_name": cfg["experiment"]["name"],
        "timestamp":       timestamp,
        "config": {
            "epsilon":    cfg["experiment"]["epsilon"],
            "delta":      cfg["experiment"]["delta"],
            "fl_rounds":  cfg["experiment"]["fl_rounds"],
            "proximal_mu": cfg["ml"]["proximal_mu"],
            # no_dp=True → baseline senza rumore DP; usato per disambiguare AUC≈0.5
            "no_dp":      cfg["experiment"].get("no_dp", False),
            # dp_mode (2026-07-22): quale placement DP — "dp-fedavg" (default),
            # "central" o "local". Vedi docs/CaseStudies.md §2.4.3 per la
            # tassonomia. Irrilevante quando no_dp=True.
            "dp_mode":    cfg["experiment"].get("dp_mode", "dp-fedavg"),
            # seed: necessario per multi-seed aggregation (mean±std) — fix M1
            "seed":       cfg["experiment"].get("seed", 42),
            # epsilon_cumulative_naive (2026-07-22, review indipendente pre-push):
            # PRIMA di questo fix, solo l'epsilon NOMINALE per-round veniva
            # esportato — mai il budget cumulativo reale su tutti i round.
            # gradient_manager.py::_compute_sigma() documenta che (a) la
            # garanzia (ε,δ)-DP formale vale solo per epochs=1 (qui epochs=50),
            # e (b) sotto composizione naive (nessuna amplificazione Rényi/
            # sub-sampling) il budget cumulativo reale su T round è
            # ε_tot ≈ T × ε_per_round — MOLTO più permissivo del solo ε
            # nominale riportato finora. Vedi docs/CaseStudies.md §2.4.3 per
            # la tabella e la spiegazione completa. Questo numero è la
            # spiegazione più rigorosa del perché LiRA continua a rilevare
            # membership anche con "DP attiva" a ε nominale basso: riportarlo
            # esplicitamente nei risultati (non solo in un commento nel
            # codice) rende il claim "c'è fuga di privacy reale nonostante la
            # DP" verificabile numero alla mano, non solo qualitativo.
            # None quando no_dp=True (nessun budget DP consumato).
            "epsilon_cumulative_naive": (
                None if cfg["experiment"].get("no_dp", False)
                else cfg["experiment"]["epsilon"] * cfg["experiment"]["fl_rounds"]
            ),
        },
        "summary": {
            # Yeom 2018 — loss-based MIA sul modello globale (baseline debole)
            "mean_auc_roc": float(np.mean(auc_values)) if auc_values else None,
            "max_auc_roc":  float(np.max(auc_values))  if auc_values else None,
            "min_auc_roc":  float(np.min(auc_values))  if auc_values else None,
            # Shadow calibrated — Carlini 2022 stile, modello globale
            "mean_shadow_auc_roc": float(np.mean(shadow_auc_values)) if shadow_auc_values else None,
            "max_shadow_auc_roc":  float(np.max(shadow_auc_values))  if shadow_auc_values else None,
            "min_shadow_auc_roc":  float(np.min(shadow_auc_values))  if shadow_auc_values else None,
            # LiRA — Carlini 2022, server-side sul singolo update di ogni client
            # PRE-aggregazione FedAvg, POST-privatizzazione DP (fix 2026-07-21c) — attacco primario
            "mean_lira_auc_roc": float(np.mean(lira_auc_values)) if lira_auc_values else None,
            "max_lira_auc_roc":  float(np.max(lira_auc_values))  if lira_auc_values else None,
            "min_lira_auc_roc":  float(np.min(lira_auc_values))  if lira_auc_values else None,
            # Privacy risk: usa l'attacco più forte (LiRA > Shadow > Yeom)
            "primary_attack": (
                "LiRA"   if lira_auc_values else
                "Shadow" if shadow_auc_values else
                "Yeom"
            ),
            # ANOMALY ha priorità su tutto il resto: un min_*_auc_roc < 0.40 indica
            # quasi certamente un bug nell'attacco (score invertito), non un dato
            # di privacy risk affidabile — va investigato, non riportato come LOW.
            "privacy_risk": (
                "ANOMALY" if _is_anomalous else
                "HIGH"    if _primary_mean is not None and _primary_mean > 0.7 else
                "MEDIUM"  if _primary_mean is not None and _primary_mean > 0.6 else
                "LOW"
            ),
        },
        "per_round": {
            # Itera sull'unione di tutti i round: FL, MIA e IDS.
            # round 0 è escluso: contiene solo raw_global_weights (init model) per IDS.
            str(r): {
                "fl":  {"mean_loss": (fl_results or {}).get(r, {}).get("mean_loss")},
                "mia": mia_results.get(r, {}),
                "ids": ids_results.get(r, {}),
            }
            for r in sorted(
                set(mia_results.keys())
                | {k for k in (fl_results or {}).keys() if k > 0}
                | set(ids_results.keys())
            )
        },
    }

    with open(result_file, "w") as f:
        json.dump(summary, f, indent=2, default=str)

    # Fix 2026-07-22 (trovato analizzando i risultati nodp/dp reali, non da una
    # review del codice): questo log stampava mean_auc_roc (la media di Yeom,
    # l'attacco PIÙ DEBOLE — quasi sempre ≈0.50, a prescindere da ε) accanto al
    # verdetto privacy_risk, che invece è calcolato da _primary_mean (LiRA
    # quando disponibile — vedi "Privacy risk basato sull'attacco più forte"
    # sopra). Il log dava l'impressione fuorviante che il numero mostrato
    # spiegasse il verdetto, mentre erano due metriche diverse (visto dal vivo:
    # "AUC-ROC medio: 0.4988 — Privacy risk: ANOMALY", con LiRA mean reale
    # ≈0.49 ma min 0.39 — il numero giusto per capire l'ANOMALY non era quello
    # stampato). Ora mostra esplicitamente l'attacco primario usato per il
    # verdetto, con Yeom a fianco per contesto.
    mean_str = f"{_primary_mean:.4f}" if _primary_mean is not None else "N/A"
    yeom_mean_str = f"{summary['summary']['mean_auc_roc']:.4f}" \
                    if summary["summary"]["mean_auc_roc"] is not None else "N/A"
    logger.info(f"Risultati salvati: {result_file.name}")
    logger.info(
        f"{summary['summary']['primary_attack']} AUC-ROC medio: {mean_str} "
        f"(Yeom: {yeom_mean_str}) — "
        f"Privacy risk: {summary['summary']['privacy_risk']}"
    )
    if summary["summary"]["privacy_risk"] == "ANOMALY":
        logger.warning(
            f"[ANOMALY] {summary['summary']['primary_attack']} ha un min AUC-ROC "
            f"< {_ANOMALY_LOW_AUC} in almeno un round — probabile bug nell'attacco "
            "(score sistematicamente invertito), NON privacy sicura. "
            "Controllare per_round prima di usare questi risultati nel paper."
        )

    # Aggiorna automaticamente il report Excel del sweep corrente
    _update_excel_report(output_dir, named_sweep=(sweep_dir is not None))

    return result_file


def _update_excel_report(sweep_dir: Path, named_sweep: bool = False) -> None:
    """
    Rigenera il report Excel a 11 sheet per il sweep corrente.

    Il nome del file Excel dipende dalla modalità:
    - sweep_dir nominata (es. experiments/nodp-sweep1) → nodp-sweep1.xlsx
    - fallback (experiments/) → ChargeShield_FL_Results.xlsx (retro-compatibilità,
      usata da `make experiment-nodp`/`experiment-dp` — run singolo-seed senza
      --sweep-dir, non parte della metodologia multi-seed)

    Usa un import Python standard invece di exec_module() per evitare il rischio
    di arbitrary code execution se il file fosse modificato da un attacker con accesso
    al filesystem. Con import standard il modulo viene caricato una sola volta e
    cachato in sys.modules — sicuro e idempotente.

    NON SOVRASCRITTURA (fix 2026-07-24, richiesta esplicita dell'utente: "i file
    excel non si devono sovrascrivere... un foglio o un file per ogni esperimento
    lanciato"): questa funzione viene chiamata una volta per ogni singolo run/seed
    completato all'interno di uno sweep. Il file "{sweep}.xlsx" continua a essere
    sovrascritto a ogni chiamata — di proposito, è la vista aggregata sempre
    aggiornata su tutti i seed raccolti finora (lo stesso principio già accettato
    per il foglio "Seed Aggregation" al suo interno). Oltre a quello, ora viene
    salvato anche uno snapshot permanente in <sweep_dir>/history/, MAI
    sovrascritto — uno per ogni esperimento lanciato, esattamente come richiesto.
    Vedi generate_excel_report.py::save_report_with_history().

    Args:
        sweep_dir:   directory dove sono i JSON del sweep (e dove salvare l'Excel)
        named_sweep: True se sweep_dir è una directory nominata (es. exp1),
                     False per backward compatibility con experiments/
    """
    try:
        from openpyxl import Workbook

        # Import diretto: generate_excel_report.py è nella stessa directory di questo script.
        # Se il file non esiste, ImportError viene catturato sotto.
        _scripts_dir = str(Path(__file__).parent)
        if _scripts_dir not in sys.path:
            sys.path.insert(0, _scripts_dir)
        import generate_excel_report as gen  # noqa: PLC0415

        records = gen.load_experiments(sweep_dir)
        if not records:
            return

        wb = Workbook()
        wb.remove(wb.active)
        gen.build_raw_data(wb.create_sheet("Raw Data"),                   records)
        gen.build_heat_map(wb.create_sheet("Heat Map"),                   records)
        gen.build_per_rounds(wb.create_sheet("Per Rounds"),               records)
        gen.build_per_epsilon(wb.create_sheet("Per Epsilon"),             records)
        gen.build_comparison(wb.create_sheet("Comparison"),               records)
        gen.build_auc_progression(wb.create_sheet("AUC Progression"),     records)
        gen.build_attack_comparison(wb.create_sheet("Attack Comparison"), records)
        gen.build_yeom_per_round(wb.create_sheet("Yeom Per Round"),         records)
        gen.build_shadow_per_round(wb.create_sheet("Shadow Per Round"),   records)
        gen.build_lira_per_round(wb.create_sheet("LiRA Per Round"),       records)
        gen.build_seed_aggregation(wb.create_sheet("Seed Aggregation"),   records)
        wb.properties.title   = "ChargeShield-FL Experiment Results"
        wb.properties.subject = "FedMIA vs Differential Privacy — DSN 2027"

        # Nome file Excel: sweep nominato → "{nome}.xlsx", fallback → nome storico
        if named_sweep:
            output_path = sweep_dir / f"{sweep_dir.name}.xlsx"
        else:
            output_path = sweep_dir / "ChargeShield_FL_Results.xlsx"

        snapshot_path = gen.save_report_with_history(wb, output_path)
        logger.info(
            f"Report Excel aggiornato: {output_path.name} "
            f"(snapshot permanente: {snapshot_path.relative_to(sweep_dir)})"
        )
    except ImportError:
        logger.warning(
            "openpyxl non trovato — report Excel non generato. "
            "Installa con: pip install openpyxl"
        )
    except Exception as exc:
        logger.warning(f"Report Excel non generato: {exc}")


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="ChargeShield-FL — FedMIA Experiment Runner"
    )
    parser.add_argument("--config",   type=Path, default=Path("config/experiment.yaml"))
    parser.add_argument("--epsilon",  type=float, default=None)
    parser.add_argument("--rounds",   type=int,   default=None)
    parser.add_argument("--skip-ids", action="store_true")
    parser.add_argument("--dry-run",  action="store_true")
    parser.add_argument(
        "--sweep-dir", type=Path, default=None,
        help=(
            "Directory del sweep corrente (es. experiments/nodp-sweep1). "
            "Se fornita, JSON e Excel vengono salvati qui con nome del sweep "
            "(es. nodp-sweep1.xlsx). Permette di isolare i risultati di sweep distinti."
        ),
    )
    parser.add_argument(
        "--seed", type=int, default=None,
        help=(
            "Seed riproducibilità (override experiment.yaml). "
            "Usato per: shuffle sessioni, DataLoader, init modello. "
            "Valori consigliati per sweep: 42 123 456 789 1234."
        ),
    )
    parser.add_argument(
        "--byzantine", action="store_true", default=False,
        help="Abilita Byzantine attack (gradient scaling) sul nodo configurato in experiment.yaml.",
    )
    parser.add_argument(
        "--byzantine-node", type=str, default=None,
        help=(
            "Override cluster attaccante (2026-07-22: synthetic_1/synthetic_2 "
            "per lo sweep IDS a 5 client — mai un sito reale caltech/jpl/office1). "
            "Default: valore da config."
        ),
    )
    parser.add_argument(
        "--scale-factor", type=float, default=None,
        help="Override scale_factor dell'attacco. Default: valore da config (10.0).",
    )
    parser.add_argument(
        "--no-dp", action="store_true", default=False,
        help=(
            "Disabilita rumore Differential Privacy (σ=0). "
            "BASELINE CRITICO per DSN 2027: distingue "
            "Scenario A (DP sopprime MIA → AUC>0.5 senza DP) da "
            "Scenario B (modello non memorizza → AUC≈0.5 anche senza DP). "
            "Usare sempre prima del full sweep per capire in quale scenario siamo."
        ),
    )
    parser.add_argument(
        "--dp-mode", type=str, default="dp-fedavg",
        choices=["dp-fedavg", "central", "local"],
        help=(
            "Placement del meccanismo DP quando --no-dp non è passato "
            "(2026-07-22, vedi docs/CaseStudies.md §2.4.3): "
            "'dp-fedavg' (default, storico) = server clippa+rumorizza ogni "
            "update PRIMA di aggregare — placement per-client non-standard, "
            "NON quello descritto dall'Algoritmo 1 di McMahan 2018; server/IDS "
            "vede l'update raw transitoriamente. "
            "'central' [McMahan 2018, Algoritmo 1 (DP-FedAvg): clip lato "
            "client, un solo draw di rumore lato server sull'aggregato] = "
            "client clippano SENZA rumorizzare, il server aggrega "
            "pulito e aggiunge UN SOLO rumore all'aggregato — atteso: LiRA sul "
            "singolo update non mostra alcuna soppressione, a nessun ε. "
            "'local' = stesso meccanismo per-client di dp-fedavg, ma server/IDS "
            "non vede MAI l'update raw (nemmeno transitoriamente) — l'IDS perde "
            "l'accesso a raw_updates, un tradeoff reale della local DP, non un bug."
        ),
    )
    parser.add_argument(
        "--n-shadow", type=int, default=None,
        help=(
            "Numero di shadow models per LiRA (override config lira.n_shadow). "
            "8 = fast demo (~10 min CPU); 16 = buona qualità; ≥32 = paper quality. "
            "Più shadow = IN/OUT distributions più stabili → AUC più affidabile."
        ),
    )
    parser.add_argument(
        "--shadow-epochs-cap", type=int, default=None,
        help=(
            "Cap massimo per shadow_epochs in LiRA (override formula automatica). "
            "Formula default: min(local_epochs × max(rounds//4, 5), 300). "
            "Usare un valore basso (es. 20) nel smoke test per ridurre il tempo: "
            "con epochs=50 e rounds=5, la formula darebbe 250 — troppo per un test rapido. "
            "Non usare nei run sperimentali reali: i shadow models sarebbero sottoadatti."
        ),
    )
    return parser.parse_args()


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    args = parse_args()

    logger.info("=" * 60)
    logger.info("ChargeShield-FL — Sprint 5 Experiment")
    logger.info("=" * 60)

    cfg = load_config(args.config, {"epsilon": args.epsilon, "rounds": args.rounds})

    # Override seed da CLI (--seed)
    if args.seed is not None:
        cfg["experiment"]["seed"] = args.seed

    # Override Byzantine attack da CLI (--byzantine, --byzantine-node, --scale-factor)
    if args.byzantine:
        cfg.setdefault("byzantine_attack", {})["enabled"] = True
    if args.byzantine_node:
        cfg.setdefault("byzantine_attack", {})["byzantine_node"] = args.byzantine_node
    if args.scale_factor is not None:
        cfg.setdefault("byzantine_attack", {})["scale_factor"] = args.scale_factor

    # Guard: --byzantine e --no-dp sono esperimenti concettualmente separati.
    # --no-dp = baseline MIA pulito (nessun attacco, nessun rumore DP).
    # --byzantine = validazione IDS (attacco attivo, non misura di privacy risk).
    # Combinarli produce un risultato privo di significato per entrambi gli obiettivi.
    if args.byzantine and args.no_dp:
        logger.error(
            "Combinazione non valida: --byzantine e --no-dp non possono essere usati insieme.\n"
            "  --no-dp   = baseline MIA pulito (solo per misurare privacy risk senza DP)\n"
            "  --byzantine = validazione IDS (solo per testare rilevamento attacchi)\n"
            "Eseguirli come esperimenti separati:\n"
            "  Privacy baseline: make experiment-nodp\n"
            "  IDS validation:   make experiment-byzantine-sweep"
        )
        sys.exit(1)

    # No-DP baseline flag: disabilita rumore DP, rinomina esperimento per distinzione
    if args.no_dp:
        cfg["experiment"]["no_dp"] = True
        cfg["experiment"]["name"] = cfg["experiment"]["name"] + "_nodp_baseline"
        if args.dp_mode != "dp-fedavg":
            logger.warning(
                f"--dp-mode {args.dp_mode} ignorato: --no-dp disabilita ogni rumore, "
                "indipendentemente dal placement DP scelto."
            )
        # Fix 2026-07-22 (review A3): normalizza args.dp_mode PRIMA che venga
        # persistito qui sotto in cfg["experiment"]["dp_mode"] — senza questo,
        # una riga con no_dp=True poteva comunque riportare dp_mode="central"/
        # "local" nel JSON/Excel, incoerente col fatto che no_dp disabilita
        # ogni rumore (il valore era innocuo a runtime — no_dp cortocircuita
        # run_fl_rounds() prima che dp_mode sia consultato — ma fuorviante nei
        # risultati salvati).
        args.dp_mode = "dp-fedavg"

    # dp_mode (2026-07-22): rinomina esperimento per distinguere le 3 modalità DP
    # nel nome/nei log — importante perché tutte e tre producono un modello con
    # rumore (auc<0.6 plausibile per tutte), ma per ragioni architetturali diverse.
    cfg["experiment"]["dp_mode"] = args.dp_mode
    if not args.no_dp and args.dp_mode != "dp-fedavg":
        cfg["experiment"]["name"] = cfg["experiment"]["name"] + f"_{args.dp_mode}_dp"

    # Warning esplicito se Byzantine è attivo senza --sweep-dir: rischio di mischiare
    # risultati IDS con risultati MIA nella directory experiments/ principale.
    if args.byzantine and not args.sweep_dir:
        logger.warning(
            "[BYZANTINE] Attacco attivo senza --sweep-dir esplicita. "
            "I risultati Byzantine andrebbero in experiments/ids_validation/ "
            "(usa 'make experiment-byzantine-sweep' per garantire la separazione). "
            "I risultati MIA in questa run NON sono validi per il privacy sweep."
        )

    exp_cfg = cfg["experiment"]
    logger.info(
        f"Config: epsilon={exp_cfg['epsilon']}, "
        f"rounds={exp_cfg['fl_rounds']}, "
        f"proximal_mu={cfg['ml']['proximal_mu']}"
    )

    sessions = load_sessions(cfg)
    sessions = enrich_sessions(sessions)
    logger.info(f"Sessioni dopo enrichment: {len(sessions)}")

    # Split hold-out PRIMA del training FL: 80% train, 20% hold-out (mai visti dai nodi FL).
    # Seed fisso per riproducibilità dei risultati — fondamentale per DSN 2027.
    seed = exp_cfg.get("seed", 42)
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    # Determinismo GPU (no overhead su CPU-only, ignorato silenziosamente se no CUDA)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    random.shuffle(sessions)
    split = max(1, int(len(sessions) * 0.8))
    train_sessions   = sessions[:split]
    holdout_sessions = sessions[split:]
    logger.info(f"Split — train: {len(train_sessions)}, hold-out: {len(holdout_sessions)}")

    # Normalizzazione min-max: calcolata SOLO su train_sessions (no leakage dal hold-out).
    # Stessa trasformazione applicata a holdout_sessions per la FedMIA.
    _FEATURES = AutoencoderTrainer.CONTINUOUS_FEATURES  # importato a livello modulo (riga 41)
    feature_stats    = compute_feature_stats(train_sessions, _FEATURES)
    train_sessions   = normalize_sessions(train_sessions,   feature_stats, _FEATURES)
    holdout_sessions = normalize_sessions(holdout_sessions, feature_stats, _FEATURES)
    logger.info(f"Feature normalizzate [0,1]: {list(feature_stats.keys())}")
    if args.dry_run:
        logger.info("Dry run completato — uscita.")
        return

    # ── Cluster reali (2026-07-22) — sostituisce lo slicing contiguo in 4 fette
    # fittizie di un unico dataset (mai stato realmente eterogeneo — vedi fix
    # mislabeling jpl/Caltech in README/CaseStudies.md). Ogni sessione porta già
    # il proprio site_id reale (ACNDataset) — group_indices_by_site() raggruppa
    # per quello, non per posizione. Vedi group_indices_by_site()/
    # inject_synthetic_client_indices() per il contratto completo.
    #
    # IMPORTANTE: i 2 client sintetici (synthetic_1/synthetic_2) esistono
    # SOLO per rendere valido il rilevamento Byzantine di Krum (che richiede
    # n≥2f+3 nodi — con f=1 servono 5, i 3 siti reali non bastano). Vengono
    # aggiunti ESCLUSIVAMENTE quando byzantine_attack.enabled=True (sweep IDS
    # dedicato, separato dallo sweep privacy). L'esperimento principale
    # (FedMIA/Shadow/LiRA, privacy_risk, tutto ciò che finisce nel paper come
    # misura di privacy) vede SEMPRE e SOLO i 3 client reali (caltech/jpl/
    # office1) — mai i sintetici, che userebbero sessioni duplicate/sovrapposte
    # tra client e invaliderebbero qualunque misura di privacy o utility.
    real_cluster_membership = group_indices_by_site(train_sessions)
    _byz_enabled_main = cfg.get("byzantine_attack", {}).get("enabled", False)
    if _byz_enabled_main:
        cluster_membership = inject_synthetic_client_indices(
            real_cluster_membership, n_synthetic=2, seed=seed,
        )
        logger.warning(
            "[BYZANTINE/IDS SWEEP] 2 client sintetici aggiunti (synthetic_1/"
            "synthetic_2) per validare Krum con n=5 — NON usare questo run per "
            "misure di privacy/FedMIA/LiRA, solo per validazione IDS."
        )
    else:
        cluster_membership = real_cluster_membership
    cluster_sessions = {
        cid: [train_sessions[i] for i in idxs] for cid, idxs in cluster_membership.items()
    }
    logger.info(f"Client attivi ({len(cluster_sessions)}): {list(cluster_sessions.keys())}")

    fl_results = run_fl_rounds(
        cfg, train_sessions, no_dp=args.no_dp, dp_mode=args.dp_mode,
        cluster_sessions=cluster_sessions,
    )

    # ── FedMIA/Shadow/LiRA: SOLO client reali, MAI durante uno sweep Byzantine/IDS ──
    # Istruzione esplicita 2026-07-22: i 2 client sintetici (synthetic_1/
    # synthetic_2) servono ESCLUSIVAMENTE a rendere valido il rilevamento Krum
    # (n≥2f+3). Il modello FL viene addestrato su di essi quando
    # byzantine_attack.enabled=True (sopra, cluster_sessions), ma qualunque
    # attacco di privacy calcolato su quel run includerebbe update di client le
    # cui sessioni si sovrappongono arbitrariamente tra loro e coi client reali
    # (vedi inject_synthetic_client_indices()) — un numero privo di significato
    # per qualunque claim di privacy, non solo "da usare con cautela". Per
    # questo FedMIA/Shadow/LiRA vengono saltati DEL TUTTO quando
    # byzantine_attack.enabled=True, invece di girare comunque su dati che poi
    # nessuno dovrebbe interpretare. Il run Byzantine resta quindi
    # esclusivamente uno strumento di validazione IDS (Krum/cosine/alert),
    # mai una fonte di numeri di privacy — coerente con quanto già indicato
    # dal warning "I risultati MIA in questa run NON sono validi per il privacy
    # sweep" più sopra in questa stessa funzione, reso qui vincolante invece
    # che solo informativo.
    mia_results: dict = {}
    if _byz_enabled_main:
        logger.warning(
            "[BYZANTINE/IDS SWEEP] FedMIA/Shadow/LiRA SALTATI — questo run usa "
            "client sintetici (synthetic_1/synthetic_2) solo per validare Krum, "
            "non è un esperimento di privacy valido. Per i numeri di privacy "
            "usare un run con byzantine_attack.enabled=false (solo 3 client reali)."
        )
    else:
        # Fix 2026-07-24: i tre attacchi (Yeom, Shadow, LiRA) non vengono più
        # chiamati per nome qui — main() itera ATTACK_REGISTRY tramite
        # run_registered_attacks() (src/plugins/attacks/, vedi la sua docstring
        # per il razionale completo e la garanzia di comportamento identico
        # alla versione pre-refactor). run_ids() resta chiamato direttamente
        # sotto: non è un "attacco" nel senso di questa interfaccia (produce
        # audit/detection, non un punteggio di membership), quindi resta fuori
        # dal registro — coerente con come i due concetti erano già separati
        # prima di questo fix.
        n_shadow = args.n_shadow if args.n_shadow is not None else cfg.get("lira", {}).get("n_shadow", 8)
        shadow_cap = args.shadow_epochs_cap  # None → local_epochs; int → override per smoke
        mia_results = run_registered_attacks(
            cfg, train_sessions, holdout_sessions, fl_results,
            n_shadow=n_shadow, shadow_epochs_cap=shadow_cap, no_dp=args.no_dp,
            dp_mode=args.dp_mode, cluster_membership=cluster_membership,
        )

    ids_results: dict = {}
    if not args.skip_ids:
        try:
            ids_results = run_ids(cfg, fl_results, no_dp=args.no_dp)
        except Exception as exc:  # noqa: BLE001
            logger.error(f"run_ids() fallita: {exc}. Continuazione senza risultati IDS.", exc_info=True)

    # sweep_dir: se fornita via --sweep-dir, i risultati vanno in quella directory
    # con Excel nominato come il sweep (es. exp1.xlsx). Altrimenti usa experiments/.
    sweep_dir = args.sweep_dir.resolve() if args.sweep_dir else None
    save_results(cfg, mia_results, ids_results, fl_results, sweep_dir=sweep_dir)

    logger.info("=" * 60)
    logger.info("Esperimento completato.")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
