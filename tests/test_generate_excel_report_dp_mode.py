# tests/test_generate_excel_report_dp_mode.py
"""
Regression tests for scripts/generate_excel_report.py's handling of `dp_mode`
(central/local DP added 2026-07-22) — specifically finding A2 from the
2026-07-22 independent review: `build_seed_aggregation()` and `build_heat_map()`
grouped/indexed records by (rounds, epsilon[, no_dp]) WITHOUT `dp_mode`, so
seeds/records from different DP mechanisms (dp-fedavg/central/local) at the
same (rounds, epsilon) would silently get pooled into one mean±std row, or
silently overwrite each other in the heat map — with no error or warning.

Unlike most of today's new code, this module has NO torch dependency (pure
Python + openpyxl), so — unlike the DP/NVFLARE code elsewhere in this session
— these tests were actually run in this sandbox (see run log at the bottom of
this docstring's companion commit message), not just syntax-checked.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

import generate_excel_report as ger  # noqa: E402


def _rec(rounds, epsilon, dp_mode="dp-fedavg", no_dp=False, seed=42, auc_roc=0.5,
         shadow_auc=0.5, lira_auc=0.5, lira_min=0.4, lira_max=0.6):
    """Minimal synthetic record matching the shape load_experiments() produces."""
    return {
        "rounds": rounds,
        "epsilon": epsilon,
        "dp_mode": dp_mode,
        "no_dp": no_dp,
        "seed": seed,
        "auc_roc": auc_roc,
        "shadow_auc": shadow_auc,
        "lira_auc": lira_auc,
        "lira_min": lira_min,
        "lira_max": lira_max,
    }


class TestSeedAggregationDpModeGrouping:
    def test_different_dp_modes_at_same_eps_are_not_pooled(self):
        """
        Core regression for A2: 5 seeds of dp-fedavg and 5 seeds of central DP,
        same (rounds=10, epsilon=1.0) — must land in TWO separate rows, each
        with N Seed == 5, not one row with N Seed == 10 (which would silently
        average two different DP mechanisms together).
        """
        records = [
            _rec(10, 1.0, dp_mode="dp-fedavg", seed=s, auc_roc=0.50)
            for s in range(5)
        ] + [
            _rec(10, 1.0, dp_mode="central", seed=s, auc_roc=0.58)
            for s in range(5, 10)
        ]

        wb = Workbook()
        ws = wb.active
        ger.build_seed_aggregation(ws, records)

        # Header è in riga 3 (righe 1-2 = titolo/sottotitolo); dati da riga 4.
        n_seed_col = 3
        # Filtra solo le vere righe dati (colonna "N Seed" numerica) — sotto le
        # righe dati il foglio prosegue con una riga di legenda testuale
        # ("Legenda N Seed: ..."), che ha colonna A non-None ma colonna C None:
        # senza questo filtro verrebbe scambiata per una riga dati in più.
        data_rows = [
            row for row in ws.iter_rows(min_row=4, max_col=n_seed_col, values_only=False)
            if isinstance(row[n_seed_col - 1].value, int)
        ]
        n_seed_values = sorted(int(row[n_seed_col - 1].value) for row in data_rows)

        assert len(data_rows) == 2, (
            f"attese 2 righe (una per dp_mode), trovate {len(data_rows)} — "
            "se è 1, i due dp_mode sono stati pooled insieme (regressione A2); "
            "se è >2 qualcos'altro è cambiato nella logica di raggruppamento"
        )
        assert n_seed_values == [5, 5], (
            f"atteso N Seed=5 per ciascuna delle 2 righe (dp-fedavg e central "
            f"separati), trovato {n_seed_values} — un [10] significherebbe che "
            "i due dp_mode sono stati mediati insieme in una sola riga"
        )

    def test_missing_dp_mode_defaults_to_dp_fedavg_for_backward_compat(self):
        """Record vecchi (pre-2026-07-22) non hanno la chiave 'dp_mode' — deve
        essere trattata come 'dp-fedavg' (default), non far esplodere il
        raggruppamento con una chiave mancante/None inconsistente."""
        records = [
            {k: v for k, v in _rec(5, 2.0, seed=s).items() if k != "dp_mode"}
            for s in range(3)
        ]
        wb = Workbook()
        ws = wb.active
        ger.build_seed_aggregation(ws, records)  # non deve sollevare eccezioni
        data_rows = [
            row for row in ws.iter_rows(min_row=4, max_col=3, values_only=False)
            if isinstance(row[2].value, int)
        ]
        assert len(data_rows) == 1
        assert int(data_rows[0][2].value) == 3


class TestHeatMapDpModeFilter:
    def test_heat_map_excludes_non_dp_fedavg_records(self):
        """
        A2 per la Heat Map: la griglia è indicizzata solo su (rounds, epsilon),
        quindi un record 'central' con lo stesso (rounds, epsilon) di un record
        'dp-fedavg' lo sovrascriverebbe silenziosamente (l'ultimo per timestamp
        vince). Fix: filtrare a monte a dp_mode == 'dp-fedavg'. Verifica che un
        valore auc_roc molto diverso su un record 'central' allo stesso
        (rounds, epsilon) NON compaia nella heat map.
        """
        records = [
            _rec(10, 1.0, dp_mode="dp-fedavg", auc_roc=0.50),
            _rec(10, 1.0, dp_mode="central", auc_roc=0.99),  # non deve vincere/mescolarsi
        ]
        wb = Workbook()
        ws = wb.active
        ger.build_heat_map(ws, records)

        # Riga 4 = primo (e unico) valore di rounds; colonna 2 = primo epsilon.
        cell_value = ws.cell(4, 2).value
        assert cell_value == pytest.approx(0.50), (
            f"atteso 0.50 (record dp-fedavg), trovato {cell_value} — il record "
            "'central' con auc_roc=0.99 sembra essere filtrato nella heat map "
            "(regressione A2)"
        )
